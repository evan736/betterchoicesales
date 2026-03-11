"""Non-pay / past-due automation API.

Upload PDF or CSV of non-pay notices → Claude extracts policy numbers →
Match to customers in DB → Send carrier-branded past-due emails.
One email per policy per 7 days max.
"""
import csv
import io
import json
import base64
import logging
import re
from datetime import datetime, timedelta
from typing import Optional

import httpx
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query, Body
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.core.database import get_db
from app.core.config import settings
from app.core.security import get_current_user
from app.models.user import User
from app.models.customer import Customer, CustomerPolicy
from app.models.nonpay import NonPayNotice, NonPayEmail
from app.services.nonpay_email import send_nonpay_email

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/nonpay", tags=["nonpay"])


@router.get("/diag")
def nonpay_diagnostic():
    """Diagnostic endpoint - no auth required."""
    diag = {"status": "ok", "xlrd": False, "openpyxl": False, "tables": False}
    try:
        import xlrd
        diag["xlrd"] = True
        diag["xlrd_version"] = xlrd.__VERSION__
    except ImportError as e:
        diag["xlrd_error"] = str(e)
    try:
        import openpyxl
        diag["openpyxl"] = True
    except ImportError as e:
        diag["openpyxl_error"] = str(e)
    try:
        from app.core.database import get_db, engine
        from sqlalchemy import inspect
        insp = inspect(engine)
        tables = insp.get_table_names()
        diag["tables"] = "nonpay_notices" in tables and "nonpay_emails" in tables
        diag["all_tables"] = [t for t in tables if "nonpay" in t]
    except Exception as e:
        diag["table_error"] = str(e)
    return diag


@router.post("/test-extract")
async def test_extract(payload: dict = Body(...)):
    """Test file extraction without auth or DB. Returns extracted policies."""
    try:
        data_b64 = payload.get("data", "")
        filename = payload.get("filename", "test.csv")
        if not data_b64:
            return {"error": "No data provided"}

        file_bytes = base64.b64decode(data_b64)
        ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

        if ext in ("xlsx", "xls"):
            policies = _extract_from_excel(file_bytes, ext)
        elif ext == "pdf":
            policies = await _extract_from_pdf(file_bytes)
        else:
            policies = _extract_from_csv(file_bytes)

        return {"filename": filename, "ext": ext, "policies_found": len(policies), "policies": policies[:5]}
    except Exception as e:
        import traceback
        return {"error": str(e), "traceback": traceback.format_exc()}

# ── Extraction prompt for Claude ──────────────────────────────────────

# Policy number patterns for carrier auto-detection
_POLICY_NUM_PATTERNS = [
    (re.compile(r'^\d{10,}\s*\d{0,2}$'), "national_general"),  # NatGen: "2027431477 00"
]

NONPAY_EXTRACTION_PROMPT = """You are an expert insurance document parser. This document is a non-pay / past-due / cancellation notice from an insurance carrier or agency management system.

Extract ALL policy numbers and associated details. Return ONLY a valid JSON array:

[
  {
    "policy_number": "The policy number exactly as shown",
    "carrier": "Insurance carrier name if visible",
    "insured_name": "Policyholder name if visible",
    "amount_due": 123.45,
    "due_date": "MM/DD/YYYY or as shown, or null",
    "cancel_reason": "The exact cancellation reason text as shown in the document",
    "notice_type": "non-pay|cancellation|reinstatement|past-due|other"
  }
]

CRITICAL — notice_type classification rules:
- "non-pay" = ONLY for non-payment of premium or NSF (bounced payment). Keywords: "non payment", "non-pay", "NSF", "insufficient funds", "past due"
- "cancellation" = policyholder requested cancellation, underwriting cancellation, carrier non-renewal, or any cancellation NOT related to non-payment. Keywords: "underwriting reasons", "policyholder request", "insured request", "non-renewal", "company cancel"
- "reinstatement" = policy being reinstated
- "past-due" = past due notice that is not yet a pending cancellation
- "other" = anything else

IMPORTANT:
- Extract EVERY policy listed in the document, even if there are hundreds
- Policy numbers may appear in columns, tables, or lists
- Amount may be labeled as "amount due", "balance", "premium due", "past due amount"
- If the document is a CSV/spreadsheet, extract from the appropriate columns
- Return ONLY the JSON array, no markdown, no explanation
- If no policies found, return an empty array: []"""


# ── Upload + Process ─────────────────────────────────────────────────

@router.post("/upload-b64")
async def upload_nonpay_b64(
    payload: dict = Body(...),
    dry_run: bool = Query(False),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Fallback upload via base64 JSON body instead of multipart form."""
    filename = payload.get("filename", "upload.csv")
    data_b64 = payload.get("data", "")
    carrier_override = payload.get("carrier_override", "")  # User-selected carrier fallback
    if not data_b64:
        raise HTTPException(status_code=400, detail="No file data provided")

    file_bytes = base64.b64decode(data_b64)
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    if ext not in ("pdf", "csv", "tsv", "txt", "xlsx", "xls"):
        raise HTTPException(status_code=400, detail="Supported formats: PDF, CSV, XLS, XLSX")

    if len(file_bytes) > 25 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 25MB)")

    try:
        # Create notice record
        notice = NonPayNotice(
            filename=filename,
            upload_type=ext,
            uploaded_by=current_user.full_name or current_user.username,
            status="processing",
        )
        db.add(notice)
        db.commit()
        db.refresh(notice)
    except Exception as e:
        db.rollback()
        import traceback
        raise HTTPException(status_code=500, detail=f"DB error creating notice: {str(e)}\n{traceback.format_exc()[-500:]}")

    try:
        if ext == "pdf":
            policies = await _extract_from_pdf(file_bytes)
        elif ext in ("xlsx", "xls"):
            policies = _extract_from_excel(file_bytes, ext)
        else:
            policies = _extract_from_csv(file_bytes)

        notice.raw_extracted = policies
        notice.policies_found = len(policies)

        # Carrier inference from filename
        FILENAME_CARRIER_MAP = {
            "trv": "travelers", "travelers": "travelers",
            "prog": "progressive", "progressive": "progressive",
            "safeco": "safeco", "geico": "geico", "grange": "grange",
            "hippo": "hippo", "branch": "branch", "next": "next",
            "gainsco": "gainsco", "steadily": "steadily", "obsidian": "steadily",
            "integrity": "integrity", "clearcover": "clearcover",
            "openly": "openly", "bristol": "bristol_west",
            "natgen": "national_general", "national_general": "national_general",
            "universal": "universal_property", "upcic": "universal_property",
            "american_modern": "american_modern", "covertree": "covertree",
        }
        has_any_carrier = any(p.get("carrier") for p in policies)
        if not has_any_carrier and filename:
            fn_lower = filename.lower()
            for pattern, carrier_key in FILENAME_CARRIER_MAP.items():
                if pattern in fn_lower:
                    for p in policies:
                        p["carrier"] = carrier_key
                    has_any_carrier = True
                    break

        # Apply user-selected carrier override if still no carrier detected
        if not has_any_carrier and carrier_override:
            for p in policies:
                if not p.get("carrier"):
                    p["carrier"] = carrier_override

        # Detect carrier from policy number patterns
        # NatGen: 10+ digit numbers, often with " 00" suffix (e.g. "2027431477 00")
        for p in policies:
            if not p.get("carrier"):
                pn = (p.get("policy_number") or "").strip()
                for pat, ckey in _POLICY_NUM_PATTERNS:
                    if pat.match(pn):
                        p["carrier"] = ckey
                        break

        results = []
        matched = 0
        sent = 0
        letters = 0
        skipped = 0

        for pol in policies:
            pnum = (pol.get("policy_number") or "").strip()
            if not pnum:
                continue

            # Filter by cancellation reason — only process non-pay/NSF
            notice_type = pol.get("notice_type", "non-pay")
            cancel_reason = pol.get("cancel_reason", "")

            # Also check reason text for non-payment keywords
            reason_lower = cancel_reason.lower() if cancel_reason else ""
            is_nonpay = any(kw in reason_lower for kw in [
                "non payment", "non-payment", "nonpayment", "non pay", "non-pay",
                "nsf", "insufficient fund", "past due", "past-due",
            ])
            is_skip = any(kw in reason_lower for kw in [
                "underwriting", "policyholder request", "insured request",
                "company cancel", "non-renewal", "nonrenewal",
            ])

            # Skip if explicitly not non-pay, or if reason text indicates skip
            if is_skip or (notice_type not in ("non-pay", "past-due") and not is_nonpay):
                # Skip non-actionable reasons
                results.append({
                    "policy_number": pnum,
                    "insured_name": pol.get("insured_name", ""),
                    "cancel_reason": cancel_reason,
                    "notice_type": notice_type,
                    "skipped_reason": True,
                    "error": f"Skipped — {cancel_reason}" if cancel_reason else f"Skipped — {notice_type}",
                })
                continue

            result = _process_single_policy(
                db=db, notice_id=notice.id, policy_number=pnum,
                carrier=pol.get("carrier", ""), insured_name=pol.get("insured_name", ""),
                amount_due=pol.get("amount_due"), due_date=pol.get("due_date"),
                cancel_date=pol.get("cancel_date"),
                dry_run=dry_run,
            )
            result["cancel_reason"] = cancel_reason
            result["notice_type"] = notice_type
            results.append(result)
            if result.get("matched"): matched += 1
            if result.get("email_sent"): sent += 1
            if result.get("letter_sent"): letters += 1
            if result.get("skipped_rate_limit"): skipped += 1

        notice.policies_matched = matched
        notice.emails_sent = sent + letters
        notice.emails_skipped = skipped
        notice.status = "dry_run" if dry_run else "completed"
        db.commit()

        return {
            "notice_id": notice.id, "filename": filename, "dry_run": dry_run,
            "policies_found": len(policies), "policies_matched": matched,
            "emails_sent": sent, "letters_sent": letters, "emails_skipped": skipped, "details": results,
        }
    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        try:
            notice.status = "error"
            notice.error_message = str(e)[:500]
            db.commit()
        except Exception:
            db.rollback()
        logger.error("Non-pay processing error: %s\n%s", e, tb)
        raise HTTPException(status_code=500, detail=f"Processing failed: {str(e)}\n{tb[-500:]}")


@router.post("/upload")
async def upload_nonpay_file(
    file: UploadFile = File(...),
    dry_run: bool = Query(False),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Upload a non-pay PDF or CSV. Extracts policy info, matches customers, sends emails.
    Set dry_run=true to preview matches without sending any emails."""
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("pdf", "csv", "tsv", "txt", "xlsx", "xls"):
        raise HTTPException(status_code=400, detail="Supported formats: PDF, CSV, XLS, XLSX")

    file_bytes = await file.read()
    if len(file_bytes) > 25 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="File too large (max 25MB)")

    # Create notice record
    try:
        notice = NonPayNotice(
            filename=file.filename,
            upload_type=ext,
            uploaded_by=current_user.full_name or current_user.username,
            status="processing",
        )
        db.add(notice)
        db.commit()
        db.refresh(notice)
    except Exception as e:
        db.rollback()
        import traceback
        raise HTTPException(status_code=500, detail=f"DB error creating notice: {str(e)}\n{traceback.format_exc()[-500:]}")

    try:
        # Extract policies from file
        if ext == "pdf":
            policies = await _extract_from_pdf(file_bytes)
        elif ext in ("xlsx", "xls"):
            # Check for Progressive-specific file formats first
            progressive_result = _detect_progressive_file(file_bytes, ext, file.filename)
            if progressive_result:
                policies = progressive_result["policies"]
                # Process UW items separately — create tasks and optionally send emails
                uw_items = progressive_result.get("uw_items", [])
                if uw_items:
                    _process_progressive_uw_items(db, uw_items, dry_run=dry_run, notice_id=notice.id)
            else:
                policies = _extract_from_excel(file_bytes, ext)
        else:
            policies = _extract_from_csv(file_bytes)

        notice.raw_extracted = policies
        notice.policies_found = len(policies)

        # If no carrier was extracted, try to infer from filename
        FILENAME_CARRIER_MAP = {
            "trv": "travelers", "travelers": "travelers",
            "prog": "progressive", "progressive": "progressive",
            "safeco": "safeco", "geico": "geico", "grange": "grange",
            "hippo": "hippo", "branch": "branch", "next": "next",
            "gainsco": "gainsco", "steadily": "steadily", "obsidian": "steadily",
            "integrity": "integrity", "clearcover": "clearcover",
            "openly": "openly", "bristol": "bristol_west",
            "natgen": "national_general", "national_general": "national_general",
            "universal": "universal_property", "upcic": "universal_property",
            "american_modern": "american_modern", "covertree": "covertree",
        }
        has_any_carrier = any(p.get("carrier") for p in policies)
        if not has_any_carrier and file.filename:
            fn_lower = file.filename.lower()
            for pattern, carrier_key in FILENAME_CARRIER_MAP.items():
                if pattern in fn_lower:
                    for p in policies:
                        p["carrier"] = carrier_key
                    break

        # Detect carrier from policy number patterns (e.g. NatGen: 10+ digit "2027431477 00")
        for p in policies:
            if not p.get("carrier"):
                pn = (p.get("policy_number") or "").strip()
                for pat, ckey in _POLICY_NUM_PATTERNS:
                    if pat.match(pn):
                        p["carrier"] = ckey
                        break

        # Process each policy
        results = []
        matched = 0
        sent = 0
        letters = 0
        skipped = 0

        for pol in policies:
            pnum = (pol.get("policy_number") or "").strip()
            if not pnum:
                continue

            # Filter by cancellation reason
            notice_type = pol.get("notice_type", "non-pay")
            cancel_reason = pol.get("cancel_reason", "")

            reason_lower = cancel_reason.lower() if cancel_reason else ""
            is_nonpay = any(kw in reason_lower for kw in [
                "non payment", "non-payment", "nonpayment", "non pay", "non-pay",
                "nsf", "insufficient fund", "past due", "past-due",
            ])
            is_skip = any(kw in reason_lower for kw in [
                "underwriting", "policyholder request", "insured request",
                "company cancel", "non-renewal", "nonrenewal",
            ])

            if is_skip or (notice_type not in ("non-pay", "past-due") and not is_nonpay):
                results.append({
                    "policy_number": pnum,
                    "insured_name": pol.get("insured_name", ""),
                    "cancel_reason": cancel_reason,
                    "notice_type": notice_type,
                    "skipped_reason": True,
                    "error": f"Skipped — {cancel_reason}" if cancel_reason else f"Skipped — {notice_type}",
                })
                continue

            result = _process_single_policy(
                db=db,
                notice_id=notice.id,
                policy_number=pnum,
                carrier=pol.get("carrier", ""),
                insured_name=pol.get("insured_name", ""),
                amount_due=pol.get("amount_due"),
                due_date=pol.get("due_date"),
                cancel_date=pol.get("cancel_date"),
                dry_run=dry_run,
            )
            result["cancel_reason"] = cancel_reason
            result["notice_type"] = notice_type
            results.append(result)
            if result.get("matched"):
                matched += 1
            if result.get("email_sent"):
                sent += 1
            if result.get("skipped_rate_limit"):
                skipped += 1

        notice.policies_matched = matched
        notice.emails_sent = sent
        notice.emails_skipped = skipped
        notice.status = "dry_run" if dry_run else "completed"
        db.commit()

        return {
            "notice_id": notice.id,
            "filename": file.filename,
            "dry_run": dry_run,
            "policies_found": len(policies),
            "policies_matched": matched,
            "emails_sent": sent,
            "emails_skipped": skipped,
            "details": results,
        }

    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        try:
            notice.status = "error"
            notice.error_message = str(e)[:500]
            db.commit()
        except Exception:
            db.rollback()
        logger.error("Non-pay processing error: %s\n%s", e, tb)
        raise HTTPException(status_code=500, detail=f"Processing failed: {str(e)}\n{tb[-500:]}")
        raise HTTPException(status_code=500, detail=f"Processing failed: {str(e)}")


def _process_single_policy(
    db: Session,
    notice_id: int,
    policy_number: str,
    carrier: str,
    insured_name: str,
    amount_due: Optional[float],
    due_date: Optional[str],
    cancel_date: Optional[str] = None,
    dry_run: bool = False,
) -> dict:
    """Match a policy to a customer and send email if within rate limit."""
    result = {
        "policy_number": policy_number,
        "carrier": carrier,
        "matched": False,
        "customer_name": None,
        "customer_email": None,
        "email_sent": False,
        "skipped_rate_limit": False,
        "error": None,
    }

    # Find policy in our DB
    policy = db.query(CustomerPolicy).filter(
        CustomerPolicy.policy_number == policy_number
    ).first()

    if not policy:
        # Try with spaces/dashes removed (NatGen: "2032293985 00" vs DB "203229398500")
        compact = policy_number.replace(" ", "").replace("-", "").replace("	", "").strip()
        if compact != policy_number:
            policy = db.query(CustomerPolicy).filter(
                CustomerPolicy.policy_number == compact
            ).first()

    if not policy:
        # Try partial match (some reports truncate policy numbers)
        policy = db.query(CustomerPolicy).filter(
            CustomerPolicy.policy_number.ilike(f"%{policy_number}%")
        ).first()

    if not policy:
        # Try partial match with compact version
        compact = policy_number.replace(" ", "").replace("-", "").replace("	", "").strip()
        if compact != policy_number:
            policy = db.query(CustomerPolicy).filter(
                CustomerPolicy.policy_number.ilike(f"%{compact}%")
            ).first()

    if not policy:
        # Try base number (strip suffix like 618207668-653-1 → 618207668)
        base_number = policy_number.split("-")[0].split()[0].strip()
        if base_number and base_number != policy_number:
            policy = db.query(CustomerPolicy).filter(
                CustomerPolicy.policy_number.ilike(f"%{base_number}%")
            ).first()

    if not policy:
        # Try reverse: maybe DB has longer number that contains our extracted number
        compact = policy_number.replace(" ", "").replace("-", "").replace("	", "")
        policy = db.query(CustomerPolicy).filter(
            CustomerPolicy.policy_number.ilike(f"{compact}%")
        ).first()

    if not policy:
        # Try matching by customer name if we have insured_name
        if insured_name:
            parts = insured_name.strip().split()
            if len(parts) >= 2:
                # Try last name match on customers
                last_name = parts[-1]
                customer = db.query(Customer).filter(
                    Customer.last_name.ilike(f"%{last_name}%")
                ).first()
                if customer:
                    result["matched"] = True
                    result["customer_name"] = customer.full_name
                    result["customer_email"] = customer.email
                    result["customer_id"] = customer.id
                    result["match_type"] = "name"
                    if not customer.email:
                        # Try Thanks.io letter for name-matched customers without email
                        if customer.address and customer.city and customer.state and customer.zip_code:
                            # Check rate limit before sending letter
                            one_week_ago = datetime.utcnow() - timedelta(days=7)
                            recent_contact = db.query(NonPayEmail).filter(
                                NonPayEmail.policy_number == policy_number,
                                NonPayEmail.email_status.in_(["sent", "letter_sent"]),
                                NonPayEmail.sent_at >= one_week_ago,
                            ).first()
                            if recent_contact:
                                result["skipped_rate_limit"] = True
                                result["last_sent"] = recent_contact.sent_at.isoformat() if recent_contact.sent_at else None
                                return result
                            if dry_run:
                                result["would_send_letter"] = True
                                result["letter_address"] = f"{customer.address}, {customer.city}, {customer.state} {customer.zip_code}"
                                result["dry_run"] = True
                                return result
                            from app.services.thanksio_letter import send_thanksio_letter
                            letter_result = send_thanksio_letter(
                                client_name=customer.full_name,
                                address=customer.address,
                                city=customer.city,
                                state=customer.state,
                                zip_code=customer.zip_code,
                                policy_number=policy_number,
                                carrier=carrier,
                                amount_due=float(amount_due) if amount_due else None,
                                due_date=due_date,
                            )
                            letter_record = NonPayEmail(
                                notice_id=notice_id, policy_number=policy_number,
                                customer_id=customer.id, customer_name=customer.full_name,
                                customer_email=None, carrier=carrier,
                                amount_due=amount_due, due_date=due_date,
                                email_status="letter_sent" if letter_result.get("success") else "letter_failed",
                                mailgun_message_id=letter_result.get("order_id"),
                                error_message=letter_result.get("error"),
                            )
                            db.add(letter_record)
                            db.commit()
                            result["letter_sent"] = letter_result.get("success", False)
                            return result
                        else:
                            result["error"] = "No email and incomplete mailing address"
                            return result
                    # Skip rate limit check and sending for name matches in case of ambiguity
                    # Actually, DO check rate limit to avoid duplicate sends
                    one_week_ago = datetime.utcnow() - timedelta(days=7)
                    recent_email = db.query(NonPayEmail).filter(
                        NonPayEmail.policy_number == policy_number,
                        NonPayEmail.email_status.in_(["sent", "letter_sent"]),
                        NonPayEmail.sent_at >= one_week_ago,
                    ).first()
                    if recent_email:
                        result["skipped_rate_limit"] = True
                        result["last_sent"] = recent_email.sent_at.isoformat() if recent_email.sent_at else None
                        return result
                    if dry_run:
                        result["would_send"] = True
                        result["dry_run"] = True
                        return result
                    # For live mode, proceed to send
                    effective_carrier = carrier or ""
                    email_result = send_nonpay_email(
                        to_email=customer.email,
                        client_name=customer.full_name,
                        policy_number=policy_number,
                        carrier=effective_carrier,
                        amount_due=amount_due,
                        due_date=due_date,
                        cancel_date=cancel_date,
                    )
                    email_record = NonPayEmail(
                        notice_id=notice_id,
                        policy_number=policy_number,
                        customer_id=customer.id,
                        customer_name=customer.full_name,
                        customer_email=customer.email,
                        carrier=effective_carrier,
                        amount_due=amount_due,
                        due_date=due_date,
                        email_status="sent" if email_result.get("success") else "failed",
                        mailgun_message_id=email_result.get("message_id"),
                        error_message=email_result.get("error"),
                    )
                    db.add(email_record)
                    db.commit()
                    result["email_sent"] = email_result.get("success", False)
                    return result

    if not policy:
        result["error"] = "Policy not found in database"
        return result

    # Get customer — try local DB first (fast), then NowCerts live if needed
    customer = None
    nowcerts_customer = None

    # First try local DB (fast path)
    policy_found = policy  # already found above
    if policy_found:
        customer = db.query(Customer).filter(Customer.id == policy_found.customer_id).first()

    # Only do NowCerts live lookup if local DB didn't find the customer
    # or if we want to verify/update the name (skip for speed during bulk uploads)
    if not customer:
        try:
            from app.services.nowcerts import get_nowcerts_client
            nc = get_nowcerts_client()
            nc_results = nc.search_by_policy_number(policy_number)
            if nc_results:
                nowcerts_customer = nc_results[0]
                logger.info(f"NowCerts live lookup for {policy_number}: {nowcerts_customer.get('name', 'unknown')}")
        except Exception as e:
            logger.warning(f"NowCerts live lookup failed for {policy_number}, falling back to local DB: {e}")

    if nowcerts_customer and not customer:
        # NowCerts found a customer that local DB didn't have
        nc_name = nowcerts_customer.get("name") or nowcerts_customer.get("commercial_name") or ""
        nc_email = nowcerts_customer.get("email") or ""
        nc_address = nowcerts_customer.get("address") or ""
        nc_city = nowcerts_customer.get("city") or ""
        nc_state = nowcerts_customer.get("state") or ""
        nc_zip = nowcerts_customer.get("zip") or nowcerts_customer.get("zip_code") or ""

        class _TempCustomer:
            pass
        customer = _TempCustomer()
        customer.id = None
        customer.full_name = nc_name or insured_name
        customer.email = nc_email or None
        customer.address = nc_address or None
        customer.city = nc_city or None
        customer.state = nc_state or None
        customer.zip_code = nc_zip or None

        result["matched"] = True
        result["customer_name"] = nc_name or insured_name
        result["customer_email"] = nc_email or None
        result["customer_id"] = None
        result["match_source"] = "nowcerts_live"
    elif customer:
        # Local DB found the customer (fast path — most common)
        result["matched"] = True
        result["customer_name"] = customer.full_name
        result["customer_email"] = customer.email
        result["customer_id"] = customer.id
        result["match_source"] = "local_db"
    else:
        result["error"] = "Customer not found"
        return result

    # Use result dict for email/name (populated from either NowCerts or local DB)
    cust_email = result.get("customer_email")
    cust_name = result.get("customer_name", insured_name)

    if not cust_email:
        # No email — try sending a physical letter via Thanks.io
        if customer.address and customer.city and customer.state and customer.zip_code:
            # Check 1x/week rate limit for letters (applies in both dry_run and live)
            one_week_ago = datetime.utcnow() - timedelta(days=7)
            recent_letter = db.query(NonPayEmail).filter(
                NonPayEmail.policy_number == policy_number,
                NonPayEmail.email_status.in_(["letter_sent", "sent"]),
                NonPayEmail.sent_at >= one_week_ago,
            ).first()
            if recent_letter:
                result["skipped_rate_limit"] = True
                result["error"] = "Already contacted this week"
                result["last_sent"] = recent_letter.sent_at.isoformat() if recent_letter.sent_at else None
                return result

            if dry_run:
                result["would_send_letter"] = True
                result["letter_address"] = f"{customer.address}, {customer.city}, {customer.state} {customer.zip_code}"
                result["dry_run"] = True
                return result

            # Check 1x/week rate limit for letters too
            # (Already checked above for both dry_run and live)

            from app.services.thanksio_letter import send_thanksio_letter
            letter_result = send_thanksio_letter(
                client_name=customer.full_name,
                address=customer.address,
                city=customer.city,
                state=customer.state,
                zip_code=customer.zip_code,
                policy_number=policy_number,
                carrier=carrier,
                amount_due=float(amount_due) if amount_due else None,
                due_date=due_date,
            )

            # Record the letter
            letter_record = NonPayEmail(
                notice_id=notice_id,
                policy_number=policy_number,
                customer_id=customer.id,
                customer_name=customer.full_name,
                customer_email=None,
                carrier=carrier,
                amount_due=amount_due,
                due_date=due_date,
                email_status="letter_sent" if letter_result.get("success") else "letter_failed",
                mailgun_message_id=letter_result.get("order_id"),
                error_message=letter_result.get("error"),
            )
            db.add(letter_record)
            db.commit()

            result["letter_sent"] = letter_result.get("success", False)
            result["letter_order_id"] = letter_result.get("order_id")
            if not letter_result.get("success"):
                result["error"] = letter_result.get("error")
            return result
        else:
            result["error"] = "No email and incomplete mailing address"
            return result

    # Check 1x/week rate limit for this policy
    one_week_ago = datetime.utcnow() - timedelta(days=7)
    recent_email = db.query(NonPayEmail).filter(
        NonPayEmail.policy_number == policy_number,
        NonPayEmail.email_status == "sent",
        NonPayEmail.sent_at >= one_week_ago,
    ).first()

    if recent_email:
        result["skipped_rate_limit"] = True
        result["last_sent"] = recent_email.sent_at.isoformat() if recent_email.sent_at else None
        return result

    # Use carrier from policy record if not in the upload
    effective_carrier = carrier or policy.carrier or ""

    # In dry_run mode, report what WOULD happen but don't send
    if dry_run:
        result["email_sent"] = False
        result["would_send"] = True
        result["dry_run"] = True
        return result

    # Send the email
    # Record the email BEFORE sending (prevents duplicates on retry/timeout)
    email_record = NonPayEmail(
        notice_id=notice_id,
        policy_number=policy_number,
        customer_id=customer.id,
        customer_name=customer.full_name,
        customer_email=customer.email,
        carrier=effective_carrier,
        amount_due=amount_due,
        due_date=due_date,
        email_status="sent",  # Optimistic — prevents retries from resending
        error_message=None,
    )
    db.add(email_record)
    db.commit()

    email_result = send_nonpay_email(
        to_email=customer.email,
        client_name=customer.full_name,
        policy_number=policy_number,
        carrier=effective_carrier,
        amount_due=amount_due,
        due_date=due_date,
        cancel_date=cancel_date,
    )

    # Update with actual result
    email_record.mailgun_message_id = email_result.get("message_id")
    if not email_result.get("success"):
        email_record.email_status = "failed"
        email_record.error_message = email_result.get("error")
    db.commit()

    result["email_sent"] = email_result.get("success", False)
    if not email_result.get("success"):
        result["error"] = email_result.get("error")

    return result


# ── File Extraction ──────────────────────────────────────────────────

async def _extract_from_pdf(pdf_bytes: bytes) -> list[dict]:
    """Use Claude API to extract policy info from a PDF."""
    if not settings.ANTHROPIC_API_KEY:
        raise ValueError("ANTHROPIC_API_KEY not configured")

    # Truncate large PDFs
    try:
        from PyPDF2 import PdfReader, PdfWriter
        reader = PdfReader(io.BytesIO(pdf_bytes))
        if len(reader.pages) > 50:
            writer = PdfWriter()
            for i in range(50):
                writer.add_page(reader.pages[i])
            buf = io.BytesIO()
            writer.write(buf)
            pdf_bytes = buf.getvalue()
    except Exception:
        pass

    pdf_b64 = base64.standard_b64encode(pdf_bytes).decode("utf-8")

    async with httpx.AsyncClient(timeout=90.0) as client:
        response = await client.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": settings.ANTHROPIC_API_KEY,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json",
            },
            json={
                "model": "claude-sonnet-4-20250514",
                "max_tokens": 4000,
                "messages": [{
                    "role": "user",
                    "content": [
                        {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_b64}},
                        {"type": "text", "text": NONPAY_EXTRACTION_PROMPT},
                    ],
                }],
            },
        )

    if response.status_code != 200:
        raise ValueError(f"Claude API error ({response.status_code}): {response.text[:300]}")

    text = ""
    for block in response.json().get("content", []):
        if block.get("type") == "text":
            text += block["text"]

    text = text.strip()
    for fence in ["```json", "```"]:
        if text.startswith(fence):
            text = text[len(fence):]
    if text.endswith("```"):
        text = text[:-3]
    text = text.strip()

    try:
        data = json.loads(text)
        return data if isinstance(data, list) else [data]
    except json.JSONDecodeError as e:
        raise ValueError(f"Failed to parse extraction: {e}\nRaw: {text[:500]}")


def _extract_from_csv(file_bytes: bytes) -> list[dict]:
    """Parse CSV/TSV to extract policy numbers and amounts."""
    text = file_bytes.decode("utf-8", errors="replace")

    # Detect delimiter
    if "\t" in text.split("\n")[0]:
        delimiter = "\t"
    else:
        delimiter = ","

    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    results = []

    # Common column name patterns
    policy_cols = ["policy_number", "policynumber", "policy #", "policy#", "policy no",
                   "policyno", "policy", "pol_number", "pol_num", "pol#", "number"]
    carrier_cols = ["carrier", "carrier_name", "carriername", "company", "insurer", "insurance_company"]
    name_cols = ["insured_name", "insuredname", "name", "client", "customer", "policyholder",
                 "insured", "named_insured", "named insured", "first name", "first_name"]
    amount_cols = ["amount_due", "amountdue", "amount", "balance", "premium_due", "premiumdue",
                   "past_due", "pastdue", "total_due", "totaldue", "premium",
                   "minimum_due", "remaining_balance", "amount_to_reinstate"]
    date_cols = ["due_date", "duedate", "cancel_date", "canceldate", "effective_date",
                 "cancellation_date", "cancellationdate",
                 "payment_due_date", "cancellation_effective_date"]

    def _find_col(fieldnames, patterns):
        for f in (fieldnames or []):
            fl = f.lower().strip().replace(" ", "_")
            if fl in patterns:
                return f
        return None

    fields = reader.fieldnames or []
    p_col = _find_col(fields, policy_cols)
    c_col = _find_col(fields, carrier_cols)
    n_col = _find_col(fields, name_cols)
    a_col = _find_col(fields, amount_cols)
    d_col = _find_col(fields, date_cols)

    for row in reader:
        pnum = row.get(p_col, "").strip() if p_col else ""
        if not pnum:
            continue

        amt = None
        if a_col and row.get(a_col):
            try:
                amt = float(row[a_col].replace(",", "").replace("$", "").strip())
            except (ValueError, AttributeError):
                pass

        results.append({
            "policy_number": pnum,
            "carrier": row.get(c_col, "").strip() if c_col else "",
            "insured_name": row.get(n_col, "").strip() if n_col else "",
            "amount_due": amt,
            "due_date": row.get(d_col, "").strip() if d_col else "",
            "notice_type": "non-pay",
        })

    return results


# ── Progressive File Detection & Processing ────────────────────────────

def _detect_progressive_file(file_bytes: bytes, ext: str, filename: str = "") -> Optional[dict]:
    """Detect and parse Progressive-specific file formats.
    
    Returns dict with 'policies' (non-pay items) and 'uw_items' (UW requirements),
    or None if this isn't a recognized Progressive format.
    
    Recognized formats:
    1. PoliciesPendingCancelOrRenewal — multi-sheet (Non-Payment, Underwriting, Pending Renewal)
    2. CustomerFollowup — single sheet with Message Subject / Action columns (UW items)
    """
    import io

    if ext != "xlsx":
        return None

    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception:
        return None

    sheet_names = [s.lower() for s in wb.sheetnames]

    # ── Check for Progressive agent code 03BXR in any sheet ──
    is_progressive_by_agent_code = False
    for sheet_name in wb.sheetnames:
        ws_check = wb[sheet_name]
        rows_check = list(ws_check.iter_rows(values_only=True))
        if len(rows_check) < 2:
            continue
        headers_check = [str(c).lower().strip() if c else "" for c in rows_check[0]]
        agent_col = next((i for i, h in enumerate(headers_check) if h == "agent code"), None)
        if agent_col is not None:
            for row in rows_check[1:]:
                if agent_col < len(row) and str(row[agent_col] or "").strip() == "03BXR":
                    is_progressive_by_agent_code = True
                    break
        if is_progressive_by_agent_code:
            break

    if not is_progressive_by_agent_code:
        return None

    logger.info("Progressive file detected (Agent Code 03BXR)")

    # ── Format 1: PoliciesPendingCancelOrRenewal ──
    if "non-payment" in sheet_names or "underwriting" in sheet_names:
        logger.info("Progressive PoliciesPendingCancel format detected")
        policies = []
        uw_items = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                continue

            headers = [str(c).strip() if c else "" for c in rows[0]]
            col_map = _progressive_col_map(headers)

            sheet_lower = sheet_name.lower()
            for row in rows[1:]:
                cells = list(row)
                record = _progressive_row_to_dict(cells, col_map, headers)
                if not record.get("policy_number"):
                    continue
                record["carrier"] = "progressive"

                if "non-payment" in sheet_lower or "non payment" in sheet_lower:
                    record["notice_type"] = "non-pay"
                    record["cancel_reason"] = "Pending Cancel for Non Payment"
                    policies.append(record)
                elif "underwriting" in sheet_lower:
                    record["notice_type"] = "underwriting"
                    record["cancel_reason"] = record.get("cancel_reason", "Underwriting")
                    record["requirement_type"] = "general_uw"
                    uw_items.append(record)
                elif "pending renewal" in sheet_lower or "renewal" in sheet_lower:
                    # Skip renewals for now — could add renewal tracking later
                    pass

        return {"policies": policies, "uw_items": uw_items}

    # ── Format 2: CustomerFollowup (UW compliance) ──
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return None

    headers = [str(c).strip() if c else "" for c in rows[0]]
    headers_lower = [h.lower() for h in headers]

    # Detect by "Message Subject" column (unique to CustomerFollowup)
    if "message subject" in headers_lower:
        logger.info("Progressive CustomerFollowup format detected")
        col_map = _progressive_col_map(headers)
        msg_col = next((i for i, h in enumerate(headers_lower) if h == "message subject"), None)
        memo_col = next((i for i, h in enumerate(headers_lower) if h == "memo sent"), None)

        uw_items = []
        for row in rows[1:]:
            cells = list(row)
            record = _progressive_row_to_dict(cells, col_map, headers)
            if not record.get("policy_number"):
                continue
            record["carrier"] = "progressive"

            # Map Message Subject to UW requirement type
            msg_subject = str(cells[msg_col]).strip() if msg_col is not None and msg_col < len(cells) and cells[msg_col] else ""
            memo_date = str(cells[memo_col]).strip() if memo_col is not None and memo_col < len(cells) and cells[memo_col] else ""

            record["message_subject"] = msg_subject
            record["memo_sent"] = memo_date
            record["requirement_type"] = _map_progressive_uw_type(msg_subject)
            record["notice_type"] = "uw_requirement"
            uw_items.append(record)

        return {"policies": [], "uw_items": uw_items}

    return None


def _progressive_col_map(headers: list[str]) -> dict:
    """Map Progressive column headers to standard field names."""
    headers_lower = [h.lower().strip() for h in headers]
    col_map = {}

    mappings = {
        "policy_number": ["policy number", "policy", "policy no"],
        "insured_name": ["full name", "named insured", "name", "insured"],
        "email": ["email address", "email"],
        "phone": ["phone number", "phone", "phone #"],
        "address": ["address"],
        "city": ["city"],
        "state": ["state", "policy state"],
        "zip": ["zip", "zip code"],
        "product": ["product"],
        "producer": ["producer"],
        "agent_code": ["agent code"],
        "amount_due": ["amount due", "amount"],
        "cancel_date": ["cancel effective date", "cancel date", "cancellation date"],
        "cancel_reason": ["cancel reason", "reason"],
        "reference_number": ["reference number"],
    }

    for field, patterns in mappings.items():
        for i, h in enumerate(headers_lower):
            if h in patterns:
                col_map[field] = i
                break

    return col_map


def _progressive_row_to_dict(cells: list, col_map: dict, headers: list) -> dict:
    """Convert a Progressive row to a standard policy dict."""
    def _get(field):
        idx = col_map.get(field)
        if idx is not None and idx < len(cells) and cells[idx] is not None:
            return str(cells[idx]).strip()
        return ""

    # Handle "Last, First" name format
    raw_name = _get("insured_name")
    if "," in raw_name:
        parts = raw_name.split(",", 1)
        name = f"{parts[1].strip()} {parts[0].strip()}"
    else:
        name = raw_name

    # Parse amount
    amount = None
    raw_amt = _get("amount_due")
    if raw_amt:
        try:
            amount = float(raw_amt.replace(",", "").replace("$", ""))
        except (ValueError, TypeError):
            pass

    # Parse cancel date and compute payment due date (day before cancellation)
    raw_cancel = _get("cancel_date")
    due_date = ""
    cancel_display = ""
    if raw_cancel:
        import datetime as _dt
        parsed = None
        # openpyxl may return datetime object → str becomes "2026-03-02 00:00:00"
        for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"]:
            try:
                parsed = _dt.datetime.strptime(raw_cancel, fmt)
                break
            except ValueError:
                continue
        if parsed:
            cancel_display = parsed.strftime("%m/%d/%Y")
            payment_due = parsed - _dt.timedelta(days=1)
            due_date = payment_due.strftime("%m/%d/%Y")
        else:
            due_date = raw_cancel  # fallback to raw string

    return {
        "policy_number": _get("policy_number"),
        "insured_name": name,
        "email": _get("email"),
        "phone": _get("phone"),
        "address": _get("address"),
        "city": _get("city"),
        "state": _get("state"),
        "zip": _get("zip"),
        "product": _get("product"),
        "producer": _get("producer"),
        "agent_code": _get("agent_code"),
        "amount_due": amount,
        "due_date": due_date,
        "cancel_date": cancel_display,
        "cancel_reason": _get("cancel_reason"),
        "carrier": "progressive",
    }


def _map_progressive_uw_type(message_subject: str) -> str:
    """Map Progressive's Message Subject to a UW requirement type."""
    subj = message_subject.lower().strip()

    if "vehicle photo" in subj:
        return "vehicle_photos"
    if "mileage" in subj or "annual mileage" in subj or "odometer" in subj:
        return "proof_of_mileage"
    if "residence" in subj or "homeowner" in subj or "renter" in subj:
        return "proof_of_residence"
    if "discount" in subj:
        return "discount_verification"
    if "continuous" in subj or "prior" in subj or "proof of insurance" in subj:
        return "proof_of_continuous_insurance"
    if "photo" in subj:
        return "vehicle_photos"

    return "general_uw"


def _process_progressive_uw_items(db: Session, uw_items: list[dict], dry_run: bool = False, notice_id: int = None):
    """Create UW tasks and optionally send emails for Progressive UW items."""
    from app.api.tasks import create_uw_requirement_task
    from app.services.uw_requirement_email import send_uw_requirement_email

    for item in uw_items:
        policy_number = item.get("policy_number", "")
        customer_name = item.get("insured_name", "")
        email = item.get("email", "")
        carrier = "progressive"
        requirement_type = item.get("requirement_type", "general_uw")
        producer = item.get("producer", "")
        memo_date = item.get("memo_sent", "")
        msg_subject = item.get("message_subject", "")

        # Create task in Compliance Center (even in dry_run — tasks are informational)
        try:
            task = create_uw_requirement_task(
                db=db,
                customer_name=customer_name,
                policy_number=policy_number,
                carrier=carrier,
                requirement_type=requirement_type,
                due_date=memo_date,  # Use memo sent date as reference
                producer_name=producer,
            )
            # Update task description with Progressive-specific details
            if task and msg_subject:
                task.description = (
                    f"Progressive UW Requirement: {msg_subject}\n"
                    f"Policy: {policy_number} ({item.get('product', '')})\n"
                    f"Customer: {customer_name}\n"
                    f"Email: {email or 'N/A'}\n"
                    f"Producer: {producer}\n"
                    f"Memo Sent: {memo_date}\n\n"
                    f"Action: Contact customer to obtain {msg_subject.lower()} documentation."
                )
                task.source = "progressive_upload"
                db.commit()
            logger.info("UW task created for %s / %s: %s", customer_name, policy_number, msg_subject)
        except Exception as e:
            logger.warning("UW task creation failed for %s: %s", policy_number, e)

        # Send UW email to customer (skip in dry_run)
        if not dry_run and email:
            try:
                send_uw_requirement_email(
                    to_email=email,
                    client_name=customer_name,
                    policy_number=policy_number,
                    carrier=carrier,
                    requirement_type=requirement_type,
                    due_date=memo_date or None,
                    producer_name=producer,
                )
                logger.info("UW email sent to %s for %s", email, policy_number)
            except Exception as e:
                logger.warning("UW email failed for %s: %s", policy_number, e)


def _extract_from_excel(file_bytes: bytes, ext: str) -> list[dict]:
    """Extract policy data from .xlsx or .xls files."""
    import io

    # Column name patterns (same as CSV)
    policy_pats = ["policy_number", "policynumber", "policy #", "policy#", "policy no",
                   "policyno", "policy", "pol_number", "pol_num", "pol#", "number"]
    carrier_pats = ["carrier", "carrier_name", "carriername", "company", "insurer", "insurance_company"]
    name_pats = ["insured_name", "insuredname", "name", "client", "customer", "policyholder",
                 "insured", "named_insured", "named insured", "first name", "first_name"]
    amount_pats = ["amount_due", "amountdue", "amount", "balance", "premium_due", "premiumdue",
                   "past_due", "pastdue", "total_due", "totaldue", "premium",
                   "minimum_due", "remaining_balance", "amount_to_reinstate"]
    date_pats = ["due_date", "duedate", "cancel_date", "canceldate", "effective_date",
                 "cancellation_date", "cancellationdate",
                 "payment_due_date", "cancellation_effective_date"]
    reason_pats = ["reason", "cancel_reason", "cancellation_reason", "cancel_type",
                   "notice_reason", "status", "cancellation_status"]
    phone_pats = ["phone", "phone_#", "phone_number", "phonenumber", "phone_no",
                  "telephone", "cell", "mobile"]

    def _match_col(headers, patterns):
        for i, h in enumerate(headers):
            if h and str(h).lower().strip().replace(" ", "_") in patterns:
                return i
        return None

    results = []

    if ext == "xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    else:  # xls
        try:
            import xlrd
            wb = xlrd.open_workbook(file_contents=file_bytes)
            ws = wb.sheet_by_index(0)
            rows = [ws.row_values(r) for r in range(ws.nrows)]
        except ImportError:
            # xlrd not installed — try reading as CSV (some .xls are actually HTML/CSV)
            try:
                text = file_bytes.decode("utf-8", errors="ignore")
                reader = csv.reader(io.StringIO(text), delimiter="\t")
                rows = [row for row in reader]
            except Exception:
                raise ValueError("XLS support requires xlrd package. Please convert to XLSX or CSV.")
        except Exception as e:
            raise ValueError(f"Failed to read XLS file: {str(e)}")

    if not rows:
        return results

    headers = [str(c).strip() if c else "" for c in rows[0]]
    p_col = _match_col(headers, policy_pats)
    c_col = _match_col(headers, carrier_pats)
    n_col = _match_col(headers, name_pats)
    a_col = _match_col(headers, amount_pats)
    d_col = _match_col(headers, date_pats)
    r_col = _match_col(headers, reason_pats)
    ph_col = _match_col(headers, phone_pats)

    for row in rows[1:]:
        cells = list(row)
        pnum = str(cells[p_col]).strip() if p_col is not None and p_col < len(cells) and cells[p_col] else ""
        if not pnum or pnum.lower() == "none":
            continue

        amt = None
        if a_col is not None and a_col < len(cells) and cells[a_col]:
            try:
                val = cells[a_col]
                if isinstance(val, (int, float)):
                    amt = float(val)
                else:
                    amt = float(str(val).replace(",", "").replace("$", "").strip())
            except (ValueError, TypeError):
                pass

        # Extract cancellation reason
        reason_raw = ""
        if r_col is not None and r_col < len(cells) and cells[r_col]:
            reason_raw = str(cells[r_col]).strip()

        # Classify the reason
        reason_lower = reason_raw.lower()
        if any(kw in reason_lower for kw in ["non payment", "non-payment", "nonpayment", "nsf",
                                              "non pay", "non-pay", "nonpay",
                                              "insufficient funds", "returned payment",
                                              "offer", "pending cancellation"]):
            notice_type = "non-pay"
        elif any(kw in reason_lower for kw in ["underwriting", "uw reason"]):
            notice_type = "underwriting"
        elif any(kw in reason_lower for kw in ["policyholder", "insured request", "customer request",
                                                "rewrite", "replacement", "policyholder's request"]):
            notice_type = "voluntary"
        elif reason_raw:
            notice_type = "other"
        else:
            notice_type = "non-pay"  # default if no reason column

        # Extract phone
        phone = ""
        if ph_col is not None and ph_col < len(cells) and cells[ph_col]:
            phone = str(cells[ph_col]).strip()

        results.append({
            "policy_number": pnum,
            "carrier": str(cells[c_col]).strip() if c_col is not None and c_col < len(cells) and cells[c_col] else "",
            "insured_name": str(cells[n_col]).strip() if n_col is not None and n_col < len(cells) and cells[n_col] else "",
            "amount_due": amt,
            "due_date": str(cells[d_col]).strip() if d_col is not None and d_col < len(cells) and cells[d_col] else "",
            "notice_type": notice_type,
            "cancel_reason": reason_raw,
            "phone": phone,
        })

    return results


# ── Inbound Email (Mailgun webhook) ─────────────────────────────────

import os
import re
from fastapi import Request, Form

# Env var to control mode: "dry_run" or "live"
# Start with dry_run, switch to live after first week
INBOUND_NONPAY_MODE = os.environ.get("INBOUND_NONPAY_MODE", "dry_run")

# Subject line keywords that indicate non-pay notices
NONPAY_SUBJECT_KEYWORDS = ["non pay", "non-pay", "nonpay", "nsf", "non payment", "non-payment"]
SKIP_SUBJECT_KEYWORDS = ["underwriting", "policyholder request", "rewrite", "replacement"]


def _parse_grangewire_html(html_body: str) -> list[dict]:
    """Parse GrangeWire Alerts HTML table into policy records."""
    from html.parser import HTMLParser

    class TableParser(HTMLParser):
        def __init__(self):
            super().__init__()
            self.in_table = False
            self.in_row = False
            self.in_cell = False
            self.current_row = []
            self.current_cell = ""
            self.rows = []
            self.table_count = 0

        def handle_starttag(self, tag, attrs):
            if tag == "table":
                self.in_table = True
                self.table_count += 1
            elif tag == "tr" and self.in_table:
                self.in_row = True
                self.current_row = []
            elif tag in ("td", "th") and self.in_row:
                self.in_cell = True
                self.current_cell = ""
            elif tag == "br" and self.in_cell:
                self.current_cell += " "

        def handle_endtag(self, tag):
            if tag == "table":
                self.in_table = False
            elif tag == "tr" and self.in_row:
                self.in_row = False
                if self.current_row:
                    self.rows.append(self.current_row)
            elif tag in ("td", "th") and self.in_cell:
                self.in_cell = False
                self.current_row.append(self.current_cell.strip())

        def handle_data(self, data):
            if self.in_cell:
                self.current_cell += data

    parser = TableParser()
    parser.feed(html_body)

    if not parser.rows:
        return []

    # Find the header row — look for one containing "POLICY" or "ACCT"
    header_idx = None
    for i, row in enumerate(parser.rows):
        row_text = " ".join(row).upper()
        if "POLICY" in row_text or "ACCT" in row_text:
            header_idx = i
            break

    if header_idx is None:
        return []

    headers = [h.upper().strip() for h in parser.rows[header_idx]]

    # Map columns
    def _find_col(headers, keywords):
        for i, h in enumerate(headers):
            for kw in keywords:
                if kw in h:
                    return i
        return None

    p_col = _find_col(headers, ["POLICY", "ACCT", "NUMBER"])
    n_col = _find_col(headers, ["INSURED", "NAME"])
    d_col = _find_col(headers, ["CANCEL", "DATE"])
    max_col = _find_col(headers, ["MAX DUE", "MAX"])
    min_col = _find_col(headers, ["MIN DUE", "MIN"])
    msg_col = _find_col(headers, ["MESSAGE"])
    phone_col = _find_col(headers, ["PHONE", "EMAIL"])

    results = []
    for row in parser.rows[header_idx + 1:]:
        if not row or len(row) <= (p_col or 0):
            continue

        pnum = row[p_col].strip() if p_col is not None and p_col < len(row) else ""
        if not pnum:
            continue

        # Parse amount — prefer MAX DUE
        amount = None
        for col in [max_col, min_col]:
            if col is not None and col < len(row) and row[col]:
                try:
                    amount = float(row[col].replace(",", "").replace("$", "").strip())
                    break
                except (ValueError, TypeError):
                    pass

        # Parse message to determine notice type
        message = row[msg_col].strip() if msg_col is not None and msg_col < len(row) else ""
        msg_lower = message.lower()
        if any(kw in msg_lower for kw in ["non pay", "non-pay", "nsf", "offer", "pending cancellation"]):
            notice_type = "non-pay"
        elif any(kw in msg_lower for kw in ["underwriting"]):
            notice_type = "underwriting"
        elif any(kw in msg_lower for kw in ["policyholder", "rewrite"]):
            notice_type = "voluntary"
        else:
            notice_type = "non-pay"  # default for Grange non-pay alerts

        results.append({
            "policy_number": pnum,
            "carrier": "grange",
            "insured_name": row[n_col].strip() if n_col is not None and n_col < len(row) else "",
            "amount_due": amount,
            "due_date": row[d_col].strip() if d_col is not None and d_col < len(row) else "",
            "notice_type": notice_type,
            "cancel_reason": message,
            "phone": row[phone_col].strip() if phone_col is not None and phone_col < len(row) else "",
        })

    return results


def _parse_generic_email_html(html_body: str, carrier: str = "") -> list[dict]:
    """Fallback parser for non-Grange carrier email tables."""
    # Try the same table parsing approach
    policies = _parse_grangewire_html(html_body)
    if policies and not carrier:
        # Try to detect carrier from email body
        body_lower = html_body.lower()
        for key in ["travelers", "progressive", "safeco", "national general",
                     "bristol west", "hippo", "branch", "clearcover"]:
            if key in body_lower:
                carrier = key.replace(" ", "_")
                break
    for p in policies:
        if carrier and not p.get("carrier"):
            p["carrier"] = carrier
    return policies


@router.post("/inbound-email")
async def inbound_email_webhook(request: Request, db: Session = Depends(get_db)):
    """
    Mailgun inbound webhook. Receives parsed email data and processes non-pay notices.
    """
    try:
        form = await request.form()
    except Exception as e:
        logger.error("Inbound email: failed to parse form data: %s", e)
        return {"status": "error", "message": str(e)}

    # ── Verify Mailgun signature ──
    from app.api.email_inbox import _verify_mailgun_signature
    mg_timestamp = form.get("timestamp", "")
    mg_token = form.get("token", "")
    mg_signature = form.get("signature", "")
    if mg_timestamp and mg_token and mg_signature:
        if not _verify_mailgun_signature(mg_token, mg_timestamp, mg_signature):
            logger.warning("⚠️ Nonpay webhook: Mailgun signature verification FAILED")
            raise HTTPException(status_code=403, detail="Invalid signature")

    sender = form.get("sender", "") or form.get("from", "")
    subject = form.get("subject", "")
    html_body = form.get("body-html", "")
    plain_body = form.get("body-plain", "")

    logger.info("Inbound email from=%s subject=%s", sender, subject[:80])

    # Determine if this is a non-pay notice
    # Check subject line first, then fall back to scanning the HTML body
    # (Forwarded emails often have generic subjects like "Fwd: GrangeWire Alerts")
    subject_lower = subject.lower()
    html_lower = (html_body or "").lower()
    plain_lower = (plain_body or "").lower()
    all_text = f"{subject_lower} {html_lower} {plain_lower}"

    # ── National General Policy Activity → smart router ──
    is_natgen_activity = (
        ("policy activity" in subject_lower and ("ngic" in all_text or "national general" in all_text))
        or ("reports@ngic.com" in (sender or "").lower())
        or ("policy activity" in subject_lower and "outstanding to do" in html_lower)
    )
    if is_natgen_activity:
        logger.info("NatGen Policy Activity detected — routing through smart parser")
        try:
            return await _handle_natgen_policy_activity(html_body, sender, db)
        except Exception as e:
            logger.error("NatGen smart router failed: %s", e)
            import traceback
            traceback.print_exc()
            # Fall through to generic handler

    # ── Carrier Inspection Follow-Ups → inspection handler ──
    # Must run BEFORE Grange non-renewal check because inspection emails
    # can contain "non renewal" in warning text
    from app.services.inspection_email import is_inspection_email, handle_inspection_email
    if is_inspection_email(sender, subject, f"{html_body} {plain_body}"):
        logger.info("Carrier inspection email detected — routing to inspection handler")
        try:
            # Collect attachments from Mailgun form data
            inspection_attachments = []
            att_count = int(form.get("attachment-count", 0) or 0)
            for i in range(1, att_count + 1):
                att = form.get(f"attachment-{i}")
                if att and hasattr(att, 'filename') and hasattr(att, 'read'):
                    att_bytes = await att.read()
                    inspection_attachments.append((att.filename, att_bytes))
            
            return await handle_inspection_email(
                sender=sender,
                subject=subject,
                html_body=html_body,
                plain_body=plain_body,
                attachments=inspection_attachments,
                db=db,
            )
        except Exception as e:
            logger.error("Inspection handler failed: %s", e)
            import traceback
            traceback.print_exc()
            # Fall through to generic handler

    # ── Grange Non-Renewal Alerts → non-renewal handler ──
    # Only match actual GrangeWire Alert emails, not inspection follow-ups
    is_grange_nonrenewal = (
        ("grangewire" in all_text or "grange" in (sender or "").lower())
        and ("non-renewal" in all_text or "nonrenewal" in all_text or "non renewal" in all_text)
        and ("grangewire" in all_text or "non-renewal alert" in all_text or "non-renewals alert" in all_text)
    )

    # Check for skip keywords in subject
    if any(kw in subject_lower for kw in SKIP_SUBJECT_KEYWORDS):
        # But only skip if the body doesn't ALSO contain non-pay content
        if not any(kw in html_lower for kw in NONPAY_SUBJECT_KEYWORDS):
            logger.info("Inbound email skipped (subject keyword): %s", subject[:80])
            return {"status": "skipped", "reason": "Subject indicates non-actionable notice type"}

    # Check for non-pay keywords in subject OR body
    is_nonpay = any(kw in all_text for kw in NONPAY_SUBJECT_KEYWORDS)
    if not is_nonpay:
        logger.info("Inbound email skipped (no non-pay keyword in subject or body): %s", subject[:80])
        return {"status": "skipped", "reason": "No non-pay keywords found in subject or body"}

    # Detect carrier from sender, forwarded-from headers, or body content
    sender_lower = sender.lower()
    carrier = ""

    # Check the actual sender and also the forwarded message headers in the body
    carrier_checks = f"{sender_lower} {html_lower}"
    if "grange" in carrier_checks:
        carrier = "grange"
    elif "travelers" in carrier_checks:
        carrier = "travelers"
    elif "progressive" in carrier_checks:
        carrier = "progressive"
    elif "safeco" in carrier_checks:
        carrier = "safeco"
    elif "national" in carrier_checks and "general" in carrier_checks:
        carrier = "national_general"

    # Parse the HTML body for policy data
    if not html_body:
        logger.warning("Inbound email has no HTML body")
        return {"status": "error", "message": "No HTML body in email"}

    if "grange" in carrier:
        policies = _parse_grangewire_html(html_body)
    else:
        policies = _parse_generic_email_html(html_body, carrier)

    if not policies:
        logger.warning("Inbound email: no policies extracted from body")
        return {"status": "error", "message": "Could not extract policy data from email"}

    # Determine mode
    mode = INBOUND_NONPAY_MODE
    dry_run = (mode != "live")

    # Create a notice record
    notice = NonPayNotice(
        filename=f"inbound-email-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}",
        upload_type="inbound-email",
        uploaded_by=f"inbound:{sender[:60]}",
        policies_found=len(policies),
        status="processing",
    )
    db.add(notice)
    db.commit()
    db.refresh(notice)

    # Process each policy (same logic as file upload)
    results = []
    matched = 0
    sent = 0
    letters = 0
    skipped = 0

    for pol in policies:
        pnum = (pol.get("policy_number") or "").strip()
        if not pnum:
            continue

        notice_type = pol.get("notice_type", "non-pay")
        cancel_reason = pol.get("cancel_reason", "")

        reason_lower = cancel_reason.lower() if cancel_reason else ""
        is_nonpay = any(kw in reason_lower for kw in [
            "non payment", "non-payment", "nonpayment", "non pay", "non-pay",
            "nsf", "insufficient fund", "past due", "past-due",
        ])
        is_skip = any(kw in reason_lower for kw in [
            "underwriting", "policyholder request", "insured request",
            "company cancel", "non-renewal", "nonrenewal",
        ])

        if is_skip or (notice_type not in ("non-pay", "past-due") and not is_nonpay):
            results.append({
                "policy_number": pnum,
                "insured_name": pol.get("insured_name", ""),
                "cancel_reason": cancel_reason,
                "notice_type": notice_type,
                "skipped_reason": True,
                "error": f"Skipped — {cancel_reason}" if cancel_reason else f"Skipped — {notice_type}",
            })
            continue

        result = _process_single_policy(
            db=db,
            notice_id=notice.id,
            policy_number=pnum,
            carrier=pol.get("carrier", carrier),
            insured_name=pol.get("insured_name", ""),
            amount_due=pol.get("amount_due"),
            due_date=pol.get("due_date"),
            cancel_date=pol.get("cancel_date"),
            dry_run=dry_run,
        )
        result["cancel_reason"] = cancel_reason
        result["notice_type"] = notice_type
        results.append(result)
        if result.get("matched"):
            matched += 1
        if result.get("email_sent"):
            sent += 1
        if result.get("letter_sent"):
            letters += 1
        if result.get("skipped_rate_limit"):
            skipped += 1

    notice.policies_matched = matched
    notice.emails_sent = sent + letters
    notice.emails_skipped = skipped
    notice.status = "dry_run" if dry_run else "completed"
    db.commit()

    summary = {
        "status": "processed",
        "mode": "dry_run" if dry_run else "live",
        "notice_id": notice.id,
        "carrier": carrier,
        "subject": subject[:100],
        "policies_found": len(policies),
        "policies_matched": matched,
        "emails_sent": sent,
        "letters_sent": letters,
        "skipped": skipped,
        "details": results,
    }

    logger.info("Inbound email processed: %s policies, %s matched, %s sent (mode=%s)",
                len(policies), matched, sent, mode)

    return summary


# ── History / Status ─────────────────────────────────────────────────

@router.get("/history")
def nonpay_history(
    limit: int = 20,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get recent non-pay notice upload history."""
    notices = (
        db.query(NonPayNotice)
        .order_by(NonPayNotice.created_at.desc())
        .limit(limit)
        .all()
    )
    return {
        "notices": [{
            "id": n.id,
            "filename": n.filename,
            "upload_type": n.upload_type,
            "uploaded_by": n.uploaded_by,
            "policies_found": n.policies_found,
            "policies_matched": n.policies_matched,
            "emails_sent": n.emails_sent,
            "emails_skipped": n.emails_skipped,
            "status": n.status,
            "error_message": n.error_message,
            "created_at": n.created_at.isoformat() if n.created_at else None,
        } for n in notices]
    }


@router.get("/emails")
def nonpay_emails(
    policy_number: Optional[str] = None,
    limit: int = 50,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get non-pay email send history, optionally filtered by policy."""
    query = db.query(NonPayEmail).order_by(NonPayEmail.sent_at.desc())
    if policy_number:
        query = query.filter(NonPayEmail.policy_number == policy_number)
    emails = query.limit(limit).all()

    return {
        "emails": [{
            "id": e.id,
            "policy_number": e.policy_number,
            "customer_name": e.customer_name,
            "customer_email": e.customer_email,
            "carrier": e.carrier,
            "amount_due": float(e.amount_due) if e.amount_due else None,
            "due_date": e.due_date,
            "email_status": e.email_status,
            "sent_at": e.sent_at.isoformat() if e.sent_at else None,
        } for e in emails]
    }


@router.get("/preview")
def preview_nonpay_email(
    carrier: str = "progressive",
    client_name: str = "John Smith",
    policy_number: str = "AUT-12345678",
    amount_due: float = 247.50,
    due_date: str = "02/28/2026",
    current_user: User = Depends(get_current_user),
):
    """Preview a non-pay email template. Returns subject + raw HTML."""
    from app.services.nonpay_email import build_nonpay_email_html
    subject, html = build_nonpay_email_html(
        client_name=client_name,
        policy_number=policy_number,
        carrier=carrier,
        amount_due=amount_due,
        due_date=due_date,
    )
    return {"subject": subject, "html": html, "carrier": carrier}


@router.get("/carriers")
def list_nonpay_carriers(
    current_user: User = Depends(get_current_user),
):
    """List all carriers that have custom email templates."""
    from app.services.welcome_email import CARRIER_INFO
    carriers = []
    for key, info in CARRIER_INFO.items():
        carriers.append({
            "key": key,
            "display_name": info.get("display_name", key),
            "accent_color": info.get("accent_color", "#1a2b5f"),
            "has_payment_url": bool(info.get("payment_url")),
        })
    carriers.sort(key=lambda c: c["display_name"])
    return {"carriers": carriers}


@router.post("/send-test")
def send_test_nonpay_email(
    to_email: str,
    carrier: str = "progressive",
    client_name: str = "John Smith",
    policy_number: str = "AUT-12345678",
    amount_due: float = 247.50,
    due_date: str = "02/28/2026",
    current_user: User = Depends(get_current_user),
):
    """Send a test non-pay email to a specific address."""
    result = send_nonpay_email(
        to_email=to_email,
        client_name=client_name,
        policy_number=policy_number,
        carrier=carrier,
        amount_due=amount_due,
        due_date=due_date,
    )
    return {"success": result.get("success", False), "to": to_email, "carrier": carrier, "error": result.get("error")}


@router.post("/send-test-internal")
async def send_test_internal(request: Request):
    """Quick test endpoint - only sends to evan@betterchoiceins.com."""
    try:
        body = await request.json()
    except Exception:
        body = {}
    
    carrier = body.get("carrier", "grange")
    client_name = body.get("client_name", "Rosa Ayala")
    policy_number = body.get("policy_number", "HM 6605796")
    amount_due = body.get("amount_due", 1494.00)
    due_date = body.get("due_date", "02/28/2026")
    
    result = send_nonpay_email(
        to_email="evan@betterchoiceins.com",
        client_name=client_name,
        policy_number=policy_number,
        carrier=carrier,
        amount_due=amount_due,
        due_date=due_date,
    )
    return {"success": result.get("success", False), "to": "evan@betterchoiceins.com", "carrier": carrier, "error": result.get("error")}


@router.post("/test-nowcerts-note")
async def test_nowcerts_note(request: Request):
    """Test NowCerts note insertion directly — debug version."""
    try:
        from app.services.nowcerts import get_nowcerts_client
        nc = get_nowcerts_client()
        if not nc.is_configured:
            return {"error": "NowCerts not configured", "username": bool(nc.username), "password": bool(nc.password)}

        body = await request.json() if request.headers.get("content-type", "").startswith("application/json") else {}
        first_name = body.get("first_name", "Rosa")
        last_name = body.get("last_name", "Ayala")
        email = body.get("email", "rosa.ayala1331@gmail.com")
        subject = body.get("subject", f"Non-Pay Notice Sent — Test | BCI CRM Automation")

        from datetime import datetime
        note_data = {
            "subject": subject,
            "insured_email": email,
            "insured_first_name": first_name,
            "insured_last_name": last_name,
            "type": "Email",
            "creator_name": "BCI Non-Pay System",
            "create_date": datetime.now().strftime("%m/%d/%Y %I:%M %p"),
        }

        # Try different "type" values to find which goes to Notes tab vs Activity
        # Manual notes show "Origin: Manual" - API returns origin:3
        # Type "Email" might route to Activity log
        import requests as req
        token = nc._authenticate()
        
        # Quick search for database ID first
        found_db_id = None
        try:
            search_results = nc.search_insureds(email or last_name, limit=3)
            if search_results:
                found_db_id = str(search_results[0].get("database_id") or "")
        except:
            pass
        
        type_results = {}
        for note_type in [None, "General", "Manual", "Note", ""]:
            test_payload = {
                "subject": f"TYPE TEST ({note_type}) — {subject[:50]}",
                "insured_database_id": found_db_id or "",
                "insured_email": email,
                "insured_first_name": first_name,
                "insured_last_name": last_name,
                "insured_commercial_name": f"{first_name} {last_name}",
                "creator_name": "BCI Non-Pay System",
            }
            if note_type is not None:
                test_payload["type"] = note_type
            
            resp = req.post(
                f"{nc.base_url}/api/Zapier/InsertNote",
                headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
                json=test_payload,
                timeout=30,
            )
            result_data = {}
            try:
                result_data = resp.json()
            except:
                result_data = {"raw": resp.text[:200]}
            
            origin = result_data.get("data", {}).get("origin") if isinstance(result_data.get("data"), dict) else None
            rtype = result_data.get("data", {}).get("type") if isinstance(result_data.get("data"), dict) else None
            type_results[f"type_{note_type}"] = {
                "status": resp.status_code,
                "origin": origin,
                "returned_type": rtype,
            }
        
        # Also try with is_sticky_note and origin field explicitly
        for origin_val in [0, 1, 2]:
            test_payload = {
                "subject": f"ORIGIN TEST ({origin_val}) — {subject[:50]}",
                "insured_database_id": found_db_id or "",
                "insured_email": email,
                "insured_first_name": first_name,
                "insured_last_name": last_name,
                "insured_commercial_name": f"{first_name} {last_name}",
                "creator_name": "BCI Non-Pay System",
                "origin": origin_val,
            }
            resp = req.post(
                f"{nc.base_url}/api/Zapier/InsertNote",
                headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
                json=test_payload,
                timeout=30,
            )
            result_data = {}
            try:
                result_data = resp.json()
            except:
                result_data = {"raw": resp.text[:200]}
            
            returned_origin = result_data.get("data", {}).get("origin") if isinstance(result_data.get("data"), dict) else None
            type_results[f"origin_{origin_val}"] = {
                "status": resp.status_code,
                "origin": returned_origin,
            }
        
        # Try multiple search strategies
        search_result = {}
        insured_db_id = None
        insured_numeric_id = None
        customer_numeric_id = None
        
        for search_query in [email, f"{first_name} {last_name}", last_name]:
            if not search_query:
                continue
            try:
                results = nc.search_insureds(search_query, limit=5)
                search_result[f"query_{search_query}"] = [
                    {
                        "id": r.get("database_id"),
                        "insured_id": r.get("insured_id"),
                        "customer_id": r.get("customer_id"),
                        "name": r.get("commercial_name"),
                        "email": r.get("email"),
                    }
                    for r in results
                ]
                if results and not insured_db_id:
                    for r in results:
                        r_email = (r.get("email") or "").lower()
                        if email and r_email == email.lower():
                            insured_db_id = r.get("database_id")
                            insured_numeric_id = r.get("insured_id")
                            customer_numeric_id = r.get("customer_id")
                            break
                    if not insured_db_id:
                        insured_db_id = results[0].get("database_id")
                        insured_numeric_id = results[0].get("insured_id")
                        customer_numeric_id = results[0].get("customer_id")
            except Exception as e:
                search_result[f"query_{search_query}"] = str(e)

        # If we found the database ID, store it for later
        found_db_id = str(insured_db_id) if insured_db_id else None

        # Also do a raw POST to see exactly what NowCerts returns
        raw_payload = {
            "subject": subject,
            "insured_database_id": found_db_id or "",
            "insured_email": email,
            "insured_first_name": first_name,
            "insured_last_name": last_name,
            "insured_commercial_name": f"{first_name} {last_name}",
            "creator_name": "BCI Non-Pay System",
            "type": "Email",
        }
        raw_resp = req.post(
            f"{nc.base_url}/api/Zapier/InsertNotesForSameInsured",
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
            json=raw_payload,
            timeout=30,
        )

        # Try with just "databaseId" instead of "insuredDatabaseId"
        alt_payload = dict(raw_payload)
        if found_db_id:
            alt_payload.pop("insuredDatabaseId", None)
            alt_payload["databaseId"] = found_db_id
        raw_resp_alt = req.post(
            f"{nc.base_url}/api/Zapier/InsertNote",
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
            json=alt_payload,
            timeout=30,
        )

        # Try with insuredId (numeric) via the Insured/InsuredNotes approach
        # Also try InsertNote with just databaseId and subject, minimal payload
        minimal_payload = {"subject": subject, "databaseId": found_db_id} if found_db_id else {}
        raw_resp_min = req.post(
            f"{nc.base_url}/api/Zapier/InsertNote",
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
            json=minimal_payload,
            timeout=30,
        ) if found_db_id else None

        # Try with insuredId (numeric) via InsertNote
        numeric_payload = dict(raw_payload)
        if insured_numeric_id:
            numeric_payload["insuredId"] = str(insured_numeric_id)
        # Also try commercial name in "Last, First" format (how NowCerts stores it)
        numeric_payload["insuredCommercialName"] = f"{last_name}, {first_name}"
        raw_resp_numeric = req.post(
            f"{nc.base_url}/api/Zapier/InsertNote",
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
            json=numeric_payload,
            timeout=30,
        )

        # Also try original endpoint for comparison
        raw_resp2 = req.post(
            f"{nc.base_url}/api/Zapier/InsertNote",
            headers={"Authorization": f"Bearer {token}", "Content-Type": "application/json"},
            json=raw_payload,
            timeout=30,
        )

        result = nc.insert_note(note_data)
        return {
            "success": True,
            "note_data_sent": note_data,
            "normalized_payload": raw_payload,
            "nowcerts_response": result,
            "search_result": {
                "insured_db_id": insured_db_id,
                "searches": search_result,
            },
            "type_origin_tests": type_results,
            "raw_debug": {
                "InsertNotesForSameInsured_withInsuredDbId": {
                    "status_code": raw_resp.status_code,
                    "body": raw_resp.text[:500],
                },
                "InsertNote_withDatabaseId": {
                    "status_code": raw_resp_alt.status_code,
                    "body": raw_resp_alt.text[:500],
                    "payload_sent": alt_payload,
                },
                "InsertNote_minimal": {
                    "status_code": raw_resp_min.status_code if raw_resp_min else "skipped",
                    "body": raw_resp_min.text[:500] if raw_resp_min else "skipped",
                    "payload_sent": minimal_payload,
                },
                "InsertNote_withInsuredId_and_CommName": {
                    "status_code": raw_resp_numeric.status_code,
                    "body": raw_resp_numeric.text[:500],
                    "payload_sent": {k: v for k, v in numeric_payload.items() if k != "subject"},
                },
                "InsertNote_withInsuredDbId": {
                    "status_code": raw_resp2.status_code,
                    "body": raw_resp2.text[:500],
                },
            },
        }
    except Exception as e:
        import traceback
        return {"error": str(e), "traceback": traceback.format_exc()}


@router.post("/backfill-nowcerts-notes")
async def backfill_nowcerts_notes():
    """One-time backfill: add NowCerts notes for all previously sent non-pay emails."""
    from app.core.database import SessionLocal
    from app.models.nonpay import NonPayNotice, NonPayEmail
    from app.services.nowcerts import get_nowcerts_client
    from datetime import datetime

    nc = get_nowcerts_client()
    if not nc.is_configured:
        return {"error": "NowCerts not configured"}

    db = SessionLocal()
    try:
        # Get all results where email was sent successfully
        results = db.query(NonPayEmail).filter(
            NonPayEmail.email_status.in_(["sent", "letter_sent"]),
        ).all()

        notes_added = 0
        errors = []
        skipped = 0

        for r in results:
            try:
                # Build note
                amt_str = f"${r.amount_due:,.2f}" if r.amount_due else "N/A"
                due_str = r.due_date or "N/A"
                carrier = (r.carrier or "unknown").replace("_", " ").title()

                send_method = "Letter (Thanks.io)" if r.email_status == "letter_sent" else "Email"

                note_body = (
                    f"Policy: {r.policy_number}\n"
                    f"Carrier: {carrier}\n"
                    f"Amount Due: {amt_str}\n"
                    f"Due Date: {due_str}\n"
                    f"Method: {send_method}\n"
                    f"Sent via BCI CRM Non-Pay Automation"
                )

                subject = f"Non-Pay Notice Sent — {r.policy_number} | {note_body}"

                parts = (r.customer_name or "").strip().split()
                first_name = parts[0] if parts else ""
                last_name = parts[-1] if len(parts) > 1 else ""

                # Use the original send date if available
                created = r.sent_at.strftime("%m/%d/%Y %I:%M %p") if r.sent_at else datetime.now().strftime("%m/%d/%Y %I:%M %p")

                note_data = {
                    "subject": subject,
                    "insured_email": r.customer_email,
                    "insured_first_name": first_name,
                    "insured_last_name": last_name,
                    "type": "Email",
                    "creator_name": "BCI Non-Pay System",
                    "create_date": created,
                }

                result = nc.insert_note(note_data)
                if result and result.get("status") == 1:
                    notes_added += 1
                else:
                    errors.append({"policy": r.policy_number, "name": r.customer_name, "response": result})
            except Exception as e:
                errors.append({"policy": r.policy_number, "name": r.customer_name, "error": str(e)})

        return {
            "total_emails_found": len(results),
            "notes_added": notes_added,
            "errors": len(errors),
            "error_details": errors[:20],
        }
    finally:
        db.close()


@router.get("/sent-list")
def list_sent_nonpay_emails():
    """List all sent non-pay emails (no auth required, read-only)."""
    from app.core.database import SessionLocal
    from app.models.nonpay import NonPayEmail

    db = SessionLocal()
    try:
        results = db.query(NonPayEmail).filter(
            NonPayEmail.email_status == "sent",
        ).order_by(NonPayEmail.sent_at.desc()).all()

        return [
            {
                "customer_name": r.customer_name,
                "customer_email": r.customer_email,
                "policy_number": r.policy_number,
                "carrier": r.carrier,
                "amount_due": f"${r.amount_due:,.2f}" if r.amount_due else "N/A",
                "due_date": r.due_date,
                "sent_at": r.sent_at.isoformat() if r.sent_at else None,
            }
            for r in results
        ]
    finally:
        db.close()


@router.get("/letter-list")
def list_sent_nonpay_letters():
    """List all sent non-pay letters via Thanks.io."""
    from app.core.database import SessionLocal
    from app.models.nonpay import NonPayEmail

    db = SessionLocal()
    try:
        results = db.query(NonPayEmail).filter(
            NonPayEmail.email_status.in_(["letter_sent", "letter_failed"]),
        ).order_by(NonPayEmail.sent_at.desc()).all()

        return [
            {
                "customer_name": r.customer_name,
                "customer_email": r.customer_email or "(no email)",
                "policy_number": r.policy_number,
                "carrier": r.carrier,
                "amount_due": f"${r.amount_due:,.2f}" if r.amount_due else "N/A",
                "due_date": r.due_date,
                "status": r.email_status,
                "sent_at": r.sent_at.isoformat() if r.sent_at else None,
            }
            for r in results
        ]
    finally:
        db.close()


# ── Grange Non-Renewal Handler ───────────────────────────────────────

async def _handle_grange_nonrenewal(html_body: str, sender: str, db: Session) -> dict:
    """Handle GrangeWire Non-Renewal Alerts.
    
    These emails have a table with: Policy Symbol | Number | Insured
    We parse the table, look up each customer in NowCerts, and trigger
    the non-renewal notification workflow.
    """
    from app.api.non_renewal import NonRenewalCreate, _send_nonrenewal_email
    
    # Parse the HTML table — reuse existing parser but handle the simpler format
    policies = _parse_grange_nonrenewal_table(html_body)
    
    if not policies:
        logger.warning("GrangeWire non-renewal: no policies extracted")
        return {"status": "ok", "message": "GrangeWire non-renewal parsed but no policies found"}
    
    logger.info(f"GrangeWire non-renewal: found {len(policies)} policies")
    
    # Create a notice record for tracking
    notice = NonPayNotice(
        filename=f"grange-nonrenewal-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}",
        upload_type="inbound-email",
        uploaded_by=f"inbound:{sender[:60]}",
        policies_found=len(policies),
        status="processing",
    )
    db.add(notice)
    db.commit()
    db.refresh(notice)
    
    results = []
    emails_sent = 0
    
    for pol in policies:
        policy_number = pol.get("policy_number", "").strip()
        insured_name = pol.get("insured_name", "").strip()
        
        if not policy_number:
            continue
        
        # Look up customer in NowCerts
        customer_email = None
        customer_phone = None
        try:
            from app.services.nowcerts import get_nowcerts_client
            nc = get_nowcerts_client()
            if nc.is_configured:
                customer_data = nc.search_insured(policy_number)
                if customer_data:
                    customer_email = customer_data.get("email") or customer_data.get("commercial_email")
                    customer_phone = customer_data.get("phone") or customer_data.get("home_phone")
                    if not insured_name and customer_data.get("insured_name"):
                        insured_name = customer_data["insured_name"]
        except Exception as e:
            logger.error(f"NowCerts lookup failed for {policy_number}: {e}")
        
        # Send non-renewal notification
        nr_data = NonRenewalCreate(
            customer_name=insured_name or "Valued Customer",
            customer_email=customer_email,
            customer_phone=customer_phone,
            policy_number=policy_number,
            carrier="grange",
            reason="Carrier non-renewal",
            send_notification=bool(customer_email),
        )
        
        email_sent = False
        if customer_email:
            email_sent = _send_nonrenewal_email(nr_data)
            if email_sent:
                emails_sent += 1
        
        # Add NowCerts note regardless
        try:
            from app.services.nowcerts import get_nowcerts_client
            nc = get_nowcerts_client()
            if nc.is_configured:
                parts = insured_name.strip().split(maxsplit=1) if insured_name else [""]
                note_body = (
                    f"Non-Renewal Notice — Grange Insurance | "
                    f"Policy: {policy_number} | "
                    f"Source: GrangeWire Alert (auto-parsed) | "
                    f"Customer notified via email: {'Yes' if email_sent else 'No — no email on file'}"
                )
                nc.insert_note({
                    "subject": note_body,
                    "insured_email": customer_email or "",
                    "insured_first_name": parts[0] if parts else "",
                    "insured_last_name": parts[1] if len(parts) > 1 else "",
                    "type": "Email",
                    "creator_name": "BCI Non-Renewal System",
                    "create_date": datetime.now().strftime("%m/%d/%Y %I:%M %p"),
                })
        except Exception as e:
            logger.error(f"NowCerts note failed for Grange non-renewal {policy_number}: {e}")
        
        results.append({
            "policy_number": policy_number,
            "insured_name": insured_name,
            "email_sent": email_sent,
            "customer_email": customer_email or "not found",
        })
    
    # Update notice
    notice.emails_sent = emails_sent
    notice.policies_matched = len([r for r in results if r["customer_email"] != "not found"])
    notice.status = "complete"
    db.commit()
    
    logger.info(f"Grange non-renewal complete: {len(results)} processed, {emails_sent} emails sent")
    return {
        "status": "ok",
        "carrier": "grange",
        "notice_type": "non_renewal",
        "policies_found": len(policies),
        "emails_sent": emails_sent,
        "results": results,
    }


def _parse_grange_nonrenewal_table(html_body: str) -> list[dict]:
    """Parse GrangeWire Non-Renewal Alerts table.
    
    Format: Policy Symbol | Number | Insured
    Policy number = Symbol + Number combined (e.g., "PHA" + "5168274 02")
    """
    from html.parser import HTMLParser

    class TableParser(HTMLParser):
        def __init__(self):
            super().__init__()
            self.in_table = False
            self.in_row = False
            self.in_cell = False
            self.current_row = []
            self.current_cell = ""
            self.rows = []

        def handle_starttag(self, tag, attrs):
            if tag == "table":
                self.in_table = True
            elif tag == "tr" and self.in_table:
                self.in_row = True
                self.current_row = []
            elif tag in ("td", "th") and self.in_row:
                self.in_cell = True
                self.current_cell = ""

        def handle_endtag(self, tag):
            if tag == "table":
                self.in_table = False
            elif tag == "tr" and self.in_row:
                self.in_row = False
                if self.current_row:
                    self.rows.append(self.current_row)
            elif tag in ("td", "th") and self.in_cell:
                self.in_cell = False
                self.current_row.append(self.current_cell.strip())

        def handle_data(self, data):
            if self.in_cell:
                self.current_cell += data

    parser = TableParser()
    parser.feed(html_body or "")

    if not parser.rows:
        return []

    # Find header row — look for "Symbol" or "Number" or "Insured"
    header_idx = None
    for i, row in enumerate(parser.rows):
        row_upper = " ".join(row).upper()
        if "SYMBOL" in row_upper or ("NUMBER" in row_upper and "INSURED" in row_upper):
            header_idx = i
            break
        # Also try: "Policy Symbol" header format
        if "POLICY" in row_upper and len(row) == 3:
            header_idx = i
            break

    if header_idx is None:
        # Fallback: if we have 3-column rows, assume first is header
        for i, row in enumerate(parser.rows):
            if len(row) == 3:
                header_idx = i
                break

    if header_idx is None:
        return []

    headers = [h.upper().strip() for h in parser.rows[header_idx]]

    # Map columns
    symbol_idx = 0
    number_idx = 1
    insured_idx = 2

    for i, h in enumerate(headers):
        if "SYMBOL" in h:
            symbol_idx = i
        elif "NUMBER" in h and "INSURED" not in h:
            number_idx = i
        elif "INSURED" in h or "NAME" in h:
            insured_idx = i

    results = []
    for row in parser.rows[header_idx + 1:]:
        if len(row) < 3:
            continue

        symbol = row[symbol_idx].strip() if symbol_idx < len(row) else ""
        number = row[number_idx].strip() if number_idx < len(row) else ""
        insured = row[insured_idx].strip() if insured_idx < len(row) else ""

        if not number and not insured:
            continue

        # Combine symbol + number for full policy number
        full_policy = f"{symbol}{number}".strip() if symbol else number.strip()

        results.append({
            "policy_number": full_policy,
            "insured_name": insured,
            "policy_symbol": symbol,
        })

    return results


# ── National General Policy Activity Smart Router ─────────────────────

async def _handle_natgen_policy_activity(html_body: str, sender: str, db: Session) -> dict:
    """Parse NatGen Policy Activity email and route each section appropriately.
    
    - Outstanding To Dos:
        - GoPaperless → skip
        - Proof of Continuous Insurance / NoPOP / ChangePriorBI / Proof of Prior BI → send UW requirement email
    - Pending Non-Renewals:
        - Send non-renewal notice to insured + CC producer for remarketing
    - Pending Cancellations:
        - Non Payment / NSF → existing non-pay flow
        - Underwriting / Voluntary → skip
    - Undeliverable Mail → notify producer
    """
    from app.services.natgen_parser import parse_natgen_policy_activity
    from app.services.uw_requirement_email import (
        send_uw_requirement_email, send_undeliverable_mail_alert,
        send_non_renewal_email,
    )

    parsed = parse_natgen_policy_activity(html_body)
    carrier = "national_general"

    # Dry-run mode: parse everything, create tasks, but send NO emails
    natgen_dry_run = os.environ.get("NATGEN_DRY_RUN", "true").lower() == "true"
    if natgen_dry_run:
        logger.info("NATGEN_DRY_RUN=true — will parse and create tasks but NOT send any emails")
    
    results = {
        "status": "processed",
        "dry_run": natgen_dry_run,
        "carrier": "national_general",
        "sections": {},
    }

    # ── 1. Outstanding To Dos ──
    todos_processed = []
    for todo in parsed["outstanding_todos"]:
        todo_type = todo.get("todo_type", "other")
        policy = todo.get("policy", "")
        insured = todo.get("insured_name", "")
        
        if todo_type == "go_paperless":
            todos_processed.append({
                "policy": policy, "insured": insured,
                "action": "skipped", "reason": "GoPaperless — not actionable"
            })
            continue
        
        if todo_type in ("proof_of_continuous_insurance", "nopop", "change_prior_bi", "proof_of_prior_bi"):
            # Look up customer in DB to get email + producer
            customer_info = _lookup_customer_for_natgen(db, policy, insured)
            
            # Create UW requirement task (always, even in dry run)
            try:
                from app.api.tasks import create_uw_requirement_task
                producer_id = None
                if customer_info and customer_info.get("sale_id"):
                    from app.models.sale import Sale as SaleModel
                    sale_rec = db.query(SaleModel).filter(SaleModel.id == customer_info["sale_id"]).first()
                    if sale_rec:
                        producer_id = sale_rec.producer_id
                uw_task = create_uw_requirement_task(
                    db=db,
                    customer_name=insured,
                    policy_number=policy,
                    carrier=carrier,
                    requirement_type=todo_type,
                    due_date=todo.get("next_action_date"),
                    producer_name=customer_info.get("producer_name") if customer_info else None,
                    assigned_to_id=producer_id,
                )
            except Exception as e:
                logger.error("Failed to create UW task for %s: %s", policy, e)

            if customer_info and customer_info.get("email"):
                # Determine requirement type
                req_type = "nopop" if todo_type == "nopop" else "proof_of_continuous_insurance"
                
                if natgen_dry_run:
                    email_result = {"success": False, "error": "DRY RUN — email not sent"}
                else:
                    email_result = send_uw_requirement_email(
                        to_email=customer_info["email"],
                        client_name=insured,
                        policy_number=policy,
                        carrier=carrier,
                        requirement_type=req_type,
                        due_date=todo.get("next_action_date"),
                        producer_name=customer_info.get("producer_name"),
                        producer_email=customer_info.get("producer_email"),
                    )
                todos_processed.append({
                    "policy": policy, "insured": insured,
                    "action": ("dry_run_uw_email" if natgen_dry_run else "uw_email_sent") if email_result.get("success") or natgen_dry_run else "uw_email_failed",
                    "would_email": customer_info["email"],
                    "requirement": req_type,
                    "error": email_result.get("error"),
                })
            else:
                todos_processed.append({
                    "policy": policy, "insured": insured,
                    "action": "no_email_found", "requirement": todo_type,
                })
            continue
        
        # Other todo types — log but skip
        todos_processed.append({
            "policy": policy, "insured": insured,
            "action": "skipped", "reason": f"Unknown todo type: {todo.get('todo_description', '')}"
        })
    
    results["sections"]["outstanding_todos"] = {
        "total": len(parsed["outstanding_todos"]),
        "details": todos_processed,
    }

    # ── 2. Pending Non-Renewals → notify insured + producer for remarketing ──
    nonrenewals_processed = []
    for nr in parsed.get("pending_non_renewals", []):
        policy = nr.get("policy", "")
        insured = nr.get("insured_name", "")
        effective = nr.get("effective_date", "")
        premium = nr.get("premium")
        product = nr.get("product", "")
        description = nr.get("description", "")
        nr_producer = nr.get("producer_name", "")

        # Check if within 60 days
        is_urgent = _is_within_days_nonpay(effective, 60)

        # Look up customer
        customer_info = _lookup_customer_for_natgen(db, policy, insured)

        # Always create a task for the retention team
        try:
            from app.api.tasks import create_non_renewal_task
            # Look up producer_id from the sale
            producer_id = None
            if customer_info and customer_info.get("sale_id"):
                from app.models.sale import Sale as SaleModel
                sale_record = db.query(SaleModel).filter(SaleModel.id == customer_info["sale_id"]).first()
                if sale_record:
                    producer_id = sale_record.producer_id
            task = create_non_renewal_task(
                db=db,
                customer_name=insured,
                policy_number=policy,
                carrier=carrier,
                effective_date=effective,
                premium=premium,
                producer_name=nr_producer,
                assigned_to_id=producer_id,
            )
            task_id = task.id
        except Exception as e:
            logger.error("Failed to create non-renewal task for %s: %s", policy, e)
            task_id = None

        if is_urgent:
            # Within 60 days → send to service@betterchoiceins.com (retention team)
            # AND send non-renewal email to insured if we have their email
            to_emails = ["service@betterchoiceins.com"]
            
            if natgen_dry_run:
                email_result = {"success": False, "error": "DRY RUN — email not sent"}
            elif customer_info and customer_info.get("email"):
                # Send customer-facing non-renewal email
                email_result = send_non_renewal_email(
                    to_email=customer_info["email"],
                    client_name=insured,
                    policy_number=policy,
                    carrier=carrier,
                    effective_date=effective,
                    premium=premium,
                    product=product,
                    description=description,
                    producer_name=nr_producer or (customer_info or {}).get("producer_name"),
                    producer_email=(customer_info or {}).get("producer_email"),
                )
            else:
                email_result = {"success": False, "error": "No customer email"}

            # Send internal alert to service@ (also gated by dry run)
            if not natgen_dry_run:
                _send_internal_nonrenewal_alert(
                insured_name=insured,
                policy_number=policy,
                carrier=carrier,
                effective_date=effective,
                premium=premium,
                product=product,
                producer_name=nr_producer,
                phone=nr.get("phone", ""),
                customer_email=(customer_info or {}).get("email", ""),
            )

            nonrenewals_processed.append({
                "policy": policy, "insured": insured,
                "action": "urgent_non_renewal" if email_result.get("success") else "urgent_no_customer_email",
                "email": (customer_info or {}).get("email", ""),
                "effective_date": effective,
                "producer": nr_producer,
                "task_id": task_id,
                "days_until": _days_until(effective),
                "error": email_result.get("error"),
            })
        else:
            # More than 60 days — still send non-renewal email to insured
            if customer_info and customer_info.get("email"):
                if natgen_dry_run:
                    email_result = {"success": False, "error": "DRY RUN — email not sent"}
                else:
                    email_result = send_non_renewal_email(
                        to_email=customer_info["email"],
                        client_name=insured,
                        policy_number=policy,
                        carrier=carrier,
                        effective_date=effective,
                        premium=premium,
                        product=product,
                        description=description,
                        producer_name=nr_producer or (customer_info or {}).get("producer_name"),
                        producer_email=(customer_info or {}).get("producer_email"),
                    )
                nonrenewals_processed.append({
                    "policy": policy, "insured": insured,
                    "action": "dry_run_non_renewal" if natgen_dry_run else ("non_renewal_email_sent" if email_result.get("success") else "non_renewal_email_failed"),
                    "would_email": customer_info["email"],
                    "effective_date": effective,
                    "producer": nr_producer,
                    "task_id": task_id,
                    "error": email_result.get("error"),
                })
            else:
                nonrenewals_processed.append({
                    "policy": policy, "insured": insured,
                    "action": "no_email_found",
                    "effective_date": effective,
                    "producer": nr_producer,
                    "task_id": task_id,
                })

    results["sections"]["pending_non_renewals"] = {
        "total": len(parsed.get("pending_non_renewals", [])),
        "details": nonrenewals_processed,
    }

    # ── 3. Pending Cancellations → existing non-pay flow ──
    cancellations_processed = []
    mode = INBOUND_NONPAY_MODE
    dry_run = (mode != "live")
    
    # Create a notice record for the non-pay portion
    nonpay_policies = [c for c in parsed["pending_cancellations"] if c["cancel_type"] in ("non_pay", "nsf")]
    if nonpay_policies:
        notice = NonPayNotice(
            filename=f"natgen-activity-{datetime.utcnow().strftime('%Y%m%d-%H%M%S')}",
            upload_type="inbound-email",
            uploaded_by=f"inbound:{sender[:60]}",
            policies_found=len(nonpay_policies),
            status="processing",
        )
        db.add(notice)
        db.commit()
        db.refresh(notice)
    
    for canc in parsed["pending_cancellations"]:
        policy = canc.get("policy", "")
        insured = canc.get("insured_name", "")
        cancel_type = canc.get("cancel_type", "other")
        
        if cancel_type in ("underwriting", "voluntary", "other"):
            cancellations_processed.append({
                "policy": policy, "insured": insured,
                "action": "skipped", "reason": canc.get("reason", cancel_type),
            })
            continue
        
        # Non-pay / NSF → process through existing flow
        if nonpay_policies:
            result = _process_single_policy(
                db=db,
                notice_id=notice.id,
                policy_number=policy,
                carrier=carrier,
                insured_name=insured,
                amount_due=canc.get("amount_due"),
                due_date=canc.get("cancel_date"),
                cancel_date=canc.get("cancel_date"),
                dry_run=dry_run,
            )
            result["cancel_type"] = cancel_type
            result["reason"] = canc.get("reason", "")
            cancellations_processed.append(result)
    
    # Update notice
    if nonpay_policies:
        sent = sum(1 for r in cancellations_processed if r.get("email_sent"))
        notice.emails_sent = sent
        notice.status = "complete"
        db.commit()
    
    results["sections"]["pending_cancellations"] = {
        "total": len(parsed["pending_cancellations"]),
        "nonpay": len(nonpay_policies),
        "skipped": len(parsed["pending_cancellations"]) - len(nonpay_policies),
        "details": cancellations_processed,
    }

    # ── 4. Undeliverable Mail → notify producer ──
    undeliverable_processed = []
    for undel in parsed["undeliverable_mail"]:
        policy = undel.get("policy", "")
        insured = undel.get("insured_name", "")
        phone = undel.get("phone", "")
        mail_desc = undel.get("mail_description", "")
        
        # Look up customer to find their producer
        customer_info = _lookup_customer_for_natgen(db, policy, insured)
        producer_email = (customer_info or {}).get("producer_email", "evan@betterchoiceins.com")
        
        if natgen_dry_run:
            alert_result = {"success": False, "error": "DRY RUN — alert not sent"}
        else:
            alert_result = send_undeliverable_mail_alert(
                producer_email=producer_email,
                client_name=insured,
                policy_number=policy,
                carrier=carrier,
                mail_description=mail_desc,
                phone=phone,
            )
        undeliverable_processed.append({
            "policy": policy, "insured": insured,
            "action": "dry_run_alert" if natgen_dry_run else ("alert_sent" if alert_result.get("success") else "alert_failed"),
            "producer": producer_email,
            "error": alert_result.get("error"),
        })
    
    results["sections"]["undeliverable_mail"] = {
        "total": len(parsed["undeliverable_mail"]),
        "details": undeliverable_processed,
    }

    logger.info(
        "NatGen Activity processed: %d todos, %d non-renewals, %d cancellations (%d nonpay), %d undeliverable",
        len(parsed["outstanding_todos"]),
        len(parsed.get("pending_non_renewals", [])),
        len(parsed["pending_cancellations"]),
        len(nonpay_policies) if nonpay_policies else 0,
        len(parsed["undeliverable_mail"]),
    )

    # ── 5. Run escalation check on ALL existing non-renewal tasks ──
    if natgen_dry_run:
        results["escalation_check"] = {"skipped": "DRY RUN — escalation emails not sent"}
    else:
        try:
            from app.services.nonrenewal_escalation import run_escalation_check
            escalation_result = run_escalation_check(db)
            results["escalation_check"] = {
                "checked": escalation_result["checked"],
                "escalated": escalation_result["escalated"],
            }
        except Exception as e:
            logger.error("Escalation check failed: %s", e)
            results["escalation_check"] = {"error": str(e)}

    # ── 6. Run compliance reminders for inspections + UW tasks ──
    if natgen_dry_run:
        results["compliance_reminders"] = {"skipped": "DRY RUN — compliance reminders not sent"}
    else:
        try:
            from app.services.compliance_reminders import run_compliance_reminders
            reminder_result = run_compliance_reminders(db)
            results["compliance_reminders"] = {
                "checked": reminder_result["checked"],
                "sent": reminder_result["reminders_sent"],
                "skipped": reminder_result["skipped"],
            }
        except Exception as e:
            logger.error("Compliance reminders failed: %s", e)
            results["compliance_reminders"] = {"error": str(e)}

    # Log full dry-run results for review
    import json
    logger.info("=== NATGEN HANDLER RESULTS ===\n%s", json.dumps(results, indent=2, default=str))

    return results


def _lookup_customer_for_natgen(db: Session, policy_number: str, insured_name: str) -> Optional[dict]:
    """Look up a customer by policy number or name to get email + producer info."""
    from app.models.sale import Sale
    from app.models.user import User

    if not policy_number:
        return None

    # Clean policy number (NatGen format: "2033776342" or with spaces)
    clean_policy = policy_number.replace(" ", "").strip()

    # Try exact match first
    sale = db.query(Sale).filter(
        Sale.policy_number.ilike(f"%{clean_policy}%")
    ).first()

    # Try without trailing " 00" or " 01" suffix
    if not sale and len(clean_policy) > 2:
        base = clean_policy[:-2] if clean_policy[-2:] in ("00", "01") else clean_policy
        sale = db.query(Sale).filter(
            Sale.policy_number.ilike(f"%{base}%")
        ).first()

    # Try name match as fallback
    if not sale and insured_name:
        parts = insured_name.strip().split()
        if len(parts) >= 2:
            first = parts[0]
            last = parts[-1]
            sale = db.query(Sale).filter(
                Sale.client_name.ilike(f"%{first}%"),
                Sale.client_name.ilike(f"%{last}%"),
            ).first()

    if not sale:
        # Fallback: check NowCerts customers table
        from app.models.customer import Customer, CustomerPolicy
        nc_policy = db.query(CustomerPolicy).filter(
            CustomerPolicy.policy_number.ilike(f"%{clean_policy}%")
        ).first()
        if not nc_policy and len(clean_policy) > 2:
            base = clean_policy[:-2] if clean_policy[-2:] in ("00", "01") else clean_policy
            nc_policy = db.query(CustomerPolicy).filter(
                CustomerPolicy.policy_number.ilike(f"%{base}%")
            ).first()
        if nc_policy:
            customer = db.query(Customer).filter(Customer.id == nc_policy.customer_id).first()
            if customer and customer.email:
                return {
                    "email": customer.email,
                    "phone": customer.phone or customer.mobile_phone,
                    "sale_id": None,
                    "source": "nowcerts",
                    "producer_name": nc_policy.agent_name,
                    "producer_email": None,
                }
        return None

    result = {
        "email": sale.client_email,
        "phone": getattr(sale, "client_phone", None),
        "sale_id": sale.id,
    }

    # If sale has no email, try to enrich from NowCerts customers table
    if not result["email"]:
        from app.models.customer import Customer, CustomerPolicy
        nc_policy = db.query(CustomerPolicy).filter(
            CustomerPolicy.policy_number.ilike(f"%{clean_policy}%")
        ).first()
        if not nc_policy and len(clean_policy) > 2:
            base2 = clean_policy[:-2] if clean_policy[-2:] in ("00", "01") else clean_policy
            nc_policy = db.query(CustomerPolicy).filter(
                CustomerPolicy.policy_number.ilike(f"%{base2}%")
            ).first()
        if nc_policy:
            customer = db.query(Customer).filter(Customer.id == nc_policy.customer_id).first()
            if customer and customer.email:
                result["email"] = customer.email
                result["phone"] = result["phone"] or customer.phone or customer.mobile_phone

    # Get producer info
    if sale.producer_id:
        producer = db.query(User).filter(User.id == sale.producer_id).first()
        if producer:
            result["producer_name"] = producer.full_name or producer.username
            result["producer_email"] = producer.email

    return result


def _is_within_days_nonpay(date_str: str, days: int) -> bool:
    """Check if a date string is within N days from now."""
    try:
        from dateutil import parser as dateparser
        dt = dateparser.parse(date_str)
        delta = (dt - datetime.now()).days
        return delta <= days
    except Exception:
        # If we can't parse the date, treat as urgent
        return True


def _days_until(date_str: str) -> Optional[int]:
    """Return number of days until a date."""
    try:
        from dateutil import parser as dateparser
        dt = dateparser.parse(date_str)
        return (dt - datetime.now()).days
    except Exception:
        return None


def _send_internal_nonrenewal_alert(
    insured_name: str,
    policy_number: str,
    carrier: str,
    effective_date: str,
    premium: Optional[float] = None,
    product: str = "",
    producer_name: str = "",
    phone: str = "",
    customer_email: str = "",
):
    """Send internal alert to service@betterchoiceins.com for urgent non-renewals."""
    import requests as req_lib

    if not settings.MAILGUN_API_KEY or not settings.MAILGUN_DOMAIN:
        logger.warning("Mailgun not configured — skipping internal non-renewal alert")
        return

    days = _days_until(effective_date)
    days_str = f"{days} days" if days is not None else "UNKNOWN"
    premium_str = f"${premium:,.2f}" if premium else "N/A"

    subject = f"🚨 URGENT Non-Renewal: {insured_name} — {days_str} until coverage ends"

    html = f"""<!DOCTYPE html><html><head><meta charset="utf-8"></head>
<body style="margin:0;padding:0;background:#f1f5f9;font-family:-apple-system,BlinkMacSystemFont,Segoe UI,Roboto,sans-serif;">
<div style="max-width:600px;margin:0 auto;padding:20px;">
<div style="background:linear-gradient(135deg,#dc2626,#991b1b);border-radius:16px 16px 0 0;padding:24px 32px;text-align:center;">
<h1 style="margin:0;font-size:20px;color:#fff;">🚨 Non-Renewal — Immediate Action Required</h1>
<p style="margin:8px 0 0;font-size:14px;color:rgba(255,255,255,0.9);">Coverage ends in <strong>{days_str}</strong></p>
</div>
<div style="background:#fff;padding:28px 32px;border-radius:0 0 16px 16px;">
<table style="width:100%;font-size:14px;color:#334155;" cellpadding="0" cellspacing="0">
<tr><td style="padding:8px 0;color:#64748b;width:140px;">Customer</td><td style="font-weight:700;font-size:15px;">{insured_name}</td></tr>
<tr><td style="padding:8px 0;color:#64748b;">Policy</td><td style="font-weight:600;">{policy_number}</td></tr>
<tr><td style="padding:8px 0;color:#64748b;">Carrier</td><td>National General</td></tr>
<tr><td style="padding:8px 0;color:#64748b;">Coverage Ends</td><td style="font-weight:700;color:#dc2626;">{effective_date}</td></tr>
<tr><td style="padding:8px 0;color:#64748b;">Premium</td><td>{premium_str}</td></tr>
<tr><td style="padding:8px 0;color:#64748b;">Product</td><td>{product}</td></tr>
<tr><td style="padding:8px 0;color:#64748b;">Producer</td><td>{producer_name}</td></tr>
<tr><td style="padding:8px 0;color:#64748b;">Phone</td><td>{phone or 'N/A'}</td></tr>
<tr><td style="padding:8px 0;color:#64748b;">Email</td><td>{customer_email or 'Not on file'}</td></tr>
</table>
<div style="margin:20px 0;padding:16px;background:#fef2f2;border-radius:10px;border:1px solid #fecaca;">
<p style="margin:0;font-size:14px;color:#991b1b;line-height:1.6;">
<strong>Action needed:</strong> Shop replacement coverage and contact the customer before {effective_date}.
A task has been created on the dashboard.</p>
</div>
</div></div></body></html>"""

    try:
        resp = req_lib.post(
            f"https://api.mailgun.net/v3/{settings.MAILGUN_DOMAIN}/messages",
            auth=("api", settings.MAILGUN_API_KEY),
            data={
                "from": f"Better Choice Insurance <service@{settings.MAILGUN_DOMAIN}>",
                "to": ["service@betterchoiceins.com"],
                "subject": subject,
                "html": html,
            },
            timeout=30,
        )
        if resp.status_code == 200:
            logger.info("Internal non-renewal alert sent for %s/%s", insured_name, policy_number)
        else:
            logger.error("Internal alert failed: %s %s", resp.status_code, resp.text[:200])
    except Exception as e:
        logger.error("Internal non-renewal alert error: %s", e)


@router.post("/test-natgen-parser")
async def test_natgen_parser(
    request: Request,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """
    Test endpoint: paste NatGen Policy Activity email HTML and see what the
    parser would do — WITHOUT sending any emails.

    POST /api/nonpay/test-natgen-parser
    Body: { "html": "<html>..." }

    Returns parsed sections + what actions WOULD be taken.
    """
    try:
        body = await request.json()
    except Exception:
        return {"error": "Send JSON with 'html' key containing the email HTML"}

    html_body = body.get("html", "")
    if not html_body:
        return {"error": "Missing 'html' field"}

    from app.services.natgen_parser import parse_natgen_policy_activity

    parsed = parse_natgen_policy_activity(html_body)

    # Build a dry-run summary of what WOULD happen
    summary = {"mode": "DRY RUN — no emails sent", "sections": {}}

    # Outstanding To Dos
    todos = []
    for todo in parsed["outstanding_todos"]:
        todo_type = todo.get("todo_type", "other")
        action = "skip (GoPaperless)" if todo_type == "go_paperless" else (
            "→ UW requirement email to insured" if todo_type in ("proof_of_continuous_insurance", "nopop", "change_prior_bi", "proof_of_prior_bi")
            else f"skip (unknown: {todo.get('todo_description', '')})"
        )
        # Try to find customer email
        customer_info = _lookup_customer_for_natgen(db, todo.get("policy", ""), todo.get("insured_name", ""))
        todos.append({
            "policy": todo.get("policy"),
            "insured": todo.get("insured_name"),
            "type": todo_type,
            "description": todo.get("todo_description"),
            "action": action,
            "customer_email_found": (customer_info or {}).get("email", "NOT FOUND"),
            "due_date": todo.get("next_action_date"),
        })
    summary["sections"]["outstanding_todos"] = {"count": len(todos), "rows": todos}

    # Pending Non-Renewals
    nonrenewals = []
    for nr in parsed.get("pending_non_renewals", []):
        effective = nr.get("effective_date", "")
        days = _days_until(effective)
        customer_info = _lookup_customer_for_natgen(db, nr.get("policy", ""), nr.get("insured_name", ""))
        is_urgent = _is_within_days_nonpay(effective, 60)

        nonrenewals.append({
            "policy": nr.get("policy"),
            "insured": nr.get("insured_name"),
            "effective_date": effective,
            "days_remaining": days,
            "premium": nr.get("premium"),
            "producer": nr.get("producer_name"),
            "within_60_days": is_urgent,
            "action": (
                "→ URGENT: internal alert to service@ + insured email + dashboard task"
                if is_urgent else
                "→ insured email + dashboard task"
            ),
            "customer_email_found": (customer_info or {}).get("email", "NOT FOUND"),
        })
    summary["sections"]["pending_non_renewals"] = {"count": len(nonrenewals), "rows": nonrenewals}

    # Pending Cancellations
    cancellations = []
    for canc in parsed["pending_cancellations"]:
        cancel_type = canc.get("cancel_type", "other")
        if cancel_type in ("non_pay", "nsf"):
            action = "→ non-pay email to insured"
        elif cancel_type == "underwriting":
            action = "skip (underwriting cancellation)"
        elif cancel_type == "voluntary":
            action = "skip (policyholder request)"
        else:
            action = f"skip ({cancel_type})"

        customer_info = _lookup_customer_for_natgen(db, canc.get("policy", ""), canc.get("insured_name", ""))
        cancellations.append({
            "policy": canc.get("policy"),
            "insured": canc.get("insured_name"),
            "reason": canc.get("reason"),
            "cancel_type": cancel_type,
            "cancel_date": canc.get("cancel_date"),
            "amount_due": canc.get("amount_due_str"),
            "action": action,
            "customer_email_found": (customer_info or {}).get("email", "NOT FOUND"),
        })
    summary["sections"]["pending_cancellations"] = {"count": len(cancellations), "rows": cancellations}

    # Undeliverable Mail
    undeliverable = []
    for undel in parsed["undeliverable_mail"]:
        customer_info = _lookup_customer_for_natgen(db, undel.get("policy", ""), undel.get("insured_name", ""))
        undeliverable.append({
            "policy": undel.get("policy"),
            "insured": undel.get("insured_name"),
            "phone": undel.get("phone"),
            "mail_description": undel.get("mail_description"),
            "action": "→ alert email to service@",
            "customer_email_found": (customer_info or {}).get("email", "NOT FOUND"),
        })
    summary["sections"]["undeliverable_mail"] = {"count": len(undeliverable), "rows": undeliverable}

    return summary


@router.get("/check-mode")
def check_mode():
    """Check current inbound email mode settings. No auth required."""
    natgen_dry = os.environ.get("NATGEN_DRY_RUN", "true")
    inbound_mode = os.environ.get("INBOUND_NONPAY_MODE", "dry_run")
    return {
        "NATGEN_DRY_RUN": natgen_dry,
        "INBOUND_NONPAY_MODE": inbound_mode,
        "emails_would_send": natgen_dry.lower() != "true" and inbound_mode == "live",
        "status": "SAFE — all sends blocked" if (natgen_dry.lower() == "true" or inbound_mode != "live") else "⚠️ LIVE — emails will send",
    }


@router.get("/db-check")
def db_check(db: Session = Depends(get_db)):
    """Quick database diagnostic - no auth required."""
    from app.models.sale import Sale
    from app.models.user import User
    from app.models.task import Task
    try:
        sale_count = db.query(Sale).count()
        user_count = db.query(User).count()
        task_count = db.query(Task).count()
        open_tasks = db.query(Task).filter(Task.status.in_(["open", "in_progress"])).count()
        latest_sale = db.query(Sale).order_by(Sale.id.desc()).first()
        return {
            "database": "connected",
            "total_sales": sale_count,
            "total_users": user_count,
            "total_tasks": task_count,
            "open_tasks": open_tasks,
            "latest_sale": {
                "id": latest_sale.id,
                "client_name": latest_sale.client_name,
                "created_at": str(latest_sale.created_at) if latest_sale.created_at else None,
            } if latest_sale else None,
        }
    except Exception as e:
        return {"database": "error", "error": str(e)}


@router.get("/nonpay-log")
def nonpay_log(db: Session = Depends(get_db)):
    """Check recent NonPayEmail records - no auth required."""
    from app.models.nonpay import NonPayEmail
    try:
        records = db.query(NonPayEmail).order_by(NonPayEmail.sent_at.desc()).limit(30).all()
        return [{
            "id": r.id,
            "policy_number": r.policy_number,
            "customer_email": r.customer_email,
            "email_status": r.email_status,
            "sent_at": str(r.sent_at) if r.sent_at else None,
            "carrier": r.carrier,
        } for r in records]
    except Exception as e:
        return {"error": str(e)}


@router.get("/customer-lookup/{policy_number}")
def customer_lookup(policy_number: str, db: Session = Depends(get_db)):
    """Look up customer by policy number - checks both sales and NowCerts tables."""
    from app.models.sale import Sale
    from app.models.user import User
    from app.models.customer import Customer, CustomerPolicy

    clean = policy_number.replace(" ", "").strip()
    base = clean[:-2] if len(clean) > 2 and clean[-2:] in ("00", "01") else clean

    # 1. Check sales table first
    sale = db.query(Sale).filter(Sale.policy_number.ilike(f"%{clean}%")).first()
    if not sale and base != clean:
        sale = db.query(Sale).filter(Sale.policy_number.ilike(f"%{base}%")).first()

    if sale:
        producer = None
        if sale.producer_id:
            producer = db.query(User).filter(User.id == sale.producer_id).first()
        return {
            "found": True, "source": "sales",
            "client_name": sale.client_name,
            "client_email": sale.client_email,
            "client_phone": getattr(sale, "client_phone", None),
            "policy_number": sale.policy_number,
            "carrier": sale.carrier,
            "producer": producer.full_name if producer else None,
        }

    # 2. Check NowCerts customer_policies table
    nc_policy = db.query(CustomerPolicy).filter(
        CustomerPolicy.policy_number.ilike(f"%{clean}%")
    ).first()
    if not nc_policy and base != clean:
        nc_policy = db.query(CustomerPolicy).filter(
            CustomerPolicy.policy_number.ilike(f"%{base}%")
        ).first()

    if nc_policy:
        customer = db.query(Customer).filter(Customer.id == nc_policy.customer_id).first()
        if customer:
            return {
                "found": True, "source": "nowcerts",
                "client_name": customer.full_name,
                "client_email": customer.email,
                "client_phone": customer.phone or customer.mobile_phone,
                "policy_number": nc_policy.policy_number,
                "carrier": nc_policy.carrier,
                "producer": nc_policy.agent_name,
            }

    return {"found": False, "policy": policy_number}


@router.post("/push-note")
async def push_note_only(request: Request, db: Session = Depends(get_db)):
    """Push a NowCerts note without sending an email. Body: {client_name, email, policy_number, carrier, note_type, requirement_type, due_date}"""
    import json
    raw = await request.body()
    body = json.loads(raw.decode()) if raw else {}
    
    note_type = body.get("note_type", "uw")
    
    if note_type == "uw":
        from app.services.uw_requirement_email import _add_uw_nowcerts_note
        _add_uw_nowcerts_note(
            client_name=body.get("client_name", ""),
            to_email=body.get("email", ""),
            policy_number=body.get("policy_number", ""),
            carrier=body.get("carrier", ""),
            requirement_type=body.get("requirement_type", "proof_of_continuous_insurance"),
            due_date=body.get("due_date"),
        )
    elif note_type == "nonpay":
        from app.services.nonpay_email import _add_nowcerts_nonpay_note
        _add_nowcerts_nonpay_note(
            client_name=body.get("client_name", ""),
            to_email=body.get("email", ""),
            policy_number=body.get("policy_number", ""),
            carrier=body.get("carrier", ""),
            amount_due=body.get("amount_due"),
            due_date=body.get("due_date"),
        )
    
    return {"status": "note_pushed", "type": note_type}
