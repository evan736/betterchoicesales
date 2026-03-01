"""Requote Campaign Engine — Re-engage former customers/leads with targeted X-date emails.

Upload lead lists (Excel/CSV), dedupe against NowCerts, schedule 2-touch drip emails
at 45 days and 15 days before X-date.
"""
import logging
import os
import re
import io
import hashlib
from datetime import datetime, timedelta, date
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query, Request
from sqlalchemy import Column, Integer, String, Text, Boolean, DateTime, Numeric, JSON, and_, or_, func as sqlfunc
from sqlalchemy.orm import Session

from app.core.database import Base, get_db
from app.core.security import get_current_user
from app.models.user import User

import httpx

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/campaigns", tags=["requote-campaigns"])

MAILGUN_API_KEY = os.getenv("MAILGUN_API_KEY", "")
MAILGUN_DOMAIN = os.getenv("MAILGUN_DOMAIN", "betterchoiceins.com")
AGENCY_FROM_EMAIL = os.getenv("AGENCY_FROM_EMAIL", "service@betterchoiceins.com")
ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")
AGENCY_REPLY_TO = os.getenv("AGENCY_REPLY_TO", "service@betterchoiceins.com")
# Campaign-specific: separate from/reply-to for requote outreach
CAMPAIGN_FROM_EMAIL = os.getenv("CAMPAIGN_FROM_EMAIL", "sales@betterchoiceins.com")
CAMPAIGN_REPLY_TO = os.getenv("CAMPAIGN_REPLY_TO", "sales@betterchoiceins.com")
NOWCERTS_USERNAME = os.getenv("NOWCERTS_USERNAME", "")
NOWCERTS_PASSWORD = os.getenv("NOWCERTS_PASSWORD", "")


# ═══════════════════════════════════════════════════════════════════
# MODELS
# ═══════════════════════════════════════════════════════════════════

class RequoteCampaign(Base):
    """A campaign batch — one upload creates one campaign."""
    __tablename__ = "requote_campaigns"

    id = Column(Integer, primary_key=True, index=True)
    name = Column(String, nullable=False)
    description = Column(Text, nullable=True)
    status = Column(String, default="active")  # active, paused, completed, archived

    # Upload tracking
    original_filename = Column(String, nullable=True)
    total_uploaded = Column(Integer, default=0)
    total_valid = Column(Integer, default=0)       # has email + X-date
    total_skipped = Column(Integer, default=0)      # current customer or no email
    total_deduped = Column(Integer, default=0)       # duplicate in this campaign

    # Drip config
    touch1_days_before = Column(Integer, default=45)  # First email X days before X-date
    touch2_days_before = Column(Integer, default=15)  # Second email X days before X-date

    # Stats
    emails_sent = Column(Integer, default=0)
    emails_opened = Column(Integer, default=0)
    responses_received = Column(Integer, default=0)
    requotes_generated = Column(Integer, default=0)

    created_by = Column(Integer, nullable=True)
    created_by_name = Column(String, nullable=True)
    created_at = Column(DateTime(timezone=True), server_default=sqlfunc.now())
    updated_at = Column(DateTime(timezone=True), onupdate=sqlfunc.now())


class RequoteLead(Base):
    """Individual lead within a campaign."""
    __tablename__ = "requote_leads"

    id = Column(Integer, primary_key=True, index=True)
    campaign_id = Column(Integer, nullable=False, index=True)

    # Contact info
    first_name = Column(String, nullable=True)
    last_name = Column(String, nullable=True)
    email = Column(String, nullable=True, index=True)
    phone = Column(String, nullable=True)
    address = Column(String, nullable=True)
    city = Column(String, nullable=True)
    state = Column(String, nullable=True)
    zip_code = Column(String, nullable=True)

    # Policy info from source file
    policy_type = Column(String, nullable=True)   # auto, home, etc.
    carrier = Column(String, nullable=True)        # their current/last carrier
    premium = Column(Numeric(10, 2), nullable=True)
    agent_name = Column(String, nullable=True)     # original agent

    # X-Date targeting
    x_date = Column(DateTime, nullable=True, index=True)
    follow_up_date = Column(DateTime, nullable=True)

    # NowCerts dedup
    is_current_customer = Column(Boolean, default=False)
    nowcerts_match_name = Column(String, nullable=True)
    nowcerts_match_id = Column(String, nullable=True)
    dedup_checked = Column(Boolean, default=False)

    # Drip status
    status = Column(String, default="pending")
    # pending, touch1_scheduled, touch1_sent, touch2_scheduled, touch2_sent,
    # responded, requoted, converted, opted_out, skipped

    # Touch 1 (45 days before X-date)
    touch1_scheduled_date = Column(DateTime, nullable=True)
    touch1_sent = Column(Boolean, default=False)
    touch1_sent_at = Column(DateTime(timezone=True), nullable=True)
    touch1_opened = Column(Boolean, default=False)

    # Touch 2 (15 days before X-date)
    touch2_scheduled_date = Column(DateTime, nullable=True)
    touch2_sent = Column(Boolean, default=False)
    touch2_sent_at = Column(DateTime(timezone=True), nullable=True)
    touch2_opened = Column(Boolean, default=False)

    # Opt-out
    unsubscribe_token = Column(String, nullable=True, index=True)
    opted_out = Column(Boolean, default=False)
    opted_out_at = Column(DateTime(timezone=True), nullable=True)

    # Source tracking
    source_row = Column(Integer, nullable=True)  # Row number in original file

    created_at = Column(DateTime(timezone=True), server_default=sqlfunc.now())
    updated_at = Column(DateTime(timezone=True), onupdate=sqlfunc.now())


class GlobalOptOut(Base):
    """Permanent email blocklist — opted-out emails are NEVER emailed again across any campaign."""
    __tablename__ = "global_opt_outs"

    id = Column(Integer, primary_key=True, index=True)
    email = Column(String, nullable=False, unique=True, index=True)
    reason = Column(String, default="unsubscribe")  # unsubscribe, manual, complaint
    source_campaign_id = Column(Integer, nullable=True)
    opted_out_at = Column(DateTime(timezone=True), server_default=sqlfunc.now())
# SELF-HEALING MIGRATION
# ═══════════════════════════════════════════════════════════════════

def run_migration(engine):
    """Create tables if they don't exist."""
    try:
        RequoteCampaign.__table__.create(engine, checkfirst=True)
        RequoteLead.__table__.create(engine, checkfirst=True)
        GlobalOptOut.__table__.create(engine, checkfirst=True)
        logger.info("Requote campaign tables ready (including global_opt_outs)")
    except Exception as e:
        logger.warning(f"Requote migration note: {e}")

    # Self-healing: add missing columns
    from sqlalchemy import inspect as sa_inspect, text
    try:
        insp = sa_inspect(engine)
        existing_cols = {c["name"] for c in insp.get_columns("requote_leads")} if insp.has_table("requote_leads") else set()
        with engine.connect() as conn:
            if "retarget_round" not in existing_cols:
                conn.execute(text("ALTER TABLE requote_leads ADD COLUMN retarget_round INTEGER DEFAULT 0"))
                conn.commit()
            if "original_lead_id" not in existing_cols:
                conn.execute(text("ALTER TABLE requote_leads ADD COLUMN original_lead_id INTEGER"))
                conn.commit()
    except Exception as e:
        logger.warning(f"Column migration note: {e}")


# ═══════════════════════════════════════════════════════════════════
# FILE PARSING — Smart column mapping for any format
# ═══════════════════════════════════════════════════════════════════

# Column name patterns → our field names
COLUMN_MAP = {
    'first_name': [r'first.?name', r'fname', r'first$'],
    'last_name': [r'last.?name', r'lname', r'last$', r'surname'],
    'email': [r'e.?mail', r'email.?address'],
    'phone': [r'^phone$', r'phone.?number', r'primary.?phone', r'home.?phone'],
    'mobile': [r'mobile', r'cell', r'cell.?phone'],
    'address': [r'street', r'address', r'street.?address', r'mailing.?address'],
    'city': [r'^city$'],
    'state': [r'^state$', r'^st$', r'state.?code'],
    'zip_code': [r'^zip', r'postal', r'zip.?code'],
    'policy_type': [r'policy.?type', r'line.?of.?business', r'lob', r'product', r'lead.?type'],
    'carrier': [r'carrier', r'company', r'insurance.?company', r'insurer'],
    'premium': [r'premium', r'annual.?premium', r'current.?premium'],
    'quote_premium': [r'quote.?premium', r'quoted'],
    'agent_name': [r'agent', r'producer', r'assigned', r'csr'],
    'x_date': [r'x.?date', r'expir', r'renewal.?date', r'eff.*date', r'policy.?exp', r'received.?date'],
    'follow_up_date': [r'follow.?up', r'callback', r'next.?contact'],
    'status': [r'^status$', r'lead.?status', r'disposition'],
}


def _map_columns(headers: list[str]) -> dict:
    """Map source file column headers to our field names.
    Returns {our_field: source_column_index}."""
    mapping = {}
    for field, patterns in COLUMN_MAP.items():
        for i, header in enumerate(headers):
            header_clean = header.strip().lower().replace(' ', '_')
            for pattern in patterns:
                if re.search(pattern, header_clean, re.IGNORECASE):
                    if field not in mapping:  # First match wins
                        mapping[field] = i
                    break
    return mapping


def _parse_date(val) -> Optional[datetime]:
    """Parse various date formats."""
    if not val:
        return None
    if isinstance(val, datetime):
        return val
    if isinstance(val, date):
        return datetime.combine(val, datetime.min.time())
    val = str(val).strip()
    if not val or val.lower() in ('nan', 'none', 'nat', ''):
        return None
    for fmt in [
        '%m/%d/%Y', '%m/%d/%y', '%Y-%m-%d', '%m-%d-%Y', '%m-%d-%y',
        '%m/%d/%Y %H:%M:%S', '%Y-%m-%d %H:%M:%S',
        '%d-%b-%Y', '%d-%b-%y', '%B %d, %Y',
    ]:
        try:
            return datetime.strptime(val, fmt)
        except ValueError:
            continue
    return None


def _parse_premium(val) -> Optional[float]:
    """Parse premium values like '$1,234.56'."""
    if not val:
        return None
    val = str(val).strip().replace('$', '').replace(',', '')
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _parse_file(file_bytes: bytes, filename: str) -> list[dict]:
    """Parse Excel or CSV file into list of lead dicts."""
    leads = []
    filename_lower = filename.lower()

    if filename_lower.endswith(('.xlsx', '.xls')):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Find header row — look for row containing 'email' or 'name' or 'first'
            header_idx = 0
            for i, row in enumerate(rows[:20]):
                row_str = ' '.join(str(c).lower() for c in row if c)
                if any(kw in row_str for kw in ['email', 'first name', 'first_name', 'fname']):
                    header_idx = i
                    break

            headers = [str(c or '').strip() for c in rows[header_idx]]
            col_map = _map_columns(headers)

            if not col_map:
                continue  # Can't map any columns

            for row_num, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
                if not row or not any(row):
                    continue
                lead = {'source_row': row_num}
                for field, col_idx in col_map.items():
                    if col_idx < len(row):
                        lead[field] = row[col_idx]
                leads.append(lead)
        wb.close()

    elif filename_lower.endswith('.csv'):
        import csv
        text = file_bytes.decode('utf-8', errors='ignore')
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return []

        headers = rows[0]
        col_map = _map_columns(headers)
        if not col_map:
            return []

        for row_num, row in enumerate(rows[1:], start=2):
            if not row or not any(row):
                continue
            lead = {'source_row': row_num}
            for field, col_idx in col_map.items():
                if col_idx < len(row):
                    lead[field] = row[col_idx]
            leads.append(lead)

    return leads


# ═══════════════════════════════════════════════════════════════════
# CUSTOMER DEDUP — Check local DB (synced from NowCerts) for active customers
# ═══════════════════════════════════════════════════════════════════

def _check_current_customer(db_session, email: str, first_name: str = "", last_name: str = "") -> dict:
    """Check local customer database for an active customer match.
    Uses the locally-synced NowCerts data (customers + customer_policies tables).
    Returns {"is_customer": bool, "match_name": str, "match_id": str, "has_active_policy": bool}.
    """
    from app.models.customer import Customer, CustomerPolicy
    from sqlalchemy import func as sqlfunc

    try:
        match = None

        # 1. Search by email (most reliable)
        if email:
            match = db_session.query(Customer).filter(
                sqlfunc.lower(Customer.email) == email.lower().strip(),
                Customer.is_active == True,
            ).first()

        # 2. Fallback: exact first + last name match
        if not match and first_name and last_name:
            match = db_session.query(Customer).filter(
                sqlfunc.lower(Customer.first_name) == first_name.lower().strip(),
                sqlfunc.lower(Customer.last_name) == last_name.lower().strip(),
                Customer.is_active == True,
            ).first()

        # 3. Fallback: phone match (strip non-digits)
        # (skipped for now — email + name should catch most)

        if not match:
            return {"is_customer": False, "match_name": None, "match_id": None, "has_active_policy": False}

        # Check if they have an active policy
        active_policy = db_session.query(CustomerPolicy).filter(
            CustomerPolicy.customer_id == match.id,
            sqlfunc.lower(CustomerPolicy.status).in_(["active", "in force", "inforce"]),
        ).first()

        return {
            "is_customer": True,
            "match_name": match.full_name or f"{match.first_name or ''} {match.last_name or ''}".strip(),
            "match_id": str(match.id),
            "has_active_policy": active_policy is not None,
        }

    except Exception as e:
        logger.warning(f"Customer dedup check failed for {email}: {e}")
        return {"is_customer": False, "match_name": None, "match_id": None, "has_active_policy": False}


# ═══════════════════════════════════════════════════════════════════
# EMAIL TEMPLATES
# ═══════════════════════════════════════════════════════════════════

def _requote_email_html(first_name: str, policy_type: str, carrier: str, x_date: str, 
                         touch_number: int, unsubscribe_url: str,
                         retarget_round: int = 0, city: str = "", state: str = "",
                         premium: float = None, use_ai: bool = True) -> str:
    """Generate branded requote campaign email — AI-powered with static fallback."""
    
    policy_label = (policy_type or 'insurance').replace('_', ' ').title()
    carrier_label = (carrier or 'your current carrier').title()

    # Build landing page URL
    import urllib.parse
    landing_params = urllib.parse.urlencode({
        'name': first_name,
        'type': policy_type or 'insurance',
        'xdate': x_date or '',
        'utm_campaign': f'requote_r{retarget_round}_t{touch_number}',
    })
    landing_url = f"https://better-choice-web.onrender.com/get-quote?{landing_params}"

    subject = ""
    body_content = ""

    # ── AI Generation ──
    if use_ai and ANTHROPIC_API_KEY:
        try:
            import anthropic
            client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

            now = datetime.utcnow()
            month_name = now.strftime("%B")
            season = "winter" if now.month in [12,1,2] else "spring" if now.month in [3,4,5] else "summer" if now.month in [6,7,8] else "fall"

            ai_prompt = f"""You are writing a marketing email for Better Choice Insurance Group. Match our website brand voice: clean, professional, confident, agency-focused.

LEAD INFO:
- First name: {first_name}
- Policy type: {policy_label}
- Renewal date: {x_date or 'upcoming'}
- Location: {city or ''} {state or ''}
- Touch: {touch_number} ({'first outreach' if touch_number == 1 else 'follow-up'})
- Retarget round: {retarget_round}
- Season: {month_name} ({season})

BRAND RULES (MUST FOLLOW):
- Write FROM "Better Choice Insurance" as a company, NOT from an individual agent
- NEVER use "[Agent Name]" or any placeholder name — the email is from the agency team
- NEVER mention a specific carrier name
- NEVER mention "self-service", "portal", "app", or "online account"
- Sign off as "The Better Choice Team" or "Better Choice Insurance Group"
- Tone: professional, friendly, concise — like a premium brand, not a local sales pitch
- Key value props: Compare rates from 15+ top carriers. Customers save avg $1,150/year. Free, fast, zero obligation.
- CTA: Reply with your declarations page, call (847) 908-5665, or click the button below
- Keep it SHORT: 2-3 short paragraphs max. Every sentence earns its place. No filler.
- Do NOT use the word "unsubscribe"

{"FOLLOW-UP: Touch 2. Be shorter and more direct. Reference that we recently reached out. Gentle urgency around their renewal." if touch_number == 2 else "FIRST OUTREACH: Introduce what Better Choice does and why it matters for them."}
{"RE-ENGAGEMENT round " + str(retarget_round) + ": Fresh angle — rate changes, new carrier partners, or " + season + " seasonal factors." if retarget_round > 0 else ""}

Respond ONLY with a JSON object (no markdown, no backticks):
{{"subject": "short specific subject line", "body": "HTML <p> tags with inline styles (font-size:16px;color:#333;line-height:1.6). <strong> for key stats only."}}"""

            response = client.messages.create(
                model=os.environ.get("SMART_INBOX_MODEL", "claude-sonnet-4-5-20250929"),
                max_tokens=800,
                messages=[{"role": "user", "content": ai_prompt}],
            )
            
            import json
            response_text = response.content[0].text.strip()
            if response_text.startswith("```"):
                response_text = response_text.split("\n", 1)[1] if "\n" in response_text else response_text[3:]
            if response_text.endswith("```"):
                response_text = response_text[:-3]
            
            ai_result = json.loads(response_text.strip())
            subject = ai_result.get("subject", "")
            body_content = ai_result.get("body", "")
            
            if subject and body_content:
                logger.info(f"AI generated email for {first_name} (touch {touch_number}, round {retarget_round})")
        except Exception as e:
            logger.warning(f"AI email generation failed, using fallback: {e}")
            subject = ""
            body_content = ""

    # ── Static Fallback ──
    if not subject or not body_content:
        if touch_number == 1:
            if retarget_round > 0:
                subject = f"{first_name}, insurance rates in your area have changed — time for a fresh look?"
                headline = "Rates Have Changed in Your Area"
                body_content = f"""
                <p style="font-size:16px;color:#333;line-height:1.6;">Hi {first_name},</p>
                <p style="font-size:16px;color:#333;line-height:1.6;">
                    We wanted to reach out again because insurance rates in your area have shifted since we last connected.
                    Many carriers have adjusted their pricing, and we've added new partners to our network — which means
                    fresh opportunities to save on your {policy_label} coverage.
                </p>
                <p style="font-size:16px;color:#333;line-height:1.6;">
                    As an independent agency representing 15+ carriers, we can run a quick comparison to see if there's
                    a better deal available for you. Our customers save an average of $1,150 per year.
                </p>
                <p style="font-size:16px;color:#333;line-height:1.6;">
                    <strong>Want us to take a look?</strong> Just reply with your current declarations page or call us at (847) 908-5665.
                </p>
                """
            else:
                subject = f"{first_name}, your {policy_label} policy is renewing soon — let's make sure you're getting the best rate"
                headline = "Your Policy is Renewing Soon"
                body_content = f"""
                <p style="font-size:16px;color:#333;line-height:1.6;">Hi {first_name},</p>
                <p style="font-size:16px;color:#333;line-height:1.6;">
                    Your {policy_label} policy is coming up for renewal 
                    around <strong>{x_date}</strong>. Insurance rates change every year, and many of our customers 
                    are surprised to find they can save significantly by shopping their renewal.
                </p>
                <p style="font-size:16px;color:#333;line-height:1.6;">
                    As an independent agency, we represent over 15 carriers and can compare rates to find you the 
                    best coverage at the best price — at no cost to you.
                </p>
                <p style="font-size:16px;color:#333;line-height:1.6;">
                    <strong>Want us to run a quick comparison?</strong> Simply reply to this email or give us a call. 
                    We just need your current declarations page and we'll do the rest.
                </p>
                """
        else:
            subject = f"Quick reminder: Your {policy_label} renewal is coming up, {first_name}"
            headline = "Don't Miss Your Renewal Window"
            body_content = f"""
            <p style="font-size:16px;color:#333;line-height:1.6;">Hi {first_name},</p>
            <p style="font-size:16px;color:#333;line-height:1.6;">
                Just a friendly reminder — your {policy_label} policy is renewing {'around <strong>' + x_date + '</strong>' if x_date else 'soon'}.
                {'That renewal is right around the corner!' if touch_number == 2 else ''}
            </p>
            <p style="font-size:16px;color:#333;line-height:1.6;">
                If you'd like us to shop your renewal and make sure you're getting the best rate, now is the 
                perfect time. We can often find savings of 15-30% by comparing multiple carriers.
            </p>
            <p style="font-size:16px;color:#333;line-height:1.6;">
                Just reply to this email with your current declarations page, or call us and we'll take care of everything.
            </p>
            """

    # ── Wrap in branded HTML template (matches landing page design) ──
    SITE = "https://better-choice-web.onrender.com"
    logo_url = f"{SITE}/carrier-logos/bci_logo_white.png"
    # Top 8 carrier logos for the strip
    carrier_row1 = ["safeco", "grange", "geico"]
    carrier_row2 = ["travelers", "branch", "hippo"]
    def logo_img(c):
        return f'<img src="{SITE}/carrier-logos/{c}.png" alt="{c}" style="height:22px;max-width:90px;display:inline-block;margin:0 14px;opacity:0.6;" />'
    row1_html = "".join(logo_img(c) for c in carrier_row1)
    row2_html = "".join(logo_img(c) for c in carrier_row2)
    carrier_imgs = f'{row1_html}<br style="line-height:32px;" />{row2_html}'

    html = f"""<!DOCTYPE html>
<html>
<head><meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"></head>
<body style="margin:0;padding:0;background:#0f172a;font-family:Arial,Helvetica,sans-serif;">
  <table width="100%" cellpadding="0" cellspacing="0" style="background:#0f172a;padding:24px 0;">
    <tr><td align="center">
      <table width="600" cellpadding="0" cellspacing="0" style="background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,0.3);">
        <!-- Header with Logo -->
        <tr><td style="background:linear-gradient(135deg, #0a1628 0%, #1a2d4a 100%);padding:32px 32px 28px;text-align:center;">
          <img src="{logo_url}" alt="Better Choice Insurance Group" style="height:56px;display:inline-block;" />
        </td></tr>
        <!-- Body -->
        <tr><td style="padding:32px 32px 16px;">{body_content}</td></tr>
        <!-- CTA -->
        <tr><td style="padding:8px 32px 28px;" align="center">
          <a href="{landing_url}" 
             style="display:inline-block;background:#2563eb;color:#fff;padding:16px 44px;border-radius:8px;text-decoration:none;font-size:16px;font-weight:700;letter-spacing:0.3px;">
            Get My Free Quote Comparison
          </a>
        </td></tr>
        <!-- Carrier Logos Strip -->
        <tr><td style="background:#f8fafc;padding:20px 24px;border-top:1px solid #e2e8f0;border-bottom:1px solid #e2e8f0;text-align:center;">
          <p style="margin:0 0 10px;font-size:11px;color:#94a3b8;text-transform:uppercase;letter-spacing:1.5px;font-weight:600;">We Compare Rates From Top Carriers</p>
          <div style="line-height:36px;">
            {carrier_imgs}
          </div>
        </td></tr>
        <!-- Stats Bar -->
        <tr><td style="padding:20px 32px;text-align:center;">
          <table width="100%" cellpadding="0" cellspacing="0">
            <tr>
              <td width="33%" style="text-align:center;">
                <p style="margin:0;font-size:22px;font-weight:800;color:#0a1628;">15+</p>
                <p style="margin:2px 0 0;font-size:11px;color:#94a3b8;">Carriers</p>
              </td>
              <td width="34%" style="text-align:center;border-left:1px solid #e2e8f0;border-right:1px solid #e2e8f0;">
                <p style="margin:0;font-size:22px;font-weight:800;color:#0a1628;">$1,150+</p>
                <p style="margin:2px 0 0;font-size:11px;color:#94a3b8;">Avg. Savings</p>
              </td>
              <td width="33%" style="text-align:center;">
                <p style="margin:0;font-size:22px;font-weight:800;color:#0a1628;">2,500+</p>
                <p style="margin:2px 0 0;font-size:11px;color:#94a3b8;">Happy Customers</p>
              </td>
            </tr>
          </table>
        </td></tr>
        <!-- Contact -->
        <tr><td style="padding:0 32px 20px;">
          <p style="font-size:14px;color:#64748b;line-height:1.6;text-align:center;margin:0;">
            <a href="tel:8479085665" style="color:#2563eb;text-decoration:none;font-weight:600;">(847) 908-5665</a>
            &nbsp;&middot;&nbsp;
            <a href="mailto:service@betterchoiceins.com" style="color:#2563eb;text-decoration:none;">service@betterchoiceins.com</a>
            <br>
            <a href="https://www.betterchoiceins.com" style="color:#94a3b8;text-decoration:none;font-size:13px;">www.betterchoiceins.com</a>
          </p>
        </td></tr>
        <!-- Footer -->
        <tr><td style="background:#f1f5f9;padding:16px 32px;border-top:1px solid #e2e8f0;">
          <p style="margin:0;font-size:11px;color:#94a3b8;text-align:center;">
            Better Choice Insurance Group &middot; Serving families across the Midwest<br>
            <a href="{unsubscribe_url}" style="color:#94a3b8;text-decoration:underline;">Unsubscribe</a>
          </p>
        </td></tr>
      </table>
    </td></tr>
  </table>
</body>
</html>"""
    return subject, html


# ═══════════════════════════════════════════════════════════════════
# SEND EMAIL VIA MAILGUN
# ═══════════════════════════════════════════════════════════════════

def _send_campaign_email(to_email: str, subject: str, html: str, 
                          reply_to: str = None) -> bool:
    """Send a campaign email via Mailgun. Returns True if sent."""
    if not MAILGUN_API_KEY:
        logger.warning("No MAILGUN_API_KEY — skipping email send")
        return False

    try:
        resp = httpx.post(
            f"https://api.mailgun.net/v3/{MAILGUN_DOMAIN}/messages",
            auth=("api", MAILGUN_API_KEY),
            data={
                "from": f"Better Choice Insurance <{CAMPAIGN_FROM_EMAIL}>",
                "to": to_email,
                "subject": subject,
                "html": html,
                "h:Reply-To": reply_to or CAMPAIGN_REPLY_TO,
                "o:tag": ["requote-campaign"],
            },
        )
        if resp.status_code == 200:
            logger.info(f"Campaign email sent to {to_email}: {subject[:50]}")
            return True
        else:
            logger.error(f"Mailgun error {resp.status_code}: {resp.text[:200]}")
            return False
    except Exception as e:
        logger.error(f"Email send failed: {e}")
        return False


# ═══════════════════════════════════════════════════════════════════
# API ENDPOINTS
# ═══════════════════════════════════════════════════════════════════

def _serialize_campaign(c: RequoteCampaign) -> dict:
    return {
        "id": c.id,
        "name": c.name,
        "description": c.description,
        "status": c.status,
        "original_filename": c.original_filename,
        "total_uploaded": c.total_uploaded,
        "total_valid": c.total_valid,
        "total_skipped": c.total_skipped,
        "total_deduped": c.total_deduped,
        "touch1_days_before": c.touch1_days_before,
        "touch2_days_before": c.touch2_days_before,
        "emails_sent": c.emails_sent,
        "emails_opened": c.emails_opened,
        "responses_received": c.responses_received,
        "requotes_generated": c.requotes_generated,
        "created_by_name": c.created_by_name,
        "created_at": c.created_at.isoformat() if c.created_at else None,
    }


def _serialize_lead(l: RequoteLead) -> dict:
    return {
        "id": l.id,
        "campaign_id": l.campaign_id,
        "first_name": l.first_name,
        "last_name": l.last_name,
        "email": l.email,
        "phone": l.phone,
        "address": l.address,
        "city": l.city,
        "state": l.state,
        "zip_code": l.zip_code,
        "policy_type": l.policy_type,
        "carrier": l.carrier,
        "premium": float(l.premium) if l.premium else None,
        "agent_name": l.agent_name,
        "x_date": l.x_date.isoformat() if l.x_date else None,
        "status": l.status,
        "is_current_customer": l.is_current_customer,
        "nowcerts_match_name": l.nowcerts_match_name,
        "dedup_checked": l.dedup_checked,
        "touch1_scheduled_date": l.touch1_scheduled_date.isoformat() if l.touch1_scheduled_date else None,
        "touch1_sent": l.touch1_sent,
        "touch1_sent_at": l.touch1_sent_at.isoformat() if l.touch1_sent_at else None,
        "touch2_scheduled_date": l.touch2_scheduled_date.isoformat() if l.touch2_scheduled_date else None,
        "touch2_sent": l.touch2_sent,
        "touch2_sent_at": l.touch2_sent_at.isoformat() if l.touch2_sent_at else None,
        "opted_out": l.opted_out,
        "created_at": l.created_at.isoformat() if l.created_at else None,
    }


# ── Campaign CRUD ─────────────────────────────────────────────────

@router.get("/")
def list_campaigns(
    status: str = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """List all campaigns."""
    q = db.query(RequoteCampaign)
    if status:
        q = q.filter(RequoteCampaign.status == status)
    campaigns = q.order_by(RequoteCampaign.created_at.desc()).all()
    return {"campaigns": [_serialize_campaign(c) for c in campaigns]}


@router.get("/preview-email")
def preview_email(
    first_name: str = Query("Sarah"),
    policy_type: str = Query("homeowners"),
    carrier: str = Query("State Farm"),
    x_date: str = Query("April 15, 2026"),
    touch: int = Query(1),
    retarget_round: int = Query(0),
    city: str = Query("Elgin"),
    state: str = Query("IL"),
    current_user: User = Depends(get_current_user),
):
    """Preview what a campaign email will look like — uses AI generation."""
    unsub_url = "https://better-choice-api.onrender.com/api/campaigns/unsubscribe/PREVIEW"
    subject, html = _requote_email_html(
        first_name, policy_type, carrier, x_date, touch, unsub_url,
        retarget_round=retarget_round, city=city, state=state,
    )
    return {"subject": subject, "html": html}


@router.post("/send-preview-emails")
async def send_preview_emails(
    to_email: str = Query("evan@betterchoiceins.com"),
    current_user: User = Depends(get_current_user),
):
    """Generate and send 5 sample AI-powered campaign emails to preview the system."""
    scenarios = [
        {"first_name": "Sarah", "policy_type": "homeowners", "carrier": "State Farm", "x_date": "April 15, 2026", "touch": 1, "retarget_round": 0, "city": "Naperville", "state": "IL"},
        {"first_name": "Mike", "policy_type": "auto", "carrier": "GEICO", "x_date": "May 1, 2026", "touch": 2, "retarget_round": 0, "city": "Schaumburg", "state": "IL"},
        {"first_name": "Jennifer", "policy_type": "homeowners", "carrier": "Allstate", "x_date": "June 10, 2026", "touch": 1, "retarget_round": 1, "city": "Elgin", "state": "IL"},
        {"first_name": "David", "policy_type": "home", "carrier": "Travelers", "x_date": "March 20, 2026", "touch": 1, "retarget_round": 2, "city": "Aurora", "state": "IL"},
        {"first_name": "Lisa", "policy_type": "bundle", "carrier": "Liberty Mutual", "x_date": "April 30, 2026", "touch": 2, "retarget_round": 1, "city": "St Charles", "state": "IL"},
    ]

    unsub_url = "https://better-choice-api.onrender.com/api/campaigns/unsubscribe/PREVIEW"
    sent = []
    for s in scenarios:
        subject, html = _requote_email_html(
            s["first_name"], s["policy_type"], s["carrier"], s["x_date"],
            s["touch"], unsub_url,
            retarget_round=s["retarget_round"], city=s["city"], state=s["state"],
        )
        label = f"[PREVIEW T{s['touch']}/R{s['retarget_round']}] {subject}"
        if _send_campaign_email(to_email, label, html):
            sent.append({"to": s["first_name"], "subject": label[:80]})

    return {"sent": len(sent), "emails": sent, "message": f"Sent {len(sent)} preview emails to {to_email}"}


@router.get("/{campaign_id}")
def get_campaign(
    campaign_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get campaign details with stats."""
    campaign = db.query(RequoteCampaign).filter(RequoteCampaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    # Get lead stats
    total_leads = db.query(RequoteLead).filter(RequoteLead.campaign_id == campaign_id).count()
    pending = db.query(RequoteLead).filter(
        RequoteLead.campaign_id == campaign_id,
        RequoteLead.status == "pending",
    ).count()
    touch1_sent = db.query(RequoteLead).filter(
        RequoteLead.campaign_id == campaign_id,
        RequoteLead.touch1_sent == True,
    ).count()
    touch2_sent = db.query(RequoteLead).filter(
        RequoteLead.campaign_id == campaign_id,
        RequoteLead.touch2_sent == True,
    ).count()
    current_customers = db.query(RequoteLead).filter(
        RequoteLead.campaign_id == campaign_id,
        RequoteLead.is_current_customer == True,
    ).count()
    opted_out = db.query(RequoteLead).filter(
        RequoteLead.campaign_id == campaign_id,
        RequoteLead.opted_out == True,
    ).count()

    # Upcoming sends (next 7 days)
    now = datetime.utcnow()
    week_ahead = now + timedelta(days=7)
    upcoming = db.query(RequoteLead).filter(
        RequoteLead.campaign_id == campaign_id,
        RequoteLead.is_current_customer == False,
        RequoteLead.opted_out == False,
        or_(
            and_(RequoteLead.touch1_sent == False, RequoteLead.touch1_scheduled_date != None,
                 RequoteLead.touch1_scheduled_date <= week_ahead),
            and_(RequoteLead.touch2_sent == False, RequoteLead.touch2_scheduled_date != None,
                 RequoteLead.touch2_scheduled_date <= week_ahead),
        ),
    ).count()

    data = _serialize_campaign(campaign)
    data["stats"] = {
        "total_leads": total_leads,
        "pending": pending,
        "touch1_sent": touch1_sent,
        "touch2_sent": touch2_sent,
        "current_customers": current_customers,
        "opted_out": opted_out,
        "upcoming_7_days": upcoming,
    }
    return data


# ── Upload & Ingest ───────────────────────────────────────────────

@router.post("/upload")
async def upload_leads(
    file: UploadFile = File(...),
    campaign_name: str = Form(""),
    touch1_days: int = Form(45),
    touch2_days: int = Form(15),
    check_nowcerts: bool = Form(True),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Upload a lead list file — creates campaign in DRAFT mode with NowCerts cross-reference."""
    file_bytes = await file.read()
    filename = file.filename or "upload"

    # Parse file
    raw_leads = _parse_file(file_bytes, filename)
    if not raw_leads:
        raise HTTPException(status_code=400, detail="Could not parse file or no data found. Supported formats: .xlsx, .xls, .csv")

    # Create campaign in DRAFT status
    if not campaign_name:
        campaign_name = f"Requote - {filename} ({datetime.now().strftime('%b %d')})"

    campaign = RequoteCampaign(
        name=campaign_name,
        original_filename=filename,
        touch1_days_before=touch1_days,
        touch2_days_before=touch2_days,
        total_uploaded=len(raw_leads),
        status="draft",
        created_by=current_user.id,
        created_by_name=current_user.full_name or current_user.username,
    )
    db.add(campaign)
    db.flush()

    # Process each lead
    valid_count = 0
    skipped_count = 0
    deduped_count = 0
    seen_emails = set()
    sample_leads = []  # For preview

    for raw in raw_leads:
        email = str(raw.get('email') or '').strip().lower()
        first_name = str(raw.get('first_name') or '').strip()
        last_name = str(raw.get('last_name') or '').strip()

        # Skip if no email
        if not email or '@' not in email:
            skipped_count += 1
            continue

        # Dedupe within this upload
        if email in seen_emails:
            deduped_count += 1
            continue
        seen_emails.add(email)

        # Check global opt-out blocklist FIRST
        if db.query(GlobalOptOut).filter(GlobalOptOut.email == email).first():
            skipped_count += 1
            continue

        # Also dedupe against existing leads in other active campaigns
        existing = db.query(RequoteLead).filter(
            RequoteLead.email == email,
            RequoteLead.opted_out == False,
        ).first()
        if existing and existing.campaign_id != campaign.id:
            deduped_count += 1
            continue

        # Parse dates
        x_date = _parse_date(raw.get('x_date'))
        follow_up_date = _parse_date(raw.get('follow_up_date'))

        # Parse premium
        premium = _parse_premium(raw.get('premium'))

        # Calculate drip schedule based on X-date
        touch1_date = None
        touch2_date = None
        lead_status = "pending"

        if x_date:
            touch1_date = x_date - timedelta(days=touch1_days)
            touch2_date = x_date - timedelta(days=touch2_days)
            now = datetime.utcnow()
            if touch2_date < now:
                lead_status = "past_xdate"
            elif touch1_date < now:
                lead_status = "touch2_scheduled"
            else:
                lead_status = "touch1_scheduled"

        # Generate unsubscribe token
        unsub_token = hashlib.sha256(f"{email}:{campaign.id}:{os.urandom(8).hex()}".encode()).hexdigest()[:24]

        lead = RequoteLead(
            campaign_id=campaign.id,
            first_name=first_name,
            last_name=last_name,
            email=email,
            phone=str(raw.get('phone') or raw.get('mobile') or '').strip(),
            address=str(raw.get('address') or '').strip(),
            city=str(raw.get('city') or '').strip(),
            state=str(raw.get('state') or '').strip(),
            zip_code=str(raw.get('zip_code') or '').strip(),
            policy_type=str(raw.get('policy_type') or '').strip(),
            carrier=str(raw.get('carrier') or '').strip(),
            premium=premium,
            agent_name=str(raw.get('agent_name') or '').strip(),
            x_date=x_date,
            follow_up_date=follow_up_date,
            status=lead_status,
            touch1_scheduled_date=touch1_date,
            touch2_scheduled_date=touch2_date,
            unsubscribe_token=unsub_token,
            source_row=raw.get('source_row'),
        )
        db.add(lead)
        valid_count += 1

        # Collect sample for preview (first 10)
        if len(sample_leads) < 10:
            sample_leads.append({
                "name": f"{first_name} {last_name}".strip(),
                "email": email,
                "carrier": str(raw.get('carrier') or '').strip(),
                "x_date": x_date.isoformat() if x_date else None,
                "premium": premium,
                "status": lead_status,
            })

    campaign.total_valid = valid_count
    campaign.total_skipped = skipped_count
    campaign.total_deduped = deduped_count
    db.commit()

    # ── Auto-run NowCerts dedup ──
    nowcerts_results = {"checked": 0, "current_customers": 0, "current_customer_list": []}

    # Check all leads against local customer database (instant — no external API calls)
    if check_nowcerts:
        leads_to_check = db.query(RequoteLead).filter(
            RequoteLead.campaign_id == campaign.id,
            RequoteLead.dedup_checked == False,
        ).all()

        for lead in leads_to_check:
            result = _check_current_customer(db, 
                lead.email or "",
                lead.first_name or "",
                lead.last_name or "",
            )
            lead.dedup_checked = True
            lead.is_current_customer = result["is_customer"]
            lead.nowcerts_match_name = result.get("match_name")
            lead.nowcerts_match_id = result.get("match_id")

            if result["is_customer"]:
                lead.status = "skipped"
                nowcerts_results["current_customers"] += 1
                nowcerts_results["current_customer_list"].append({
                    "name": f"{lead.first_name} {lead.last_name}".strip(),
                    "email": lead.email,
                    "nowcerts_match": result.get("match_name", ""),
                })
            nowcerts_results["checked"] += 1

        # Update campaign skip count
        campaign.total_skipped = db.query(RequoteLead).filter(
            RequoteLead.campaign_id == campaign.id,
            or_(RequoteLead.is_current_customer == True, RequoteLead.status == "skipped"),
        ).count()
        db.commit()

    # Calculate how many would actually receive emails
    sendable = db.query(RequoteLead).filter(
        RequoteLead.campaign_id == campaign.id,
        RequoteLead.is_current_customer == False,
        RequoteLead.opted_out == False,
        RequoteLead.status.in_(["pending", "touch1_scheduled", "touch2_scheduled"]),
    ).count()

    past_xdate = db.query(RequoteLead).filter(
        RequoteLead.campaign_id == campaign.id,
        RequoteLead.status == "past_xdate",
    ).count()

    unchecked = db.query(RequoteLead).filter(
        RequoteLead.campaign_id == campaign.id,
        RequoteLead.dedup_checked == False,
    ).count()

    return {
        "campaign_id": campaign.id,
        "campaign_name": campaign.name,
        "status": "draft",
        "total_uploaded": len(raw_leads),
        "total_valid": valid_count,
        "total_skipped": skipped_count,
        "total_deduped": deduped_count,
        "would_receive_email": sendable,
        "past_xdate": past_xdate,
        "nowcerts_check": nowcerts_results,
        "nowcerts_unchecked": unchecked,
        "sample_leads": sample_leads,
        "message": f"Campaign created in DRAFT mode. {sendable} leads would receive emails. {nowcerts_results['current_customers']} current customers excluded." + (f" {unchecked} leads still need NowCerts check — run Dedup to finish." if unchecked > 0 else ""),
    }


# ── NowCerts Dedup ────────────────────────────────────────────────

@router.post("/{campaign_id}/dedup")
async def run_dedup(
    campaign_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Run NowCerts dedup check on all unchecked leads in a campaign."""
    campaign = db.query(RequoteCampaign).filter(RequoteCampaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")

    leads = db.query(RequoteLead).filter(
        RequoteLead.campaign_id == campaign_id,
        RequoteLead.dedup_checked == False,
    ).limit(200).all()  # Process in batches of 200

    checked = 0
    matched = 0
    for lead in leads:
        result = _check_current_customer(db, 
            lead.email or "",
            lead.first_name or "",
            lead.last_name or "",
        )
        lead.dedup_checked = True
        lead.is_current_customer = result["is_customer"]
        lead.nowcerts_match_name = result.get("match_name")
        lead.nowcerts_match_id = result.get("match_id")

        if result["is_customer"]:
            lead.status = "skipped"
            matched += 1
        checked += 1

    # Update campaign stats
    campaign.total_skipped = db.query(RequoteLead).filter(
        RequoteLead.campaign_id == campaign_id,
        or_(RequoteLead.is_current_customer == True, RequoteLead.status == "skipped"),
    ).count()

    db.commit()

    remaining = db.query(RequoteLead).filter(
        RequoteLead.campaign_id == campaign_id,
        RequoteLead.dedup_checked == False,
    ).count()

    return {
        "checked": checked,
        "current_customers_found": matched,
        "remaining_unchecked": remaining,
        "message": f"Checked {checked} leads, found {matched} current customers. {f'{remaining} more to check.' if remaining > 0 else 'Dedup complete!'}",
    }


# ── Leads List ────────────────────────────────────────────────────

@router.get("/{campaign_id}/leads")
def list_leads(
    campaign_id: int,
    status: str = Query(None),
    search: str = Query(None),
    page: int = Query(1),
    per_page: int = Query(50),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """List leads in a campaign with filters."""
    q = db.query(RequoteLead).filter(RequoteLead.campaign_id == campaign_id)

    if status:
        if status == "current_customer":
            q = q.filter(RequoteLead.is_current_customer == True)
        elif status == "ready":
            q = q.filter(
                RequoteLead.is_current_customer == False,
                RequoteLead.opted_out == False,
                RequoteLead.status.in_(["pending", "touch1_scheduled", "touch2_scheduled"]),
            )
        else:
            q = q.filter(RequoteLead.status == status)

    if search:
        pattern = f"%{search}%"
        q = q.filter(or_(
            RequoteLead.first_name.ilike(pattern),
            RequoteLead.last_name.ilike(pattern),
            RequoteLead.email.ilike(pattern),
            RequoteLead.carrier.ilike(pattern),
        ))

    total = q.count()
    leads = q.order_by(RequoteLead.x_date.asc().nullslast()).offset((page - 1) * per_page).limit(per_page).all()

    return {
        "leads": [_serialize_lead(l) for l in leads],
        "total": total,
        "page": page,
        "per_page": per_page,
        "pages": (total + per_page - 1) // per_page,
    }


# ── Send Drip Emails ──────────────────────────────────────────────

@router.post("/{campaign_id}/send-due")
async def send_due_emails(
    campaign_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Send all emails that are due today for a campaign."""
    if current_user.role.lower() not in ("admin", "manager", "owner"):
        raise HTTPException(status_code=403, detail="Manager/admin only")

    campaign = db.query(RequoteCampaign).filter(RequoteCampaign.id == campaign_id).first()
    if not campaign or campaign.status != "active":
        raise HTTPException(status_code=400, detail="Campaign not found or not active")

    now = datetime.utcnow()
    sent_count = 0
    errors = 0

    # Touch 1: due today or overdue, not yet sent
    touch1_due = db.query(RequoteLead).filter(
        RequoteLead.campaign_id == campaign_id,
        RequoteLead.is_current_customer == False,
        RequoteLead.opted_out == False,
        RequoteLead.touch1_sent == False,
        RequoteLead.touch1_scheduled_date != None,
        RequoteLead.touch1_scheduled_date <= now,
        RequoteLead.email != None,
    ).limit(100).all()

    for lead in touch1_due:
        x_date_str = lead.x_date.strftime('%B %d, %Y') if lead.x_date else "soon"
        unsub_url = f"https://better-choice-api.onrender.com/api/campaigns/unsubscribe/{lead.unsubscribe_token}"
        retarget_round = getattr(lead, 'retarget_round', 0) or 0
        subject, html = _requote_email_html(
            lead.first_name or "Valued Customer",
            lead.policy_type, lead.carrier, x_date_str, 1, unsub_url,
            retarget_round=retarget_round,
            city=lead.city or "", state=lead.state or "",
            premium=float(lead.premium) if lead.premium else None,
        )
        if _send_campaign_email(lead.email, subject, html):
            lead.touch1_sent = True
            lead.touch1_sent_at = now
            lead.status = "touch1_sent"
            sent_count += 1
        else:
            errors += 1

    # Touch 2: due today or overdue, touch1 already sent, touch2 not yet sent
    touch2_due = db.query(RequoteLead).filter(
        RequoteLead.campaign_id == campaign_id,
        RequoteLead.is_current_customer == False,
        RequoteLead.opted_out == False,
        RequoteLead.touch1_sent == True,
        RequoteLead.touch2_sent == False,
        RequoteLead.touch2_scheduled_date != None,
        RequoteLead.touch2_scheduled_date <= now,
        RequoteLead.email != None,
    ).limit(100).all()

    for lead in touch2_due:
        x_date_str = lead.x_date.strftime('%B %d, %Y') if lead.x_date else "soon"
        unsub_url = f"https://better-choice-api.onrender.com/api/campaigns/unsubscribe/{lead.unsubscribe_token}"
        retarget_round = getattr(lead, 'retarget_round', 0) or 0
        subject, html = _requote_email_html(
            lead.first_name or "Valued Customer",
            lead.policy_type, lead.carrier, x_date_str, 2, unsub_url,
            retarget_round=retarget_round,
            city=lead.city or "", state=lead.state or "",
            premium=float(lead.premium) if lead.premium else None,
        )
        if _send_campaign_email(lead.email, subject, html):
            lead.touch2_sent = True
            lead.touch2_sent_at = now
            lead.status = "touch2_sent"
            sent_count += 1
        else:
            errors += 1

    # Update campaign stats
    campaign.emails_sent = db.query(RequoteLead).filter(
        RequoteLead.campaign_id == campaign_id,
        or_(RequoteLead.touch1_sent == True, RequoteLead.touch2_sent == True),
    ).count()
    db.commit()

    return {
        "sent": sent_count,
        "errors": errors,
        "message": f"Sent {sent_count} emails ({len(touch1_due)} touch 1, {len(touch2_due)} touch 2). {f'{errors} failed.' if errors else ''}",
    }


# ── Unsubscribe ───────────────────────────────────────────────────

@router.get("/unsubscribe/{token}")
def unsubscribe(token: str, db: Session = Depends(get_db)):
    """Handle email unsubscribe — adds to GLOBAL blocklist so they're never emailed again."""
    lead = db.query(RequoteLead).filter(RequoteLead.unsubscribe_token == token).first()
    if not lead:
        from fastapi.responses import HTMLResponse
        return HTMLResponse(content="""
        <html><body style="font-family:Arial;text-align:center;padding:60px;">
        <h2>Invalid or expired link</h2>
        <p>If you need help, contact us at service@betterchoiceins.com</p>
        </body></html>
        """)

    lead.opted_out = True
    lead.opted_out_at = datetime.utcnow()
    lead.status = "opted_out"

    # Add to GLOBAL opt-out blocklist — this email will never be emailed again
    if lead.email:
        existing = db.query(GlobalOptOut).filter(GlobalOptOut.email == lead.email.lower()).first()
        if not existing:
            db.add(GlobalOptOut(
                email=lead.email.lower(),
                reason="unsubscribe",
                source_campaign_id=lead.campaign_id,
            ))

        # Also opt out this email from ALL other active campaigns
        other_leads = db.query(RequoteLead).filter(
            RequoteLead.email == lead.email,
            RequoteLead.id != lead.id,
            RequoteLead.opted_out == False,
        ).all()
        for ol in other_leads:
            ol.opted_out = True
            ol.opted_out_at = datetime.utcnow()
            ol.status = "opted_out"

    db.commit()

    from fastapi.responses import HTMLResponse
    return HTMLResponse(content="""
    <html><body style="font-family:Arial;text-align:center;padding:60px;">
    <h2>You've been unsubscribed</h2>
    <p>You won't receive any more emails from Better Choice Insurance campaigns.</p>
    <p style="color:#666;font-size:14px;">If this was a mistake, please contact us at service@betterchoiceins.com</p>
    </body></html>
    """)


# ── Lead Actions ──────────────────────────────────────────────────

@router.post("/{campaign_id}/leads/{lead_id}/skip")
def skip_lead(
    campaign_id: int, lead_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Manually skip a lead (don't send emails)."""
    lead = db.query(RequoteLead).filter(
        RequoteLead.id == lead_id, RequoteLead.campaign_id == campaign_id,
    ).first()
    if not lead:
        raise HTTPException(status_code=404)
    lead.status = "skipped"
    db.commit()
    return {"ok": True}


@router.post("/{campaign_id}/leads/{lead_id}/mark-responded")
def mark_responded(
    campaign_id: int, lead_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Mark a lead as having responded."""
    lead = db.query(RequoteLead).filter(
        RequoteLead.id == lead_id, RequoteLead.campaign_id == campaign_id,
    ).first()
    if not lead:
        raise HTTPException(status_code=404)
    lead.status = "responded"
    db.commit()
    return {"ok": True}


@router.post("/{campaign_id}/leads/{lead_id}/mark-requoted")
def mark_requoted(
    campaign_id: int, lead_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Mark a lead as having been requoted."""
    lead = db.query(RequoteLead).filter(
        RequoteLead.id == lead_id, RequoteLead.campaign_id == campaign_id,
    ).first()
    if not lead:
        raise HTTPException(status_code=404)
    lead.status = "requoted"
    # Update campaign counter
    campaign = db.query(RequoteCampaign).filter(RequoteCampaign.id == campaign_id).first()
    if campaign:
        campaign.requotes_generated = (campaign.requotes_generated or 0) + 1
    db.commit()
    return {"ok": True}


# ── Campaign Actions ──────────────────────────────────────────────

@router.post("/{campaign_id}/activate")
def activate_campaign(campaign_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Move a draft campaign to active status — emails will start sending."""
    campaign = db.query(RequoteCampaign).filter(RequoteCampaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if campaign.status not in ("draft", "paused"):
        raise HTTPException(status_code=400, detail=f"Campaign is already {campaign.status}")
    campaign.status = "active"
    db.commit()

    sendable = db.query(RequoteLead).filter(
        RequoteLead.campaign_id == campaign_id,
        RequoteLead.is_current_customer == False,
        RequoteLead.opted_out == False,
        RequoteLead.status.in_(["pending", "touch1_scheduled", "touch2_scheduled"]),
    ).count()

    return {"status": "active", "would_receive_email": sendable, "message": f"Campaign activated! {sendable} leads will receive emails on schedule."}

@router.post("/{campaign_id}/delete-draft")
def delete_draft(campaign_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Delete a draft campaign and all its leads."""
    campaign = db.query(RequoteCampaign).filter(RequoteCampaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404, detail="Campaign not found")
    if campaign.status != "draft":
        raise HTTPException(status_code=400, detail="Can only delete draft campaigns")
    db.query(RequoteLead).filter(RequoteLead.campaign_id == campaign_id).delete()
    db.delete(campaign)
    db.commit()
    return {"message": "Draft campaign deleted."}

@router.post("/{campaign_id}/pause")
def pause_campaign(campaign_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    campaign = db.query(RequoteCampaign).filter(RequoteCampaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404)
    campaign.status = "paused"
    db.commit()
    return {"ok": True, "status": "paused"}


@router.post("/{campaign_id}/resume")
def resume_campaign(campaign_id: int, current_user: User = Depends(get_current_user), db: Session = Depends(get_db)):
    campaign = db.query(RequoteCampaign).filter(RequoteCampaign.id == campaign_id).first()
    if not campaign:
        raise HTTPException(status_code=404)
    campaign.status = "active"
    db.commit()
    return {"ok": True, "status": "active"}


# ── Calendar View Data ────────────────────────────────────────────

@router.get("/{campaign_id}/calendar")
def campaign_calendar(
    campaign_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get send schedule grouped by date for calendar view."""
    leads = db.query(RequoteLead).filter(
        RequoteLead.campaign_id == campaign_id,
        RequoteLead.is_current_customer == False,
        RequoteLead.opted_out == False,
    ).all()

    schedule = {}  # date_str -> {touch1: count, touch2: count}
    for lead in leads:
        if lead.touch1_scheduled_date and not lead.touch1_sent:
            d = lead.touch1_scheduled_date.strftime('%Y-%m-%d')
            schedule.setdefault(d, {"touch1": 0, "touch2": 0, "date": d})
            schedule[d]["touch1"] += 1
        if lead.touch2_scheduled_date and not lead.touch2_sent:
            d = lead.touch2_scheduled_date.strftime('%Y-%m-%d')
            schedule.setdefault(d, {"touch1": 0, "touch2": 0, "date": d})
            schedule[d]["touch2"] += 1

    return {
        "schedule": sorted(schedule.values(), key=lambda x: x["date"]),
    }


# ═══════════════════════════════════════════════════════════════════
# LANDING PAGE LEAD CAPTURE
# ═══════════════════════════════════════════════════════════════════

@router.post("/landing-lead")
async def capture_landing_lead(request: Request, db: Session = Depends(get_db)):
    """Capture a lead from the requote landing page (no auth required)."""
    try:
        data = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid JSON")

    name = (data.get("name") or "").strip()
    phone = (data.get("phone") or "").strip()
    email = (data.get("email") or "").strip()
    message = (data.get("message") or "").strip()
    policy_type = (data.get("policy_type") or "").strip()
    current_carrier = (data.get("current_carrier") or "").strip()
    renewal_date = (data.get("renewal_date") or "").strip()
    utm_campaign = (data.get("utm_campaign") or "").strip()

    if not name or not phone:
        raise HTTPException(status_code=400, detail="Name and phone are required")

    # Send alert email to Evan
    if MAILGUN_API_KEY:
        # Pre-compute values outside f-string (Python 3.11 doesn't allow backslashes in f-string expressions)
        phone_digits = phone.replace('(','').replace(')','').replace('-','').replace(' ','')
        first_name_only = name.split()[0] if name else 'Lead'

        email_row = f"<tr><td style='padding:8px 0;color:#64748b;'>Email</td><td style='padding:8px 0;'>{email}</td></tr>" if email else ""
        policy_row = f"<tr><td style='padding:8px 0;color:#64748b;'>Policy Type</td><td style='padding:8px 0;'>{policy_type}</td></tr>" if policy_type else ""
        carrier_row = f"<tr><td style='padding:8px 0;color:#64748b;'>Current Carrier</td><td style='padding:8px 0;'>{current_carrier}</td></tr>" if current_carrier else ""
        renewal_row = f"<tr><td style='padding:8px 0;color:#64748b;'>Renewal Date</td><td style='padding:8px 0;'>{renewal_date}</td></tr>" if renewal_date else ""
        utm_row = f"<tr><td style='padding:8px 0;color:#64748b;'>UTM Campaign</td><td style='padding:8px 0;font-size:12px;color:#94a3b8;'>{utm_campaign}</td></tr>" if utm_campaign else ""
        message_block = f'<div style="margin:16px 0;padding:12px 16px;background:#f0f9ff;border-radius:8px;border:1px solid #bae6fd;"><p style="margin:0;font-size:13px;color:#0369a1;"><strong>Message:</strong> {message}</p></div>' if message else ""

        alert_html = f"""<!DOCTYPE html><html><body style="font-family:Arial,sans-serif;margin:0;padding:20px;background:#f1f5f9;">
        <div style="max-width:600px;margin:0 auto;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,0.08);">
            <div style="background:linear-gradient(135deg,#059669,#10b981);padding:20px 28px;">
                <h2 style="margin:0;color:#fff;font-size:18px;">🎯 New Requote Lead from Landing Page</h2>
            </div>
            <div style="padding:24px 28px;">
                <table style="width:100%;font-size:14px;color:#334155;" cellpadding="0" cellspacing="0">
                    <tr><td style="padding:8px 0;color:#64748b;width:130px;">Name</td><td style="padding:8px 0;font-weight:700;">{name}</td></tr>
                    <tr><td style="padding:8px 0;color:#64748b;">Phone</td><td style="padding:8px 0;font-weight:700;">{phone}</td></tr>
                    {email_row}
                    {policy_row}
                    {carrier_row}
                    {renewal_row}
                    {utm_row}
                </table>
                {message_block}
                <div style="margin-top:20px;">
                    <a href="tel:{phone_digits}"
                       style="display:inline-block;background:#2563eb;color:#fff;padding:10px 24px;border-radius:6px;text-decoration:none;font-weight:700;font-size:14px;">
                        📞 Call {first_name_only} Now
                    </a>
                </div>
            </div>
        </div></body></html>"""

        try:
            httpx.post(
                f"https://api.mailgun.net/v3/{MAILGUN_DOMAIN}/messages",
                auth=("api", MAILGUN_API_KEY),
                data={
                    "from": f"ORBIT Lead Alert <{AGENCY_FROM_EMAIL}>",
                    "to": "evan@betterchoiceins.com",
                    "subject": f"🎯 New Requote Lead: {name} — {phone}",
                    "html": alert_html,
                    "o:tag": ["landing-page-lead"],
                },
            )
        except Exception as e:
            logger.warning(f"Failed to send landing lead alert: {e}")

    logger.info(f"Landing page lead captured: {name} / {phone} / {email}")
    return {"status": "ok", "message": "Lead captured successfully"}



# ═══════════════════════════════════════════════════════════════
# AI COVERAGE ANALYSIS — Reads dec pages with Claude Vision
# ═══════════════════════════════════════════════════════════════

@router.post("/coverage-analysis")
async def coverage_analysis(
    file: UploadFile = File(...),
    name: str = Form(""),
    phone: str = Form(""),
    email: str = Form(""),
):
    """
    Upload a declarations page (PDF or image). Claude Vision reads it,
    identifies the policy type, extracts key coverages, finds gaps,
    and recommends improvements with estimated savings.
    """
    import anthropic
    import base64
    import subprocess
    import tempfile

    ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
    if not ANTHROPIC_API_KEY:
        raise HTTPException(status_code=500, detail="AI analysis not configured")

    # Read the uploaded file
    file_bytes = await file.read()
    filename = file.filename or "upload"
    content_type = file.content_type or ""

    # Convert PDF to images if needed
    images_b64 = []

    if "pdf" in content_type.lower() or filename.lower().endswith(".pdf"):
        # Convert PDF pages to PNG using pdftoppm
        with tempfile.TemporaryDirectory() as tmpdir:
            pdf_path = os.path.join(tmpdir, "upload.pdf")
            with open(pdf_path, "wb") as f:
                f.write(file_bytes)

            try:
                subprocess.run(
                    ["pdftoppm", "-r", "200", "-png", pdf_path, os.path.join(tmpdir, "page")],
                    check=True, timeout=30, capture_output=True,
                )
            except (subprocess.CalledProcessError, FileNotFoundError) as e:
                logger.warning(f"pdftoppm failed: {e}, falling back to direct PDF")
                # Fallback: send raw PDF as document
                b64 = base64.standard_b64encode(file_bytes).decode("utf-8")
                images_b64.append({"type": "document", "media_type": "application/pdf", "data": b64})

            if not images_b64:
                # Collect converted page images (max 5 pages)
                import glob
                pages = sorted(glob.glob(os.path.join(tmpdir, "page-*.png")))[:5]
                for page_path in pages:
                    with open(page_path, "rb") as pf:
                        b64 = base64.standard_b64encode(pf.read()).decode("utf-8")
                        images_b64.append({"type": "image", "media_type": "image/png", "data": b64})
    else:
        # Image file (PNG, JPG)
        media_type = "image/png" if "png" in content_type.lower() else "image/jpeg"
        b64 = base64.standard_b64encode(file_bytes).decode("utf-8")
        images_b64.append({"type": "image", "media_type": media_type, "data": b64})

    if not images_b64:
        raise HTTPException(status_code=400, detail="Could not process uploaded file")

    # Build Claude Vision request
    content_blocks = []
    for img_data in images_b64:
        if img_data["type"] == "document":
            content_blocks.append({
                "type": "document",
                "source": {"type": "base64", "media_type": img_data["media_type"], "data": img_data["data"]},
            })
        else:
            content_blocks.append({
                "type": "image",
                "source": {"type": "base64", "media_type": img_data["media_type"], "data": img_data["data"]},
            })

    content_blocks.append({
        "type": "text",
        "text": """You are an expert insurance analyst at Better Choice Insurance Group, an independent agency. You're analyzing a customer's declarations page to help them understand their coverage and show why they should let one of our agents review their policy.

Extract and respond with ONLY a JSON object (no markdown, no backticks, no preamble):

{
  "policy_type": "homeowners|auto|renters|landlord|condo|umbrella|other",
  "carrier": "Name of the insurance carrier",
  "policy_number": "Policy number if visible",
  "effective_dates": "Policy period if visible",
  "insured_name": "Named insured if visible",
  "property_address": "Insured property address if visible",
  "key_coverages": {
    "description of coverage": "limit/amount"
  },
  "gaps": [
    {
      "area": "Specific coverage area name",
      "severity": "high|medium",
      "detail": "2-3 sentence explanation. Be specific to the actual numbers on the dec page."
    }
  ],
  "savings_estimate": "$X - $Y per year",
  "recommendation": "2-3 sentence personalized recommendation. ALWAYS mention the premium looks high for a home/vehicle in that specific city/area and recommend an immediate review by one of our licensed agents who can compare across 15+ carriers."
}

CRITICAL ANALYSIS RULES - READ CAREFULLY:

1. NEVER claim a coverage is missing if you can see it on the declarations page. Read EVERY line carefully. Water backup, sump pump, service line, equipment breakdown — these often appear as endorsements or additional coverages. If they are listed, acknowledge them.

2. HIGH DEDUCTIBLES ARE A SELLING POINT, NOT A GAP. We often recommend higher deductibles ($2,500-$5,000+) as a smart way to LOWER premiums. If the deductible is high, frame it positively: "Your $X deductible is a smart choice that keeps your premium lower. We can model different deductible scenarios to find your ideal balance of premium vs out-of-pocket risk."

3. PREMIUM ALWAYS LOOKS HIGH. No matter what the premium is, your recommendation MUST mention that it appears high for a property/vehicle in that specific city/area. Say something like "At $X/year for a home in [City], this premium appears elevated compared to what we typically see. We strongly recommend an immediate review — our agents often find significant savings by shopping across 15+ carriers."

4. For HOMEOWNERS gaps, focus on:
   - Liability limits relative to home value (if home is worth $500K+, recommend $500K-$1M liability or umbrella)
   - Whether replacement cost or ACV (if ACV, flag as HIGH priority)
   - Loss of use adequacy
   - Personal property coverage relative to dwelling
   - Equipment breakdown / service line coverage if not visible
   - DO NOT flag water backup as missing unless you've carefully confirmed it's truly not listed anywhere on the page

5. For AUTO gaps, focus on:
   - Liability limits (recommend 100/300/100 minimum)
   - Uninsured/underinsured motorist matching liability
   - Rental reimbursement
   - Roadside assistance
   - Gap coverage for newer/financed vehicles

6. Always have 2-4 gaps. At least one should be "high" severity.

7. Savings estimate: Be realistic but optimistic. Use 15-30% of current premium as the range. If premium is $3,000+, savings of $500-$1,100 is realistic.

8. The recommendation MUST always end with urgency — "We recommend scheduling a review with one of our agents as soon as possible" or similar."""
    })

    try:
        client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)
        response = client.messages.create(
            model=os.environ.get("SMART_INBOX_MODEL", "claude-sonnet-4-5-20250929"),
            max_tokens=2000,
            messages=[{"role": "user", "content": content_blocks}],
        )

        # Parse JSON from response
        response_text = response.content[0].text.strip()
        # Clean any markdown wrapping
        if response_text.startswith("```"):
            response_text = response_text.split("\n", 1)[1] if "\n" in response_text else response_text[3:]
        if response_text.endswith("```"):
            response_text = response_text[:-3]
        response_text = response_text.strip()

        import json
        analysis = json.loads(response_text)

        # Ensure required fields
        if "gaps" not in analysis:
            analysis["gaps"] = []
        if "savings_estimate" not in analysis:
            analysis["savings_estimate"] = "$300 - $800"
        if "recommendation" not in analysis:
            analysis["recommendation"] = "Based on our review, we see opportunities to improve your coverage. A licensed agent will compare rates across 15+ carriers."
        if "policy_type" not in analysis:
            analysis["policy_type"] = "unknown"

        logger.info(f"AI coverage analysis complete: {analysis.get('policy_type')} policy from {analysis.get('carrier', 'unknown')}, {len(analysis.get('gaps', []))} gaps found")

        # Send alert email to Evan with the analysis results
        if MAILGUN_API_KEY and (name or phone):
            gaps_html = ""
            for gap in analysis.get("gaps", []):
                severity_color = "#ef4444" if gap.get("severity") == "high" else "#f59e0b"
                severity_label = "HIGH" if gap.get("severity") == "high" else "MEDIUM"
                gaps_html += f'<tr><td style="padding:6px 8px;"><span style="background:{severity_color};color:#fff;padding:2px 6px;border-radius:3px;font-size:11px;font-weight:700;">{severity_label}</span></td><td style="padding:6px 8px;font-weight:700;">{gap.get("area","")}</td><td style="padding:6px 8px;color:#64748b;font-size:13px;">{gap.get("detail","")}</td></tr>'

            coverages_html = ""
            for cov_name, cov_val in analysis.get("key_coverages", {}).items():
                coverages_html += f'<tr><td style="padding:4px 8px;color:#64748b;">{cov_name}</td><td style="padding:4px 8px;font-weight:600;">{cov_val}</td></tr>'

            lead_name = name or "Unknown"
            lead_phone = phone or "N/A"
            lead_email = email or "N/A"
            carrier_found = analysis.get("carrier", "Unknown")
            policy_type_found = analysis.get("policy_type", "Unknown")
            savings = analysis.get("savings_estimate", "TBD")

            alert_html = f"""<!DOCTYPE html><html><body style="font-family:Arial,sans-serif;margin:0;padding:20px;background:#f1f5f9;">
<div style="max-width:650px;margin:0 auto;background:#fff;border-radius:12px;overflow:hidden;">
  <div style="background:linear-gradient(135deg,#7c3aed,#6366f1);padding:20px 28px;">
    <h2 style="margin:0;color:#fff;font-size:18px;">🧠 AI Coverage Analysis Lead</h2>
  </div>
  <div style="padding:24px 28px;">
    <table style="width:100%;font-size:14px;color:#334155;margin-bottom:16px;" cellpadding="0" cellspacing="0">
      <tr><td style="padding:6px 0;color:#64748b;width:120px;">Name</td><td style="font-weight:700;">{lead_name}</td></tr>
      <tr><td style="padding:6px 0;color:#64748b;">Phone</td><td style="font-weight:700;">{lead_phone}</td></tr>
      <tr><td style="padding:6px 0;color:#64748b;">Email</td><td>{lead_email}</td></tr>
      <tr><td style="padding:6px 0;color:#64748b;">File</td><td>{filename}</td></tr>
      <tr><td style="padding:6px 0;color:#64748b;">Carrier</td><td style="font-weight:700;">{carrier_found}</td></tr>
      <tr><td style="padding:6px 0;color:#64748b;">Policy Type</td><td>{policy_type_found}</td></tr>
      <tr><td style="padding:6px 0;color:#64748b;">Est. Savings</td><td style="font-weight:700;color:#2563eb;">{savings}</td></tr>
    </table>
    <h3 style="font-size:14px;color:#0f172a;margin:16px 0 8px;">Key Coverages Found</h3>
    <table style="width:100%;font-size:13px;border-collapse:collapse;">{coverages_html}</table>
    <h3 style="font-size:14px;color:#0f172a;margin:16px 0 8px;">Gaps Identified ({len(analysis.get('gaps',[]))})</h3>
    <table style="width:100%;font-size:13px;border-collapse:collapse;">{gaps_html}</table>
    <div style="margin-top:16px;padding:12px;background:#eff6ff;border-radius:8px;font-size:13px;color:#1e40af;">
      <strong>AI Recommendation:</strong> {analysis.get('recommendation','')}
    </div>
  </div>
</div></body></html>"""

            try:
                httpx.post(
                    f"https://api.mailgun.net/v3/{MAILGUN_DOMAIN}/messages",
                    auth=("api", MAILGUN_API_KEY),
                    data={
                        "from": f"ORBIT AI Analysis <{AGENCY_FROM_EMAIL}>",
                        "to": "evan@betterchoiceins.com",
                        "subject": f"🧠 AI Coverage Analysis: {lead_name} — {carrier_found} {policy_type_found} — {savings} savings",
                        "html": alert_html,
                        "o:tag": ["ai-coverage-analysis"],
                    },
                    files={"attachment": (filename, file_bytes)} if file_bytes else None,
                )
            except Exception as e:
                logger.warning(f"Failed to send AI analysis alert: {e}")

        return analysis

    except json.JSONDecodeError as e:
        logger.error(f"Failed to parse AI response as JSON: {e}")
        logger.error(f"Raw response: {response_text[:500]}")
        # Return a generic analysis if parsing fails
        return {
            "policy_type": "unknown",
            "carrier": "Unknown",
            "gaps": [
                {"area": "Coverage Review Needed", "severity": "high", "detail": "We were able to read your declarations page but need a licensed agent to complete the full analysis. There are likely coverage gaps and savings opportunities available."},
                {"area": "Rate Comparison", "severity": "medium", "detail": "Rates vary significantly across carriers. Without a full comparison across our 15+ partners, you could be overpaying by $500+ per year."},
            ],
            "savings_estimate": "$400 - $1,200",
            "recommendation": "We recommend speaking with one of our licensed agents who can manually review your policy and run a full rate comparison across all our carrier partners.",
        }
    except Exception as e:
        logger.error(f"AI coverage analysis failed: {e}")
        raise HTTPException(status_code=500, detail=f"Analysis failed: {str(e)}")


# ═══════════════════════════════════════════════════════════════════
# PIPELINE STATS — Where leads are in the system
# ═══════════════════════════════════════════════════════════════════

@router.get("/pipeline/stats")
def pipeline_stats(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get a system-wide view of where all leads are across all campaigns."""
    from sqlalchemy import case

    total = db.query(RequoteLead).count()

    # Status breakdown
    statuses = db.query(
        RequoteLead.status,
        sqlfunc.count(RequoteLead.id),
    ).group_by(RequoteLead.status).all()

    status_map = {s: c for s, c in statuses}

    # Current customers
    current_customers = db.query(RequoteLead).filter(RequoteLead.is_current_customer == True).count()

    # Global opt-outs
    global_optouts = db.query(GlobalOptOut).count()

    # Leads by campaign (active only)
    campaign_breakdown = db.query(
        RequoteCampaign.id,
        RequoteCampaign.name,
        RequoteCampaign.status,
        sqlfunc.count(RequoteLead.id).label("total_leads"),
        sqlfunc.sum(case((RequoteLead.status.in_(["touch1_sent", "touch2_sent"]), 1), else_=0)).label("emails_sent"),
        sqlfunc.sum(case((RequoteLead.status == "responded", 1), else_=0)).label("responded"),
        sqlfunc.sum(case((RequoteLead.status == "requoted", 1), else_=0)).label("requoted"),
        sqlfunc.sum(case((RequoteLead.is_current_customer == True, 1), else_=0)).label("current_customers"),
        sqlfunc.sum(case((RequoteLead.opted_out == True, 1), else_=0)).label("opted_out"),
    ).join(RequoteLead, RequoteLead.campaign_id == RequoteCampaign.id).group_by(
        RequoteCampaign.id, RequoteCampaign.name, RequoteCampaign.status,
    ).order_by(RequoteCampaign.created_at.desc()).limit(20).all()

    campaigns = []
    for c in campaign_breakdown:
        campaigns.append({
            "id": c.id, "name": c.name, "status": c.status,
            "total_leads": c.total_leads, "emails_sent": c.emails_sent,
            "responded": c.responded, "requoted": c.requoted,
            "current_customers": c.current_customers, "opted_out": c.opted_out,
        })

    return {
        "total_leads": total,
        "pipeline": {
            "pending": status_map.get("pending", 0),
            "touch1_scheduled": status_map.get("touch1_scheduled", 0),
            "touch1_sent": status_map.get("touch1_sent", 0),
            "touch2_scheduled": status_map.get("touch2_scheduled", 0),
            "touch2_sent": status_map.get("touch2_sent", 0),
            "responded": status_map.get("responded", 0),
            "requoted": status_map.get("requoted", 0),
            "opted_out": status_map.get("opted_out", 0),
            "skipped": status_map.get("skipped", 0),
            "past_xdate": status_map.get("past_xdate", 0),
        },
        "current_customers_excluded": current_customers,
        "global_opt_outs": global_optouts,
        "campaigns": campaigns,
    }


# ═══════════════════════════════════════════════════════════════════
# RE-CHECK NOWCERTS — Sweep active campaigns for new customers
# ═══════════════════════════════════════════════════════════════════

@router.post("/recheck-nowcerts")
async def recheck_nowcerts_all(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Re-check all active campaign leads against NowCerts. Remove any that became customers."""
    active_campaigns = db.query(RequoteCampaign).filter(
        RequoteCampaign.status.in_(["active", "draft"]),
    ).all()

    total_checked = 0
    new_customers = 0

    for campaign in active_campaigns:
        leads = db.query(RequoteLead).filter(
            RequoteLead.campaign_id == campaign.id,
            RequoteLead.is_current_customer == False,
            RequoteLead.opted_out == False,
            RequoteLead.status.notin_(["skipped", "opted_out"]),
        ).limit(500).all()

        for lead in leads:
            result = _check_current_customer(db, 
                lead.email or "",
                lead.first_name or "",
                lead.last_name or "",
            )
            if result["is_customer"]:
                lead.is_current_customer = True
                lead.nowcerts_match_name = result.get("match_name")
                lead.nowcerts_match_id = result.get("match_id")
                lead.status = "skipped"
                new_customers += 1
            total_checked += 1

    db.commit()

    return {
        "total_checked": total_checked,
        "new_customers_found": new_customers,
        "message": f"Checked {total_checked} leads across {len(active_campaigns)} campaigns. Found {new_customers} new customers to exclude.",
    }


# ═══════════════════════════════════════════════════════════════════
# AUTO-RETARGET — Bi-annual re-engagement for unconverted leads
# ═══════════════════════════════════════════════════════════════════

@router.post("/auto-retarget")
async def auto_retarget(
    min_age_days: int = Query(180, description="Min days since last campaign to retarget"),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Auto-create retarget campaigns from completed campaigns with unconverted leads.
    
    Runs bi-annually (every 6 months). For each completed campaign older than min_age_days:
    1. Find leads that completed both touches but didn't respond/convert
    2. Re-check NowCerts — skip anyone who became a customer
    3. Check global opt-out — skip permanently opted-out emails
    4. Create a new campaign with fresh X-dates (6 months from now)
    5. AI will generate completely new email content for each lead
    """
    now = datetime.utcnow()
    cutoff = now - timedelta(days=min_age_days)

    # Find campaigns that are old enough and have unconverted leads
    eligible_campaigns = db.query(RequoteCampaign).filter(
        RequoteCampaign.status.in_(["active", "completed"]),
        RequoteCampaign.created_at <= cutoff,
    ).all()

    created_campaigns = []

    for old_campaign in eligible_campaigns:
        # Find unconverted leads: both touches sent, not responded/requoted/opted_out
        unconverted = db.query(RequoteLead).filter(
            RequoteLead.campaign_id == old_campaign.id,
            RequoteLead.is_current_customer == False,
            RequoteLead.opted_out == False,
            RequoteLead.status.in_(["touch1_sent", "touch2_sent"]),
        ).all()

        if not unconverted:
            continue

        # Filter out global opt-outs and re-check NowCerts
        retarget_leads = []
        for lead in unconverted:
            email = (lead.email or "").lower()
            if not email:
                continue

            # Check global opt-out
            if db.query(GlobalOptOut).filter(GlobalOptOut.email == email).first():
                lead.opted_out = True
                lead.status = "opted_out"
                continue

            # Re-check NowCerts
            nc_result = _check_current_customer(db, email, lead.first_name or "", lead.last_name or "")
            if nc_result["is_customer"]:
                lead.is_current_customer = True
                lead.nowcerts_match_name = nc_result.get("match_name")
                lead.status = "skipped"
                continue

            retarget_leads.append(lead)

        if not retarget_leads:
            db.commit()
            continue

        # Determine retarget round
        max_round = max((getattr(l, 'retarget_round', 0) or 0) for l in retarget_leads)
        new_round = max_round + 1

        # Create new retarget campaign
        new_campaign = RequoteCampaign(
            name=f"Retarget R{new_round} — {old_campaign.name}",
            description=f"Auto-generated bi-annual retarget from '{old_campaign.name}' (round {new_round})",
            original_filename=old_campaign.original_filename,
            touch1_days_before=45,
            touch2_days_before=15,
            total_uploaded=len(retarget_leads),
            status="draft",  # Always draft — Evan reviews before activating
            created_by=current_user.id,
            created_by_name="ORBIT Auto-Retarget",
        )
        db.add(new_campaign)
        db.flush()

        valid = 0
        for old_lead in retarget_leads:
            # Set new X-date to 6 months from now
            new_xdate = now + timedelta(days=180)
            touch1_date = new_xdate - timedelta(days=45)
            touch2_date = new_xdate - timedelta(days=15)

            unsub_token = hashlib.sha256(f"{old_lead.email}:{new_campaign.id}:{os.urandom(8).hex()}".encode()).hexdigest()[:24]

            new_lead = RequoteLead(
                campaign_id=new_campaign.id,
                first_name=old_lead.first_name,
                last_name=old_lead.last_name,
                email=old_lead.email,
                phone=old_lead.phone,
                address=old_lead.address,
                city=old_lead.city,
                state=old_lead.state,
                zip_code=old_lead.zip_code,
                policy_type=old_lead.policy_type,
                carrier=old_lead.carrier,
                premium=old_lead.premium,
                agent_name=old_lead.agent_name,
                x_date=new_xdate,
                status="touch1_scheduled",
                touch1_scheduled_date=touch1_date,
                touch2_scheduled_date=touch2_date,
                unsubscribe_token=unsub_token,
            )
            # Set retarget tracking
            try:
                new_lead.retarget_round = new_round
                new_lead.original_lead_id = old_lead.id
            except Exception:
                pass  # columns may not exist yet

            db.add(new_lead)
            valid += 1

        new_campaign.total_valid = valid
        db.commit()

        # Mark old campaign as completed
        old_campaign.status = "completed"
        db.commit()

        created_campaigns.append({
            "campaign_id": new_campaign.id,
            "campaign_name": new_campaign.name,
            "leads": valid,
            "retarget_round": new_round,
            "source_campaign": old_campaign.name,
        })

    return {
        "retarget_campaigns_created": len(created_campaigns),
        "campaigns": created_campaigns,
        "message": f"Created {len(created_campaigns)} retarget campaigns in draft mode. Review and activate when ready.",
    }
