"""BEACON Knowledge Base — learnable knowledge from PDFs, screenshots, corrections, and conversations."""
import logging
import os
import re
import base64
import hashlib
import json
from datetime import datetime
from typing import Optional, List

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from sqlalchemy import Column, Integer, String, Text, Boolean, DateTime, Float
from sqlalchemy.orm import Session
from sqlalchemy.sql import func

from app.core.database import Base, get_db
from app.core.security import get_current_user
from app.models.user import User

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/beacon-kb", tags=["beacon-knowledge"])

ANTHROPIC_API_KEY = os.getenv("ANTHROPIC_API_KEY", "")


# ── Models ───────────────────────────────────────────────────────

class BeaconKnowledge(Base):
    __tablename__ = "beacon_knowledge"

    id = Column(Integer, primary_key=True, index=True)
    source_type = Column(String, nullable=False)       # pdf, screenshot, correction, conversation
    title = Column(String, nullable=False)              # Display title
    content = Column(Text, nullable=False)              # Extracted/written text content
    summary = Column(Text, nullable=True)               # AI-generated summary for quick matching
    tags = Column(String, nullable=True)                # Comma-separated: "natgen,auto,florida"
    carrier = Column(String, nullable=True)             # Carrier name if applicable
    status = Column(String, default="pending")          # pending, approved, rejected
    submitted_by = Column(Integer, nullable=True)       # User ID who submitted
    submitted_by_name = Column(String, nullable=True)   # User name
    reviewed_by = Column(Integer, nullable=True)        # Manager who approved/rejected
    reviewed_by_name = Column(String, nullable=True)
    reviewed_at = Column(DateTime, nullable=True)
    review_note = Column(Text, nullable=True)           # Manager's note on approval/rejection
    original_filename = Column(String, nullable=True)   # Original upload filename
    file_hash = Column(String, nullable=True)           # SHA256 of uploaded file (dedup)
    created_at = Column(DateTime, server_default=func.now())
    updated_at = Column(DateTime, server_default=func.now(), onupdate=func.now())


# ── Helpers ──────────────────────────────────────────────────────

def _serialize(k: BeaconKnowledge) -> dict:
    return {
        "id": k.id,
        "source_type": k.source_type,
        "title": k.title,
        "content": k.content[:500] + ("..." if len(k.content) > 500 else ""),
        "full_content": k.content,
        "summary": k.summary,
        "tags": k.tags,
        "carrier": k.carrier,
        "status": k.status,
        "submitted_by": k.submitted_by,
        "submitted_by_name": k.submitted_by_name,
        "reviewed_by_name": k.reviewed_by_name,
        "reviewed_at": k.reviewed_at.isoformat() if k.reviewed_at else None,
        "review_note": k.review_note,
        "original_filename": k.original_filename,
        "created_at": k.created_at.isoformat() if k.created_at else None,
    }


def _is_manager(user: User) -> bool:
    return user.role.lower() in ("admin", "manager", "owner")


def _extract_text_from_pdf(pdf_bytes: bytes) -> str:
    """Extract text from PDF bytes using pdfplumber, with OCR fallback for scanned PDFs."""
    text = ""
    num_pages = 0
    
    # Phase 1: Try text extraction with pdfplumber
    try:
        import pdfplumber
        import io
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            pages = []
            num_pages = len(pdf.pages)
            for page in pdf.pages:
                page_text = page.extract_text()
                if page_text:
                    pages.append(page_text)
            text = "\n\n".join(pages)
    except ImportError:
        logger.warning("pdfplumber not installed, trying PyPDF2")
        try:
            import PyPDF2
            import io
            reader = PyPDF2.PdfReader(io.BytesIO(pdf_bytes))
            pages = []
            num_pages = len(reader.pages)
            for page in reader.pages:
                page_text = page.extract_text()
                if page_text:
                    pages.append(page_text)
            text = "\n\n".join(pages)
        except Exception as e:
            logger.error(f"PDF extraction failed: {e}")
    
    # Phase 2: Check if we got enough text — if not, this is likely a scanned PDF
    # Heuristic: if we got <1000 chars from a multi-page PDF, it's probably scanned
    chars_per_page = len(text) / max(num_pages, 1)
    if chars_per_page < 200 and num_pages > 0:
        logger.info(f"PDF appears scanned: {len(text)} chars from {num_pages} pages ({chars_per_page:.0f} chars/page). Attempting OCR via Claude Vision...")
        ocr_text = _ocr_pdf_with_vision(pdf_bytes, num_pages)
        if ocr_text and len(ocr_text) > len(text) * 2:
            logger.info(f"OCR extracted {len(ocr_text):,} chars (vs {len(text)} from text extraction)")
            return ocr_text
        elif ocr_text:
            # Combine both
            text = text + "\n\n" + ocr_text
    
    return text


def _ocr_pdf_with_vision(pdf_bytes: bytes, num_pages: int) -> str:
    """Convert PDF pages to images and use Claude Vision to OCR them.
    Processes up to 20 pages to stay within reasonable API costs."""
    if not ANTHROPIC_API_KEY:
        return ""
    
    try:
        import io
        from pdf2image import convert_from_bytes
        import httpx
        
        # Convert PDF pages to images (limit to 20 pages for cost control)
        max_pages = min(num_pages, 20)
        logger.info(f"Converting {max_pages} PDF pages to images for OCR...")
        
        images = convert_from_bytes(
            pdf_bytes, 
            first_page=1, 
            last_page=max_pages,
            dpi=200,  # Good balance of quality vs size
            fmt="jpeg",
        )
        
        all_text = []
        
        # Process pages in batches of 4 (Claude can handle multiple images per request)
        batch_size = 4
        for batch_start in range(0, len(images), batch_size):
            batch_images = images[batch_start:batch_start + batch_size]
            page_range = f"{batch_start + 1}-{batch_start + len(batch_images)}"
            
            # Build content array with all images in this batch
            content = []
            for i, img in enumerate(batch_images):
                buf = io.BytesIO()
                img.save(buf, format="JPEG", quality=85)
                b64 = base64.b64encode(buf.getvalue()).decode("utf-8")
                content.append({
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": "image/jpeg",
                        "data": b64,
                    },
                })
            
            content.append({
                "type": "text",
                "text": f"Extract ALL text from these insurance document pages (pages {page_range}). "
                       f"This is an underwriting guidelines PDF. Extract everything verbatim — "
                       f"every rule, eligibility requirement, coverage detail, exclusion, "
                       f"claims limit, loss history threshold, and any tables or lists. "
                       f"Preserve the document structure with headers and sections. "
                       f"Do NOT summarize — extract the complete text.",
            })
            
            try:
                with httpx.Client(timeout=120.0) as client:
                    resp = client.post(
                        "https://api.anthropic.com/v1/messages",
                        headers={
                            "x-api-key": ANTHROPIC_API_KEY,
                            "anthropic-version": "2023-06-01",
                            "content-type": "application/json",
                        },
                        json={
                            "model": "claude-sonnet-4-5-20250929",  # Sonnet for accurate OCR
                            "max_tokens": 4096,
                            "messages": [{"role": "user", "content": content}],
                        },
                    )
                    
                    if resp.status_code == 200:
                        data = resp.json()
                        for block in data.get("content", []):
                            if block.get("type") == "text":
                                all_text.append(block["text"])
                        logger.info(f"OCR pages {page_range}: extracted {sum(len(t) for t in all_text):,} chars so far")
                    else:
                        logger.error(f"Vision OCR error for pages {page_range}: {resp.status_code} - {resp.text[:200]}")
            except Exception as e:
                logger.error(f"Vision OCR failed for pages {page_range}: {e}")
        
        result = "\n\n".join(all_text)
        if num_pages > max_pages:
            result += f"\n\n[Note: OCR processed first {max_pages} of {num_pages} pages]"
        
        return result
        
    except ImportError as e:
        logger.error(f"pdf2image not available for OCR: {e}")
        return ""
    except Exception as e:
        logger.error(f"OCR PDF conversion failed: {e}")
        return ""


def _extract_text_from_image(image_bytes: bytes, media_type: str = "image/png") -> str:
    """Use Claude vision to extract text/info from a screenshot."""
    if not ANTHROPIC_API_KEY:
        return "[Image uploaded but no API key for extraction]"
    
    try:
        import httpx
        b64 = base64.b64encode(image_bytes).decode("utf-8")
        
        with httpx.Client(timeout=60.0) as client:
            resp = client.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key": ANTHROPIC_API_KEY,
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json",
                },
                json={
                    "model": "claude-haiku-4-5-20251001",
                    "max_tokens": 2048,
                    "messages": [{
                        "role": "user",
                        "content": [
                            {
                                "type": "image",
                                "source": {
                                    "type": "base64",
                                    "media_type": media_type,
                                    "data": b64,
                                },
                            },
                            {
                                "type": "text",
                                "text": "Extract ALL text and important information from this insurance-related screenshot. Include carrier names, policy details, guidelines, rates, rules, contact info, and any other relevant data. Format clearly with headers and bullet points.",
                            },
                        ],
                    }],
                },
            )
            
            if resp.status_code == 200:
                data = resp.json()
                text = ""
                for block in data.get("content", []):
                    if block.get("type") == "text":
                        text += block["text"]
                return text
            else:
                logger.error(f"Vision API error: {resp.status_code}")
                return "[Failed to extract text from image]"
    except Exception as e:
        logger.error(f"Image extraction failed: {e}")
        return f"[Image extraction error: {str(e)[:100]}]"


def _extract_text_from_excel(file_bytes: bytes, filename: str = "") -> str:
    """Extract text from Excel files (.xlsx/.xls)."""
    try:
        import openpyxl
        import io
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        parts = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(h) if h is not None else "" for h in rows[0]]
            parts.append(f"=== Sheet: {sheet_name} ===")
            parts.append(f"Columns: {', '.join(h for h in headers if h)}")
            parts.append("")
            for row in rows[1:]:
                row_data = "; ".join(
                    f"{headers[j]}: {row[j]}"
                    for j in range(len(row))
                    if j < len(headers) and headers[j] and row[j] is not None
                )
                if row_data.strip():
                    parts.append(row_data)
            parts.append("")
        wb.close()
        return "\n".join(parts)
    except ImportError:
        logger.warning("openpyxl not installed — cannot parse Excel")
        return ""
    except Exception as e:
        logger.error(f"Excel extraction failed: {e}")
        return ""


def _extract_text_from_csv(file_bytes: bytes) -> str:
    """Extract text from CSV files."""
    try:
        import csv
        import io
        text = file_bytes.decode("utf-8", errors="replace")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return ""
        headers = rows[0]
        parts = [f"Columns: {', '.join(headers)}", ""]
        for row in rows[1:]:
            row_data = "; ".join(
                f"{headers[j]}: {row[j]}"
                for j in range(len(row))
                if j < len(headers) and row[j]
            )
            if row_data.strip():
                parts.append(row_data)
        return "\n".join(parts)
    except Exception as e:
        logger.error(f"CSV extraction failed: {e}")
        return ""


def _generate_summary(content: str, title: str) -> str:
    """Generate a short summary + tags using AI."""
    if not ANTHROPIC_API_KEY or len(content) < 50:
        return ""
    
    try:
        import httpx
        with httpx.Client(timeout=30.0) as client:
            resp = client.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key": ANTHROPIC_API_KEY,
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json",
                },
                json={
                    "model": "claude-haiku-4-5-20251001",
                    "max_tokens": 300,
                    "messages": [{
                        "role": "user",
                        "content": f"Summarize this insurance knowledge entry in 1-2 sentences for quick reference. Title: {title}\n\nContent:\n{content[:3000]}",
                    }],
                },
            )
            if resp.status_code == 200:
                data = resp.json()
                for block in data.get("content", []):
                    if block.get("type") == "text":
                        return block["text"]
        return ""
    except Exception:
        return ""


def _generate_smart_title(content: str, carrier: str, filename: str) -> str:
    """Generate a meaningful title from PDF/document content."""
    first_500 = content[:500].upper()
    
    # Detect state
    STATES = [
        'ALABAMA', 'ALASKA', 'ARIZONA', 'ARKANSAS', 'CALIFORNIA', 'COLORADO',
        'CONNECTICUT', 'DELAWARE', 'FLORIDA', 'GEORGIA', 'HAWAII', 'IDAHO',
        'ILLINOIS', 'INDIANA', 'IOWA', 'KANSAS', 'KENTUCKY', 'LOUISIANA',
        'MAINE', 'MARYLAND', 'MASSACHUSETTS', 'MICHIGAN', 'MINNESOTA',
        'MISSISSIPPI', 'MISSOURI', 'MONTANA', 'NEBRASKA', 'NEVADA',
        'NEW HAMPSHIRE', 'NEW JERSEY', 'NEW MEXICO', 'NEW YORK',
        'NORTH CAROLINA', 'NORTH DAKOTA', 'OHIO', 'OKLAHOMA', 'OREGON',
        'PENNSYLVANIA', 'RHODE ISLAND', 'SOUTH CAROLINA', 'SOUTH DAKOTA',
        'TENNESSEE', 'TEXAS', 'UTAH', 'VERMONT', 'VIRGINIA', 'WASHINGTON',
        'WEST VIRGINIA', 'WISCONSIN', 'WYOMING', 'COUNTRYWIDE',
    ]
    state = ""
    for s in STATES:
        if s in first_500:
            state = s.title()
            break
    
    # Detect product type
    content_lower = content[:2000].lower()
    product = ""
    if 'personal auto' in content_lower and 'rv' in content_lower:
        product = "Auto & RV"
    elif 'personal auto' in content_lower:
        product = "Auto"
    elif 'homeowner' in content_lower and 'auto' in content_lower:
        product = "Auto & Home"
    elif 'homeowner' in content_lower or 'dwelling' in content_lower:
        product = "Home"
    elif 'rv' in content_lower or 'motorhome' in content_lower:
        product = "RV"
    elif 'motorcycle' in content_lower:
        product = "Motorcycle"
    elif 'commercial' in content_lower:
        product = "Commercial"
    
    # Detect carrier from content if not provided
    carrier_label = carrier or _detect_carrier_from_content(content)
    
    # Build carrier short name
    CARRIER_SHORT = {
        'national general': 'NatGen', 'travelers': 'Travelers', 'safeco': 'Liberty Mutual',
        'liberty mutual': 'Liberty Mutual',
        'progressive': 'Progressive', 'grange': 'Grange', 'geico': 'GEICO',
        'hartford': 'Hartford', 'openly': 'Openly', 'foremost': 'Foremost',
    }
    short = carrier_label
    for full, abbr in CARRIER_SHORT.items():
        if full in carrier_label.lower():
            short = abbr
            break
    
    parts = [p for p in [short, state, product, "Guidelines"] if p]
    title = " ".join(parts) if len(parts) > 1 else filename.rsplit(".", 1)[0].replace("_", " ").title()
    return title[:120]


def _detect_carrier_from_content(content: str) -> str:
    """Try to detect carrier name from document content."""
    first_1000 = content[:1000].lower()
    CARRIERS = {
        'national general': 'National General',
        'integon': 'National General',
        'encompass': 'National General',
        'new south insurance': 'National General',
        'imperial fire': 'National General',
        'travelers': 'Travelers',
        'safeco': 'Liberty Mutual',
        'liberty mutual': 'Liberty Mutual',
        'progressive': 'Progressive',
        'grange': 'Grange',
        'geico': 'GEICO',
        'hartford': 'Hartford',
        'openly': 'Openly',
        'foremost': 'Foremost',
        'bristol west': 'Bristol West',
        'obsidian': 'Steadily',
        'steadily': 'Steadily',
    }
    for pattern, name in CARRIERS.items():
        if pattern in first_1000:
            return name
    return ""


# ── Knowledge Retrieval (used by BEACON) ─────────────────────────

def _ai_extract_intent(query: str, available_titles: list[str]) -> dict:
    """Use a fast Haiku call to understand what the user is actually asking about.
    
    Returns structured intent:
    {
        "carriers": ["national general"],
        "states": ["oklahoma"],
        "products": ["home", "homeowners"],
        "topics": ["claims", "loss history", "eligibility", "maximum claims"],
        "search_terms": ["claims limit", "loss history", "chargeable losses", "homeowners eligibility"],
        "best_titles": ["NatGen Oklahoma Auto RV & Home Guidelines"]
    }
    """
    if not ANTHROPIC_API_KEY:
        return {}
    
    import httpx
    
    titles_list = "\n".join(f"- {t}" for t in available_titles[:40])  # Cap at 40 titles
    
    prompt = f"""You are an insurance knowledge base search assistant. An insurance agent asked a question, and you need to extract their intent so we can find the right documents.

Available documents in our knowledge base:
{titles_list}

Agent's question: "{query}"

Extract the following as JSON (no markdown, no explanation, just valid JSON):
{{
  "carriers": ["full carrier names mentioned or implied, e.g. national general, travelers, safeco, grange, geico, progressive"],
  "states": ["full state names, expand abbreviations: OK=oklahoma, IL=illinois, etc."],
  "products": ["product lines: home, homeowners, auto, rv, renters, condo, umbrella, motorcycle, flood, life"],
  "topics": ["specific underwriting topics: claims, loss history, eligibility, deductible, coverage limits, binding, cancellation, non-renewal, roof age, credit score, trampoline, pool, breed restrictions, prior insurance, lapse in coverage, etc."],
  "search_terms": ["5-10 keywords and phrases to search for in documents, including insurance jargon equivalents of what the agent asked. Think about what words would actually appear in an underwriting guidelines PDF for this topic."],
  "best_titles": ["pick 1-3 document titles from the list above that are MOST likely to contain the answer, in order of relevance. If none seem relevant, return empty array."]
}}

IMPORTANT: 
- Always expand state abbreviations to full names (OK → oklahoma, IL → illinois, OH → ohio, etc.)
- "NatGen" = "National General", "Travelers" = "Travelers", etc.
- Think about what the underwriting guidelines PDF would actually say. If someone asks "how many claims can I have", the PDF probably says "loss history", "chargeable losses", "claims count", "experience period", "ineligible if X or more losses"
- For product lines, "home" and "homeowners" are interchangeable. Include both.
- Return ONLY valid JSON, nothing else."""

    try:
        haiku_model = os.getenv("BEACON_HAIKU_MODEL", "claude-haiku-4-5-20251001")
        with httpx.Client(timeout=10.0) as client:
            resp = client.post(
                "https://api.anthropic.com/v1/messages",
                headers={
                    "x-api-key": ANTHROPIC_API_KEY,
                    "anthropic-version": "2023-06-01",
                    "content-type": "application/json",
                },
                json={
                    "model": haiku_model,
                    "max_tokens": 512,
                    "messages": [{"role": "user", "content": prompt}],
                },
            )
            if resp.status_code == 200:
                data = resp.json()
                text = data.get("content", [{}])[0].get("text", "")
                # Strip markdown fences if present
                text = text.strip()
                if text.startswith("```"):
                    text = re.sub(r'^```(?:json)?\s*', '', text)
                    text = re.sub(r'\s*```$', '', text)
                intent = json.loads(text)
                logger.info(f"AI intent extraction for '{query[:60]}': carriers={intent.get('carriers')}, states={intent.get('states')}, products={intent.get('products')}, topics={intent.get('topics', [])[:3]}")
                return intent
            else:
                logger.warning(f"AI intent extraction failed: {resp.status_code}")
                return {}
    except Exception as e:
        logger.warning(f"AI intent extraction error: {e}")
        return {}


def get_relevant_knowledge(query: str, db: Session, limit: int = 5) -> str:
    """Search approved knowledge entries relevant to a query.
    Uses AI intent extraction for smart retrieval, with keyword fallback.
    Returns formatted context string to inject into BEACON's prompt."""
    
    # Get all approved entries (exclude superseded)
    entries = db.query(BeaconKnowledge).filter(
        BeaconKnowledge.status == "approved"
    ).order_by(BeaconKnowledge.created_at.desc()).all()
    
    if not entries:
        return ""
    
    # Deduplicate: if multiple entries have the same title+carrier, keep only the newest
    seen_titles = {}
    deduped = []
    for entry in entries:
        dedup_key = f"{(entry.title or '').lower()}|{(entry.carrier or '').lower()}"
        if dedup_key not in seen_titles:
            seen_titles[dedup_key] = True
            deduped.append(entry)
    entries = deduped
    
    # ── Phase 1: AI Intent Extraction ──
    # Ask Haiku to understand what the agent is really asking about
    available_titles = [e.title for e in entries if e.title]
    intent = _ai_extract_intent(query, available_titles)
    
    ai_carriers = [c.lower() for c in intent.get("carriers", [])]
    ai_states = [s.lower() for s in intent.get("states", [])]
    ai_products = [p.lower() for p in intent.get("products", [])]
    ai_topics = [t.lower() for t in intent.get("topics", [])]
    ai_search_terms = [t.lower() for t in intent.get("search_terms", [])]
    ai_best_titles = [t.lower() for t in intent.get("best_titles", [])]
    
    # ── Phase 2: Keyword fallback (always runs alongside AI) ──
    query_lower = query.lower()
    query_words = set(re.findall(r'\w+', query_lower))
    
    # State abbreviation expansion (fast fallback)
    STATE_ABBREVS = {
        'al': 'alabama', 'ak': 'alaska', 'az': 'arizona', 'ar': 'arkansas',
        'ca': 'california', 'co': 'colorado', 'ct': 'connecticut', 'de': 'delaware',
        'fl': 'florida', 'ga': 'georgia', 'hi': 'hawaii', 'id': 'idaho',
        'il': 'illinois', 'in': 'indiana', 'ia': 'iowa', 'ks': 'kansas',
        'ky': 'kentucky', 'la': 'louisiana', 'me': 'maine', 'md': 'maryland',
        'ma': 'massachusetts', 'mi': 'michigan', 'mn': 'minnesota', 'ms': 'mississippi',
        'mo': 'missouri', 'mt': 'montana', 'ne': 'nebraska', 'nv': 'nevada',
        'nh': 'new hampshire', 'nj': 'new jersey', 'nm': 'new mexico', 'ny': 'new york',
        'nc': 'north carolina', 'nd': 'north dakota', 'oh': 'ohio', 'ok': 'oklahoma',
        'or': 'oregon', 'pa': 'pennsylvania', 'ri': 'rhode island', 'sc': 'south carolina',
        'sd': 'south dakota', 'tn': 'tennessee', 'tx': 'texas', 'ut': 'utah',
        'vt': 'vermont', 'va': 'virginia', 'wa': 'washington', 'wv': 'west virginia',
        'wi': 'wisconsin', 'wy': 'wyoming',
    }
    STATE_NAMES = {v: k for k, v in STATE_ABBREVS.items()}
    
    expanded_query_words = set(query_words)
    detected_state_names = set(ai_states)  # Start with AI-detected states
    for word in query_words:
        word_l = word.lower()
        if word_l in STATE_ABBREVS:
            full_name = STATE_ABBREVS[word_l]
            expanded_query_words.add(full_name)
            for part in full_name.split():
                expanded_query_words.add(part)
            detected_state_names.add(full_name)
        elif word_l in STATE_NAMES:
            expanded_query_words.add(STATE_NAMES[word_l])
            detected_state_names.add(word_l)
    
    for state_name in STATE_NAMES:
        if state_name in query_lower:
            detected_state_names.add(state_name)
    
    # Add AI search terms to expanded words
    for term in ai_search_terms:
        for word in re.findall(r'\w+', term):
            if len(word) >= 3:
                expanded_query_words.add(word)
    
    # Merge AI products with keyword detection
    PRODUCT_KEYWORDS = {
        'home': ['home', 'homeowner', 'homeowners', 'dwelling', 'ho3', 'ho5', 'ho6'],
        'auto': ['auto', 'automobile', 'vehicle', 'car', 'driver'],
        'rv': ['rv', 'recreational vehicle', 'motorhome'],
        'renters': ['renter', 'renters', 'tenant'],
        'condo': ['condo', 'condominium'],
        'umbrella': ['umbrella', 'excess liability'],
        'motorcycle': ['motorcycle'],
        'flood': ['flood'],
        'life': ['life', 'term life', 'whole life'],
    }
    detected_products = set(ai_products)
    for product, keywords in PRODUCT_KEYWORDS.items():
        for kw in keywords:
            if kw in query_lower:
                detected_products.add(product)
                break
    
    # ── Phase 3: Score each entry using combined AI + keyword signals ──
    scored = []
    for entry in entries:
        score = 0
        title_lower = (entry.title or '').lower()
        carrier_lower = (entry.carrier or '').lower()
        searchable = f"{entry.title} {entry.content} {entry.tags or ''} {entry.carrier or ''} {entry.summary or ''}".lower()
        
        # ── AI Title Match (strongest signal) ──
        # If AI explicitly picked this doc as most relevant, massive boost
        for i, best_title in enumerate(ai_best_titles):
            if best_title == title_lower:
                score += 50 - (i * 10)  # 50 for #1 pick, 40 for #2, 30 for #3
                break
            # Fuzzy: check if AI title is contained in actual title or vice versa
            elif best_title in title_lower or title_lower in best_title:
                score += 40 - (i * 10)
                break
        
        # ── AI Carrier Match ──
        for ai_carrier in ai_carriers:
            if ai_carrier in carrier_lower or carrier_lower in ai_carrier:
                score += 20
                break
            # Handle aliases: "natgen" = "national general"
            carrier_aliases = {
                'national general': ['natgen', 'national general', 'integon', 'encompass'],
                'travelers': ['travelers', 'standard fire', 'travco', 'charter oak'],
                'safeco': ['safeco'],
                'grange': ['grange'],
                'progressive': ['progressive'],
                'geico': ['geico'],
                'hippo': ['hippo', 'spinnaker', 'spinnaker insurance'],
            }
            for canonical, aliases in carrier_aliases.items():
                if any(a in ai_carrier for a in aliases) and any(a in carrier_lower for a in aliases):
                    score += 20
                    break
        
        # ── AI State Match in Title ──
        for state_name in detected_state_names:
            if state_name in title_lower:
                score += 30
            if state_name in STATE_NAMES:
                abbrev = STATE_NAMES[state_name]
                if f" {abbrev} " in f" {title_lower} " or title_lower.endswith(f" {abbrev}"):
                    score += 30
        
        # ── AI Product Match in Title ──
        for product in detected_products:
            if product in title_lower:
                score += 15
            for kw in PRODUCT_KEYWORDS.get(product, []):
                if kw in title_lower:
                    score += 10
                    break
        
        # ── AI Search Terms in Content (deeper semantic matching) ──
        for term in ai_search_terms:
            term_words = term.split()
            if len(term_words) > 1:
                # Multi-word phrase — check if full phrase exists in content
                if term in searchable:
                    score += 8
            else:
                if len(term) >= 3 and term in searchable:
                    score += 3
        
        # ── AI Topics in Content ──
        for topic in ai_topics:
            topic_words = topic.split()
            if len(topic_words) > 1 and topic in searchable:
                score += 5
            elif len(topic) >= 3 and topic in searchable:
                score += 2
        
        # ── Original keyword matching (ensures we don't miss anything) ──
        for word in expanded_query_words:
            if len(word) >= 3:
                count = searchable.count(word)
                score += count * 1  # Reduced weight since AI signals are stronger
        
        # Carrier name from original query
        if entry.carrier and entry.carrier.lower() in query_lower:
            score += 15
        
        # Tag matches
        if entry.tags:
            for tag in entry.tags.split(","):
                tag = tag.strip().lower()
                if tag and tag in query_lower:
                    score += 10
        
        # Correction entries get a boost
        if entry.source_type == "correction":
            score += 5
        
        if score > 0:
            scored.append((score, entry))
    
    # Sort by score, take top N
    scored.sort(key=lambda x: x[0], reverse=True)
    top = scored[:limit]
    
    if not top:
        return ""
    
    # Log scoring results for debugging
    for score, entry in top[:3]:
        logger.info(f"KB match: score={score} title='{entry.title}'")
    
    # ── Phase 4: Format context with AI-enhanced chunk extraction ──
    # Combine original query words + AI search terms for better chunk extraction
    all_search_words = expanded_query_words.copy()
    for term in ai_search_terms + ai_topics:
        for word in re.findall(r'\w+', term):
            if len(word) >= 3:
                all_search_words.add(word)
    
    parts = ["\n## Agency Knowledge Base (approved entries from your team)\n"]
    for score, entry in top:
        source_label = {
            "pdf": "📄 PDF",
            "screenshot": "📸 Screenshot",
            "correction": "✏️ Correction",
            "conversation": "💬 Learned",
            "spreadsheet": "📊 Spreadsheet",
            "document": "📝 Document",
        }.get(entry.source_type, "📝 Note")
        
        parts.append(f"### {source_label}: {entry.title}")
        
        content = entry.content
        if len(content) <= 12000:
            parts.append(content)
        else:
            # Use enriched search words for chunk extraction
            chunks = _extract_relevant_chunks(content, all_search_words, max_total=15000)
            if chunks:
                parts.append(f"[Relevant sections from {len(content):,} character document]\n")
                parts.append(chunks)
            else:
                parts.append(content[:6000] + "\n[... document continues, use specific questions to find more]")
        parts.append("")
    
    return "\n".join(parts)


def _extract_relevant_chunks(content: str, query_words: set, max_total: int = 8000) -> str:
    """Extract sections of a large document that contain query keywords.
    Uses synonym expansion, phrase matching, and section awareness."""
    
    lines = content.split('\n')
    
    # Synonym expansions for insurance terms
    SYNONYMS = {
        'claim': ['claim', 'claims', 'loss', 'losses', 'chargeable', 'loss history'],
        'limit': ['limit', 'limits', 'maximum', 'max', 'exceed', 'exceeding', 'ineligible'],
        'many': ['many', 'number', 'count', 'how many', 'maximum'],
        'home': ['home', 'homeowner', 'homeowners', 'dwelling', 'residence'],
        'auto': ['auto', 'automobile', 'vehicle', 'car', 'driver'],
        'cancel': ['cancel', 'cancellation', 'non-renewal', 'nonrenewal'],
        'cover': ['cover', 'coverage', 'covered', 'covers'],
        'deduct': ['deduct', 'deductible', 'deductibles'],
        'eligible': ['eligible', 'eligibility', 'ineligible', 'qualify', 'acceptable'],
        'require': ['require', 'requirement', 'requirements', 'required'],
    }
    
    expanded = set(query_words)
    for word in query_words:
        for key, syns in SYNONYMS.items():
            if word.startswith(key) or key.startswith(word):
                expanded.update(syns)
    
    # State abbreviation expansion for chunk matching too
    STATE_ABBREVS_CHUNK = {
        'al': 'alabama', 'ak': 'alaska', 'az': 'arizona', 'ar': 'arkansas',
        'ca': 'california', 'co': 'colorado', 'ct': 'connecticut', 'de': 'delaware',
        'fl': 'florida', 'ga': 'georgia', 'hi': 'hawaii', 'id': 'idaho',
        'il': 'illinois', 'in': 'indiana', 'ia': 'iowa', 'ks': 'kansas',
        'ky': 'kentucky', 'la': 'louisiana', 'me': 'maine', 'md': 'maryland',
        'ma': 'massachusetts', 'mi': 'michigan', 'mn': 'minnesota', 'ms': 'mississippi',
        'mo': 'missouri', 'mt': 'montana', 'ne': 'nebraska', 'nv': 'nevada',
        'nh': 'new hampshire', 'nj': 'new jersey', 'nm': 'new mexico', 'ny': 'new york',
        'nc': 'north carolina', 'nd': 'north dakota', 'oh': 'ohio', 'ok': 'oklahoma',
        'or': 'oregon', 'pa': 'pennsylvania', 'ri': 'rhode island', 'sc': 'south carolina',
        'sd': 'south dakota', 'tn': 'tennessee', 'tx': 'texas', 'ut': 'utah',
        'vt': 'vermont', 'va': 'virginia', 'wa': 'washington', 'wv': 'west virginia',
        'wi': 'wisconsin', 'wy': 'wyoming',
    }
    for word in query_words:
        if word.lower() in STATE_ABBREVS_CHUNK:
            full = STATE_ABBREVS_CHUNK[word.lower()]
            expanded.add(full)
            for part in full.split():
                expanded.add(part)
    
    scored_lines = []
    for i, line in enumerate(lines):
        line_lower = line.lower()
        if not line_lower.strip():
            continue
        score = 0
        matched = set()
        for word in expanded:
            if len(word) >= 3 and word in line_lower:
                score += 2
                matched.add(word)
        # Section header bonus
        stripped = line.strip()
        if stripped and (stripped.isupper() or stripped.startswith(('I.', 'II.', 'III.', 'IV.', 'A.', 'B.', 'C.', 'D.'))):
            if matched:
                score += 5
        # Multi-word proximity bonus
        if len(matched) >= 3:
            score += 10
        elif len(matched) >= 2:
            score += 5
        # Key phrase bonus
        for phrase in ['ineligible', 'loss history', 'claim count', 'exceeding', 'chargeable loss', 'experience period', 'not eligible']:
            if phrase in line_lower:
                score += 8
        if score > 0:
            scored_lines.append((i, score))
    
    if not scored_lines:
        return ""
    
    scored_lines.sort(key=lambda x: x[1], reverse=True)
    
    CONTEXT_LINES = 30
    chunks = []
    used_lines = set()
    total_chars = 0
    
    for line_idx, score in scored_lines[:25]:
        start = max(0, line_idx - CONTEXT_LINES)
        end = min(len(lines), line_idx + CONTEXT_LINES + 1)
        overlap = sum(1 for i in range(start, end) if i in used_lines)
        if overlap > (end - start) * 0.5:
            continue
        chunk_text = '\n'.join(lines[start:end])
        if total_chars + len(chunk_text) > max_total:
            break
        chunks.append(f"[...section near line {start+1}...]\n{chunk_text}")
        used_lines.update(range(start, end))
        total_chars += len(chunk_text)
    
    return "\n\n".join(chunks)



# ── API Endpoints ────────────────────────────────────────────────

@router.post("/upload")
async def upload_knowledge(
    file: UploadFile = File(...),
    title: str = Form(""),
    carrier: str = Form(""),
    tags: str = Form(""),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Upload a PDF or screenshot to the knowledge base."""
    file_bytes = await file.read()
    file_hash = hashlib.sha256(file_bytes).hexdigest()
    
    # Check for exact duplicate (same file hash)
    existing = db.query(BeaconKnowledge).filter(BeaconKnowledge.file_hash == file_hash).first()
    if existing:
        return {"error": "This exact file has already been uploaded", "existing_id": existing.id}
    
    # Check for superseding — same title+carrier means this is an updated version
    # Mark older entries as superseded so only the newest is used
    if title and carrier:
        older = db.query(BeaconKnowledge).filter(
            BeaconKnowledge.title == title,
            BeaconKnowledge.carrier == carrier,
            BeaconKnowledge.status.in_(["approved", "pending"]),
        ).all()
        for old in older:
            old.status = "superseded"
            logger.info(f"Superseded older knowledge entry: {old.title} (id={old.id})")
        if older:
            db.commit()
    
    filename = file.filename or "upload"
    content_type = file.content_type or ""
    
    # Determine source type and extract text
    if "pdf" in content_type or filename.lower().endswith(".pdf"):
        source_type = "pdf"
        content = _extract_text_from_pdf(file_bytes)
        if not content:
            return {"error": "Could not extract text from PDF. The file may be scanned or empty."}
    elif any(ext in filename.lower() for ext in [".png", ".jpg", ".jpeg", ".gif", ".webp"]) or "image" in content_type:
        source_type = "screenshot"
        media = content_type if content_type else "image/png"
        content = _extract_text_from_image(file_bytes, media)
        if not content or content.startswith("["):
            return {"error": "Could not extract information from image."}
    elif filename.lower().endswith((".xlsx", ".xls")) or "spreadsheet" in content_type or "excel" in content_type:
        source_type = "spreadsheet"
        content = _extract_text_from_excel(file_bytes, filename)
        if not content:
            return {"error": "Could not extract data from Excel file."}
    elif filename.lower().endswith(".csv") or "csv" in content_type:
        source_type = "spreadsheet"
        content = _extract_text_from_csv(file_bytes)
        if not content:
            return {"error": "Could not extract data from CSV file."}
    elif filename.lower().endswith((".txt", ".md", ".text")):
        source_type = "document"
        content = file_bytes.decode("utf-8", errors="replace")
        if not content.strip():
            return {"error": "Text file is empty."}
    else:
        return {"error": f"Unsupported file type: {content_type or filename}. Supported: PDF, images (PNG/JPG), Excel (XLSX/XLS), CSV, text files."}
    
    # Auto-generate title if not provided or if filename is generic
    is_generic_filename = (
        not title or 
        title.lower().startswith("download") or 
        title.lower().startswith("untitled") or
        re.match(r'^[\d\s\(\)]+$', title)
    )
    if is_generic_filename:
        title = _generate_smart_title(content, carrier, filename)
    elif not title:
        title = filename.rsplit(".", 1)[0].replace("_", " ").replace("-", " ").title()
    
    # Auto-detect carrier from content if not provided
    if not carrier:
        carrier = _detect_carrier_from_content(content)
    
    # Generate summary
    summary = _generate_summary(content, title)
    
    # Auto-approve if manager, otherwise pending
    status = "approved" if _is_manager(current_user) else "pending"
    
    entry = BeaconKnowledge(
        source_type=source_type,
        title=title,
        content=content,
        summary=summary,
        tags=tags,
        carrier=carrier,
        status=status,
        submitted_by=current_user.id,
        submitted_by_name=current_user.full_name or current_user.username,
        original_filename=filename,
        file_hash=file_hash,
        reviewed_by=current_user.id if status == "approved" else None,
        reviewed_by_name=(current_user.full_name or current_user.username) if status == "approved" else None,
        reviewed_at=datetime.utcnow() if status == "approved" else None,
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    
    logger.info(f"Knowledge uploaded: {title} ({source_type}) by {current_user.username} → {status}")
    return _serialize(entry)


@router.post("/correction")
def add_correction(
    data: dict,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Add a correction from chat (BEACON was wrong about something)."""
    content = data.get("content", "").strip()
    title = data.get("title", "").strip()
    carrier = data.get("carrier", "")
    tags = data.get("tags", "")
    
    if not content:
        raise HTTPException(400, "Content is required")
    if not title:
        title = f"Correction: {content[:60]}..."
    
    status = "approved" if _is_manager(current_user) else "pending"
    
    entry = BeaconKnowledge(
        source_type="correction",
        title=title,
        content=content,
        summary=content[:200],
        tags=tags,
        carrier=carrier,
        status=status,
        submitted_by=current_user.id,
        submitted_by_name=current_user.full_name or current_user.username,
        reviewed_by=current_user.id if status == "approved" else None,
        reviewed_by_name=(current_user.full_name or current_user.username) if status == "approved" else None,
        reviewed_at=datetime.utcnow() if status == "approved" else None,
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    
    logger.info(f"Correction added: {title} by {current_user.username} → {status}")
    return _serialize(entry)


@router.post("/from-conversation")
def add_from_conversation(
    data: dict,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Save a useful BEACON conversation exchange as knowledge."""
    content = data.get("content", "").strip()
    title = data.get("title", "").strip()
    carrier = data.get("carrier", "")
    tags = data.get("tags", "")
    
    if not content:
        raise HTTPException(400, "Content is required")
    if not title:
        title = f"Learned: {content[:60]}..."
    
    summary = _generate_summary(content, title)
    status = "approved" if _is_manager(current_user) else "pending"
    
    entry = BeaconKnowledge(
        source_type="conversation",
        title=title,
        content=content,
        summary=summary,
        tags=tags,
        carrier=carrier,
        status=status,
        submitted_by=current_user.id,
        submitted_by_name=current_user.full_name or current_user.username,
        reviewed_by=current_user.id if status == "approved" else None,
        reviewed_by_name=(current_user.full_name or current_user.username) if status == "approved" else None,
        reviewed_at=datetime.utcnow() if status == "approved" else None,
    )
    db.add(entry)
    db.commit()
    db.refresh(entry)
    
    logger.info(f"Conversation knowledge saved: {title} by {current_user.username}")
    return _serialize(entry)


@router.get("/entries")
def list_entries(
    status: str = Query(None),
    source_type: str = Query(None),
    search: str = Query(None),
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """List knowledge base entries with optional filters."""
    q = db.query(BeaconKnowledge)
    
    if status:
        q = q.filter(BeaconKnowledge.status == status)
    if source_type:
        q = q.filter(BeaconKnowledge.source_type == source_type)
    if search:
        pattern = f"%{search}%"
        q = q.filter(
            (BeaconKnowledge.title.ilike(pattern)) |
            (BeaconKnowledge.content.ilike(pattern)) |
            (BeaconKnowledge.tags.ilike(pattern)) |
            (BeaconKnowledge.carrier.ilike(pattern))
        )
    
    # Non-managers only see their own pending + all approved
    if not _is_manager(current_user):
        q = q.filter(
            (BeaconKnowledge.status == "approved") |
            (BeaconKnowledge.submitted_by == current_user.id)
        )
    # Managers see everything including superseded
    
    entries = q.order_by(BeaconKnowledge.created_at.desc()).limit(100).all()
    
    # Count pending for badge
    pending_count = db.query(BeaconKnowledge).filter(
        BeaconKnowledge.status == "pending"
    ).count()
    
    return {
        "entries": [_serialize(e) for e in entries],
        "pending_count": pending_count,
    }


@router.get("/entries/{entry_id}")
def get_entry(
    entry_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get full details of a knowledge entry."""
    entry = db.query(BeaconKnowledge).filter(BeaconKnowledge.id == entry_id).first()
    if not entry:
        raise HTTPException(404, "Entry not found")
    return _serialize(entry)


@router.post("/entries/{entry_id}/approve")
def approve_entry(
    entry_id: int,
    data: dict = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Approve a pending knowledge entry (managers only)."""
    if not _is_manager(current_user):
        raise HTTPException(403, "Only managers can approve knowledge entries")
    
    entry = db.query(BeaconKnowledge).filter(BeaconKnowledge.id == entry_id).first()
    if not entry:
        raise HTTPException(404, "Entry not found")
    
    entry.status = "approved"
    entry.reviewed_by = current_user.id
    entry.reviewed_by_name = current_user.full_name or current_user.username
    entry.reviewed_at = datetime.utcnow()
    entry.review_note = (data or {}).get("note", "")
    db.commit()
    
    logger.info(f"Knowledge approved: {entry.title} by {current_user.username}")
    return _serialize(entry)


@router.post("/entries/{entry_id}/reject")
def reject_entry(
    entry_id: int,
    data: dict = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Reject a pending knowledge entry (managers only)."""
    if not _is_manager(current_user):
        raise HTTPException(403, "Only managers can reject knowledge entries")
    
    entry = db.query(BeaconKnowledge).filter(BeaconKnowledge.id == entry_id).first()
    if not entry:
        raise HTTPException(404, "Entry not found")
    
    entry.status = "rejected"
    entry.reviewed_by = current_user.id
    entry.reviewed_by_name = current_user.full_name or current_user.username
    entry.reviewed_at = datetime.utcnow()
    entry.review_note = (data or {}).get("note", "")
    db.commit()
    
    logger.info(f"Knowledge rejected: {entry.title} by {current_user.username}")
    return _serialize(entry)


@router.put("/entries/{entry_id}")
def update_entry(
    entry_id: int,
    data: dict,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Edit a knowledge entry (managers, or original submitter if still pending)."""
    entry = db.query(BeaconKnowledge).filter(BeaconKnowledge.id == entry_id).first()
    if not entry:
        raise HTTPException(404, "Entry not found")
    
    if not _is_manager(current_user) and entry.submitted_by != current_user.id:
        raise HTTPException(403, "Not authorized to edit this entry")
    
    for field in ("title", "content", "tags", "carrier"):
        if field in data:
            setattr(entry, field, data[field])
    
    db.commit()
    return _serialize(entry)


@router.delete("/entries/{entry_id}")
def delete_entry(
    entry_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Delete a knowledge entry (managers only)."""
    if not _is_manager(current_user):
        raise HTTPException(403, "Only managers can delete knowledge entries")
    
    entry = db.query(BeaconKnowledge).filter(BeaconKnowledge.id == entry_id).first()
    if not entry:
        raise HTTPException(404, "Entry not found")
    
    db.delete(entry)
    db.commit()
    return {"deleted": True}


@router.get("/stats")
def kb_stats(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Get knowledge base statistics."""
    total = db.query(BeaconKnowledge).filter(BeaconKnowledge.status != "superseded").count()
    approved = db.query(BeaconKnowledge).filter(BeaconKnowledge.status == "approved").count()
    pending = db.query(BeaconKnowledge).filter(BeaconKnowledge.status == "pending").count()
    superseded = db.query(BeaconKnowledge).filter(BeaconKnowledge.status == "superseded").count()
    
    by_type = {}
    for st in ("pdf", "screenshot", "correction", "conversation", "spreadsheet", "document"):
        by_type[st] = db.query(BeaconKnowledge).filter(
            BeaconKnowledge.source_type == st,
            BeaconKnowledge.status == "approved",
        ).count()
    
    return {
        "total": total,
        "approved": approved,
        "pending": pending,
        "superseded": superseded,
        "by_type": by_type,
    }


@router.post("/re-extract/{entry_id}")
async def re_extract_entry(
    entry_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Re-extract text from a knowledge base entry's original file.
    Useful for entries that were uploaded before OCR was available."""
    if current_user.role not in ("admin", "ADMIN"):
        raise HTTPException(status_code=403, detail="Admin only")
    
    entry = db.query(BeaconKnowledge).filter(BeaconKnowledge.id == entry_id).first()
    if not entry:
        raise HTTPException(status_code=404, detail="Entry not found")
    
    if not entry.file_hash:
        return {"error": "No file hash stored — can't re-extract without original file"}
    
    old_len = len(entry.content or "")
    return {
        "id": entry.id,
        "title": entry.title,
        "old_content_length": old_len,
        "message": "To re-extract, please re-upload the original PDF. The new OCR pipeline will now handle scanned documents automatically.",
    }


@router.post("/bulk-health-check")
async def bulk_health_check(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Check all KB entries for suspiciously low content (likely failed extractions).
    Returns list of entries that need re-upload."""
    if current_user.role not in ("admin", "ADMIN"):
        raise HTTPException(status_code=403, detail="Admin only")
    
    entries = db.query(BeaconKnowledge).filter(
        BeaconKnowledge.status.in_(["approved", "pending"])
    ).all()
    
    issues = []
    healthy = []
    for entry in entries:
        content_len = len(entry.content or "")
        if content_len < 1000 and entry.source_type == "pdf":
            issues.append({
                "id": entry.id,
                "title": entry.title,
                "carrier": entry.carrier,
                "content_length": content_len,
                "issue": "Likely scanned PDF — needs re-upload with new OCR pipeline",
            })
        else:
            healthy.append({
                "id": entry.id,
                "title": entry.title,
                "content_length": content_len,
            })
    
    return {
        "total_entries": len(entries),
        "healthy_count": len(healthy),
        "issue_count": len(issues),
        "issues": issues,
        "message": f"{len(issues)} entries have very little content and need re-uploading. "
                   f"The new OCR pipeline will automatically detect scanned PDFs and use "
                   f"Claude Vision to extract text from page images." if issues else "All entries look healthy!",
    }
